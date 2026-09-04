# =========================
# IMPORTAÇÕES
# =========================
from fastapi import FastAPI, HTTPException, Depends, Request, Response, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel, EmailStr
from sqlalchemy import create_engine, text
from sqlalchemy.exc import IntegrityError
from passlib.context import CryptContext
from datetime import datetime, timedelta, date, timezone

from funil import STATUS_GANHO, agregar_funil, janela_meses
from typing import Optional
from apscheduler.schedulers.background import BackgroundScheduler
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import uuid
import jwt
import functools
import difflib
import os
import re
import time
import hmac
import hashlib
import secrets as _secrets
import threading
import unicodedata
import math
import io
import httpx
import resend
import base64
import json
import asyncio
import pyotp
import requests as http_requests
from urllib.parse import urlparse


def _url_sem_credencial(url: str | None) -> str:
    """Host e banco da URL de conexão, sem usuário nem senha.

    O boot loga a URL para dar uma pista rápida de "apontei pro banco certo?",
    mas o log do serviço fica visível para todo mundo com acesso ao projeto no
    Railway — e a senha do Postgres ia inteira para lá a cada deploy.
    """
    if not url:
        return "(não definida)"
    try:
        partes = urlparse(url)
        porta = f":{partes.port}" if partes.port else ""
        return f"{partes.scheme}://{partes.hostname or '?'}{porta}{partes.path}"
    except ValueError:
        return "(formato inválido)"


print("🔥 ENV DATABASE_URL:", _url_sem_credencial(os.getenv("DATABASE_URL")))

# =========================
# CONFIG
# =========================
DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise ValueError("🚨 DATABASE_URL não encontrada!")

# Chave de assinatura do JWT — NUNCA hardcoded. Deve vir do ambiente (Railway Variables).
SECRET_KEY = os.getenv("JWT_SECRET") or os.getenv("SECRET_KEY")
if not SECRET_KEY:
    SECRET_KEY = _secrets.token_urlsafe(64)
    print(
        "⚠️  JWT_SECRET não definido no ambiente — usando chave aleatória temporária "
        "(todos os tokens invalidam a cada restart). Defina JWT_SECRET no Railway."
    )
ALGORITHM = "HS256"
# Access token curto + refresh token de vida longa (cookie httpOnly).
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "30"))
REFRESH_TOKEN_EXPIRE_DAYS = int(os.getenv("REFRESH_TOKEN_EXPIRE_DAYS", "7"))

# Origem(ns) do frontend autorizadas (CORS). Configurável por env (lista separada por vírgula).
FRONTEND_ORIGINS = [
    o.strip()
    for o in os.getenv(
        "FRONTEND_ORIGINS",
        "https://frontend-crm-xi-plum.vercel.app,http://localhost:3000",
    ).split(",")
    if o.strip()
]

resend.api_key = os.getenv("RESEND_API_KEY")

MICROSOFT_CLIENT_ID = os.getenv("MICROSOFT_CLIENT_ID")
MICROSOFT_CLIENT_SECRET = os.getenv("MICROSOFT_CLIENT_SECRET")
MICROSOFT_TENANT_ID = os.getenv("MICROSOFT_TENANT_ID")
MICROSOFT_REDIRECT_URI = os.getenv("MICROSOFT_REDIRECT_URI")

GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET")
GOOGLE_REDIRECT_URI = os.getenv("GOOGLE_REDIRECT_URI")

BACKEND_URL = os.getenv("BACKEND_URL", "https://backend-crm-production-157b.up.railway.app")
OUTLOOK_WEBHOOK_SECRET = os.getenv("OUTLOOK_WEBHOOK_SECRET", "crm-webhook-secret")
GMAIL_PUBSUB_TOPIC = os.getenv("GMAIL_PUBSUB_TOPIC", "projects/SEU_PROJECT_ID/topics/gmail-crm-push")
GOOGLE_PLACES_API_KEY = os.getenv("GOOGLE_PLACES_API_KEY")

# --- Teto de chamadas PAGAS ao Google Places -------------------------------
# O cache absorve a repeticao de segmento/cidade, mas quem paga a conta e o
# cache miss: cada um vira uma chamada de verdade ao Text Search. Ate aqui o
# unico limite era o do proprio Google, ou seja, so descobriamos o estouro
# depois de gastar. Este contador barra ANTES da chamada, por USUARIO e por MES:
# o balde e individual, entao um vendedor sozinho nao consome a folga dos outros
# -- e o gasto maximo da conta e este limite vezes o numero de vendedores.
# Acerto de cache nao consome nada: so o cache miss vira chamada paga.
#   PLACES_LIMITE_MENSAL=0 -> desliga a busca paga (a tela so responde do cache).
PLACES_LIMITE_MENSAL = int(os.getenv("PLACES_LIMITE_MENSAL", "50"))
# O mes da cota vira a meia-noite do dia 1 deste fuso, em offset fixo de horas.
# -3 e o horario de Brasilia; o Brasil nao tem horario de verao desde 2019,
# entao o offset fixo evita depender do tzdata dentro do container.
PLACES_FUSO_HORAS = float(os.getenv("PLACES_FUSO_HORAS", "-3"))
# Quando quem recusa e o Google, nao da para saber o reset -- pode ser limite
# por minuto ou o teto da conta no Google Cloud. Seguramos a tela essa janela.
PLACES_ESPERA_GOOGLE_MIN = int(os.getenv("PLACES_ESPERA_GOOGLE_MIN", "15"))

# =========================================================================
# GUARDA DE AMBIENTE: nao subir contra producao a partir da maquina de alguem
# =========================================================================
# O Postgres de producao e alcancavel de qualquer lugar pelo proxy TCP do
# Railway, e as migracoes deste arquivo sao LAZY: disparam sozinhas na primeira
# chamada autenticada, sem passo manual e sem confirmacao. Basta subir o app
# local com a DATABASE_URL errada para um ALTER TABLE cair nos dados reais.
#
# A checagem e por HOST, nao por nome de arquivo de env: renomear o `.env`
# ajuda, mas nao impede ninguem de exportar a variavel na mao. Dentro do Railway
# a plataforma injeta RAILWAY_ENVIRONMENT; fora dela a variavel nao existe, e e
# esse o sinal de "isto aqui e a maquina de alguem".
HOSTS_PRODUCAO = tuple(
    h.strip().lower()
    for h in os.getenv(
        "HOSTS_PRODUCAO", "shortline.proxy.rlwy.net,.railway.internal"
    ).split(",")
    if h.strip()
)

_LINHA = "=" * 72

_TEXTO_RECUSA = """
  RECUSANDO SUBIR: a DATABASE_URL aponta para o banco de PRODUCAO
  (%s)
  e este processo nao esta rodando no Railway.

  As migracoes deste arquivo sao lazy: subir assim escreve nos dados
  reais na primeira chamada autenticada.

  Para desenvolver, aponte para o banco de dev:
      uvicorn main:app --reload --env-file .env.dev

  Se voce REALMENTE quer producao a partir daqui, seja explicito:
      PERMITIR_PROD_LOCAL=1 uvicorn main:app --env-file .env.production
"""

_TEXTO_PERMITIDO = """
  ATENCAO: conectado ao Postgres de PRODUCAO fora do Railway.
  PERMITIR_PROD_LOCAL=1 esta setado -- as migracoes lazy vao rodar
  contra os dados reais na primeira chamada autenticada.
"""


def _e_host_de_producao(url):
    """True quando a URL de conexao aponta para um host de producao conhecido."""
    if not url:
        return False
    try:
        host = (urlparse(url).hostname or "").lower()
    except ValueError:
        return False
    return any(host == h or host.endswith(h) for h in HOSTS_PRODUCAO)


_FORA_DO_RAILWAY = not (
    os.getenv("RAILWAY_ENVIRONMENT") or os.getenv("RAILWAY_SERVICE_ID")
)

if _e_host_de_producao(DATABASE_URL) and _FORA_DO_RAILWAY:
    if os.getenv("PERMITIR_PROD_LOCAL") == "1":
        print(_LINHA + _TEXTO_PERMITIDO + _LINHA)
    else:
        raise SystemExit(
            _LINHA + (_TEXTO_RECUSA % _url_sem_credencial(DATABASE_URL)) + _LINHA
        )


engine = create_engine(DATABASE_URL)

# =========================================================================
# MIGRACOES: uma vez por processo, nao por requisicao
# =========================================================================
# As funcoes garantir_* sao idempotentes, mas nao sao de graca. So a
# garantir_campos_pipeline dispara 6 ALTER TABLE -- que pegam ACCESS EXCLUSIVE
# em `empresas` mesmo quando nao tem nada a fazer -- mais um UPDATE de backfill.
# Chamadas no inicio de cada rota, isso significava escrita e disputa de lock em
# TODA leitura, e o front agora rele /empresas e /orcamentos de 5 em 5 segundos
# por aba aberta.
_MIGRACOES_FEITAS: set[str] = set()
_TRAVA_MIGRACAO = threading.Lock()


def uma_vez(fn):
    """Executa a migracao so na primeira chamada bem-sucedida deste processo.

    So marca como feita quando a funcao volta sem erro: se falhar, a proxima
    requisicao tenta de novo, exatamente como era antes.
    """

    @functools.wraps(fn)
    def wrapper(conn, *args, **kwargs):
        if fn.__name__ in _MIGRACOES_FEITAS:
            return None
        # A trava importa: rota sincrona no FastAPI roda em threadpool, entao
        # duas requisicoes simultaneas entrariam aqui juntas.
        with _TRAVA_MIGRACAO:
            if fn.__name__ in _MIGRACOES_FEITAS:
                return None
            resultado = fn(conn, *args, **kwargs)
            _MIGRACOES_FEITAS.add(fn.__name__)
            return resultado

    return wrapper

security = HTTPBearer()

# Rate limiting em nível de aplicação (in-memory; para múltiplas instâncias trocar
# o storage por Redis via storage_uri=os.getenv("REDIS_URL")).
limiter = Limiter(key_func=get_remote_address, default_limits=[])

app = FastAPI()
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

app.add_middleware(
    CORSMiddleware,
    allow_origins=FRONTEND_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
)

SEGMENTOS_PADRAO = [
    "Academias e Fitness",
    "Administracao de Condominios",
    "Advocacia",
    "Agencia de Marketing",
    "Agencia de Publicidade",
    "Agronegocio",
    "Alimentos e Bebidas",
    "Arquitetura e Urbanismo",
    "Assistencia Tecnica",
    "Atacado e Distribuicao",
    "Automacao Industrial",
    "Autopecas",
    "Bares e Restaurantes",
    "Beleza e Estetica",
    "Biotecnologia",
    "Clinicas Medicas",
    "Comercio Exterior",
    "Comercio Varejista",
    "Concessionarias",
    "Construcao Civil",
    "Consultoria Empresarial",
    "Contabilidade",
    "Coworking",
    "Cursos e Treinamentos",
    "Decoracao",
    "Distribuidora",
    "E-commerce",
    "Educacao",
    "Energia",
    "Energia Solar",
    "Engenharia",
    "Entretenimento",
    "Escritorio de Projetos",
    "Eventos",
    "Farmacias e Drogarias",
    "Financeiro",
    "Franquias",
    "Gestao de Pessoas",
    "Hotelaria",
    "Imobiliarias",
    "Industria Alimenticia",
    "Industria Automotiva",
    "Industria Farmaceutica",
    "Industria Metalurgica",
    "Industria Textil",
    "Logistica e Transporte",
    "Manutencao Predial",
    "Maquinas e Equipamentos",
    "Materiais de Construcao",
    "Moda e Vestuario",
    "Moveis Planejados",
    "Odontologia",
    "Pet Shop",
    "Produtos Agropecuarios",
    "Recursos Humanos",
    "Saude",
    "Seguranca Eletronica",
    "Seguros",
    "Servicos de Limpeza",
    "Servicos Financeiros",
    "Software e SaaS",
    "Supermercados",
    "Tecnologia da Informacao",
    "Telecomunicacoes",
    "Turismo",
    "Venda de Gado",
    "Vendas B2B",
    "Veterinaria",
    "Agropecuaria",
    "Clinicas Odontologicas",
    "Confeitaria",
    "Delivery",
    "Grafica",
    "Hospitais",
    "Jardinagem e Paisagismo",
    "Laboratorios",
    "Laticinios",
    "Lavanderias",
    "Marcenaria",
    "Padarias",
    "Papelarias",
    "Postos de Combustivel",
    "Serralheria",
    "Transportadoras",
]

PALAVRAS_CHAVE_SEGMENTO = {
    "academia","administracao","advocacia","agencia","agro","agronegocio","alimento","arquitetura","assistencia","atacado","automacao",
    "autopecas","bar","beleza","biotecnologia","clinica","comercio","condominio","concessionaria","construcao","consultoria","contabilidade","coworking",
    "curso","decoracao","distribuicao","distribuidora","ecommerce","educacao","energia","engenharia","entretenimento","escola",
    "evento","farmacia","financeiro","fitness","franquia","gado","gestao","hotel","imobiliaria","industria","limpeza","logistica",
    "manutencao","maquinas","marketing","materiais","medica","metalurgica","moda","moveis","odontologia","oficina","papelaria","pet",
    "projetos","publicidade","recursos","restaurante","rh","saas","saude",
    "seguranca",
    "seguros",
    "servicos",
    "software",
    "solar",
    "supermercado",
    "tecnologia",
    "telecomunicacoes",
    "textil",
    "transporte",
    "turismo",
    "varejo",
    "vendas",
    "veterinaria",
    "agropecuaria",
    "combustivel",
    "confeitaria",
    "contabil",
    "delivery",
    "frigorifico",
    "grafica",
    "hospital",
    "jardinagem",
    "juridico",
    "laboratorio",
    "laticinios",
    "lavanderia",
    "marcenaria",
    "padaria",
    "paisagismo",
    "panificadora",
    "pecuaria",
    "posto",
    "rural",
    "serralheria",
    "transportadora",
}

# =========================
# SEGURANÇA
# =========================
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


def hash_senha(senha: str) -> str:
    return pwd_context.hash(senha[:72])


def verificar_senha(senha: str, senha_hash: str) -> bool:
    return pwd_context.verify(senha[:72], senha_hash)


def criar_token_acesso(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
        if not email:
            raise HTTPException(401, "Token inválido")
        return email
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Token inválido")


# =========================================================================
# AUTH HARDENING — lockout, auditoria, refresh tokens, MFA
# =========================================================================

# Hash "dummy" para comparar senha mesmo quando o usuário não existe.
# Evita timing attack que revela se um e-mail está cadastrado (enumeração).
_DUMMY_HASH = pwd_context.hash("dummy-password-para-timing-constante")

# Mensagem genérica única para login — não revela se é e-mail ou senha o erro.
CREDENCIAIS_INVALIDAS = "E-mail ou senha inválidos."

# --- Bloqueio progressivo por conta (in-memory; single instance) ---
_login_lock = threading.Lock()
_login_attempts: dict[str, dict] = {}
MAX_TENTATIVAS = 5
JANELA_LOCKOUT_SEG = 15 * 60


def _now() -> float:
    return time.time()


def checar_lockout(email: str):
    """Bloqueia por 15 min após MAX_TENTATIVAS falhas na janela. Levanta 429."""
    email = (email or "").lower()
    with _login_lock:
        rec = _login_attempts.get(email)
        if not rec:
            return
        if rec["count"] >= MAX_TENTATIVAS and _now() < rec["until"]:
            restante = int((rec["until"] - _now()) / 60) + 1
            raise HTTPException(
                429,
                f"Muitas tentativas de login. Tente novamente em {restante} min.",
            )


def registrar_falha_login(email: str) -> int:
    email = (email or "").lower()
    with _login_lock:
        rec = _login_attempts.get(email)
        if not rec or _now() >= rec.get("until", 0):
            rec = {"count": 0, "until": 0.0}
        rec["count"] += 1
        if rec["count"] >= MAX_TENTATIVAS:
            rec["until"] = _now() + JANELA_LOCKOUT_SEG
        else:
            rec["until"] = _now() + JANELA_LOCKOUT_SEG
        _login_attempts[email] = rec
        return rec["count"]


def limpar_falhas_login(email: str):
    with _login_lock:
        _login_attempts.pop((email or "").lower(), None)


def client_ip(request: Optional[Request]) -> Optional[str]:
    """IP real do cliente considerando proxies (Cloudflare/Railway)."""
    if not request:
        return None
    xff = request.headers.get("x-forwarded-for")
    if xff:
        return xff.split(",")[0].strip()
    return request.client.host if request.client else None


# --- Auditoria ---
def registrar_auditoria(
    *,
    usuario: Optional[dict] = None,
    email: Optional[str] = None,
    acao: str,
    recurso: Optional[str] = None,
    recurso_id: Optional[str] = None,
    quantidade: Optional[int] = None,
    request: Optional[Request] = None,
    meta: Optional[dict] = None,
    conn=None,
):
    """Grava um evento no audit_log. Nunca inclui senha/segredos."""
    def _do(c):
        c.execute(
            text(
                """
                INSERT INTO audit_log
                    (usuario_id, usuario_email, conta_id, acao, recurso, recurso_id,
                     quantidade, ip, user_agent, meta)
                VALUES
                    (:uid, :email, :cid, :acao, :recurso, :rid,
                     :qtd, :ip, :ua, CAST(:meta AS JSONB))
                """
            ),
            {
                "uid": (usuario or {}).get("usuario_id"),
                "email": (usuario or {}).get("email") or email,
                "cid": (usuario or {}).get("conta_id"),
                "acao": acao,
                "recurso": recurso,
                "rid": str(recurso_id) if recurso_id is not None else None,
                "qtd": quantidade,
                "ip": client_ip(request),
                "ua": request.headers.get("user-agent") if request else None,
                "meta": json.dumps(meta or {}),
            },
        )
    try:
        if conn is not None:
            _do(conn)
        else:
            with engine.begin() as c:
                _do(c)
    except Exception as e:
        # Auditoria nunca deve derrubar a requisição principal.
        print(f"⚠️ Falha ao gravar auditoria ({acao}): {e}")


# --- Refresh tokens (opacos, armazenados com hash) ---
def _hash_token(raw: str) -> str:
    return hashlib.sha256(raw.encode()).hexdigest()


def criar_refresh_token(conn, usuario_id: str, familia: Optional[str] = None,
                        request: Optional[Request] = None) -> str:
    raw = _secrets.token_urlsafe(48)
    familia = familia or str(uuid.uuid4())
    conn.execute(
        text(
            """
            INSERT INTO refresh_tokens
                (usuario_id, token_hash, familia, user_agent, ip, expira_em)
            VALUES (:uid, :h, :fam, :ua, :ip,
                    NOW() + (:dias || ' days')::interval)
            """
        ),
        {
            "uid": usuario_id,
            "h": _hash_token(raw),
            "fam": familia,
            "ua": request.headers.get("user-agent") if request else None,
            "ip": client_ip(request),
            "dias": str(REFRESH_TOKEN_EXPIRE_DAYS),
        },
    )
    return f"{familia}.{raw}"


def revogar_refresh_familia(conn, familia: str):
    conn.execute(
        text("UPDATE refresh_tokens SET revogado = TRUE WHERE familia = :f"),
        {"f": familia},
    )


def revogar_refresh_usuario(conn, usuario_id: str):
    conn.execute(
        text("UPDATE refresh_tokens SET revogado = TRUE WHERE usuario_id = :u"),
        {"u": usuario_id},
    )


# --- MFA / TOTP ---
def mfa_gerar_backup_codes(n: int = 10):
    """Gera N códigos de uso único; retorna (lista_plana, lista_hash)."""
    codes = [f"{_secrets.randbelow(10**8):08d}" for _ in range(n)]
    hashes = [_hash_token(c) for c in codes]
    return codes, hashes


def mfa_verificar_totp(secret: str, code: str) -> bool:
    if not secret or not code:
        return False
    return pyotp.TOTP(secret).verify(str(code).strip(), valid_window=1)


# --- Migration de segurança (lazy, roda junto do schema multiusuário) ---
@uma_vez
def garantir_seguranca(conn):
    conn.execute(
        text(
            """
        CREATE TABLE IF NOT EXISTS audit_log (
            id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            usuario_id UUID,
            usuario_email TEXT,
            conta_id UUID,
            acao TEXT NOT NULL,
            recurso TEXT,
            recurso_id TEXT,
            quantidade INTEGER,
            ip TEXT,
            user_agent TEXT,
            meta JSONB DEFAULT '{}',
            criado_em TIMESTAMP DEFAULT NOW()
        )
        """
        )
    )
    conn.execute(
        text("CREATE INDEX IF NOT EXISTS idx_audit_conta_data ON audit_log (conta_id, criado_em DESC)")
    )
    conn.execute(
        text(
            """
        CREATE TABLE IF NOT EXISTS refresh_tokens (
            id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            usuario_id UUID NOT NULL,
            token_hash TEXT NOT NULL UNIQUE,
            familia UUID NOT NULL,
            user_agent TEXT,
            ip TEXT,
            revogado BOOLEAN DEFAULT FALSE,
            expira_em TIMESTAMP NOT NULL,
            criado_em TIMESTAMP DEFAULT NOW(),
            usado_em TIMESTAMP
        )
        """
        )
    )
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_refresh_hash ON refresh_tokens (token_hash)"))
    conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS mfa_secret TEXT"))
    conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS mfa_ativado BOOLEAN DEFAULT FALSE"))
    conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS mfa_backup_codes JSONB"))


# Flag de processo: garante que o schema multiusuário (contas/role/conta_id)
# foi criado/migrado uma vez antes do primeiro acesso autenticado.
_schema_multiusuario_pronto = False


ROLES_VALIDAS = ("vendedor", "supervisor", "gerente")
# Como cada função aparece para o usuário final ("Função", não "Papel").
ROTULO_FUNCAO = {"vendedor": "Vendedor", "supervisor": "Supervisor", "gerente": "Gerente"}


# =========================================================================
# RECURSOS POR PLANO (feature flags da assinatura)
# =========================================================================
# Fonte unica do que cada plano libera. A UI NAO decide isso: ela recebe a
# lista pronta em /me e pergunta `temRecurso("insights")` num helper so. Sem
# isto, "recurso pago" vira `if (plano === 'x')` espalhado por tela, e a
# primeira mudanca de pacote passa a exigir cacar ocorrencia.
#
# Insights e o primeiro recurso adicional: a tela de analise que o gerente usa
# para cravar meta. Todo mundo esta em 'completo' hoje, entao ninguem perde
# nada -- o gating existe, ligado.
RECURSOS_POR_PLANO: dict[str, set[str]] = {
    "completo": {"insights"},
    "essencial": set(),
}
PLANO_PADRAO = "completo"


def recursos_do_plano(plano: str | None) -> list[str]:
    """Recursos liberados. Plano desconhecido cai no padrao em vez de bloquear:
    conta com plano digitado errado no banco perdendo tela e pior do que conta
    vendo tela a mais."""
    return sorted(RECURSOS_POR_PLANO.get((plano or "").strip().lower(),
                                         RECURSOS_POR_PLANO[PLANO_PADRAO]))


def exigir_recurso(nome: str):
    """Dependencia para fechar uma rota atras do plano.

    Preparada e nao aplicada: nenhuma rota e gateada hoje porque toda conta
    esta em 'completo'. Quando houver plano sem insights, e so somar
    `Depends(exigir_recurso("insights"))` na rota -- a checagem ja e server
    side, entao esconder o item no menu nao vira a unica barreira.
    """

    def _checar(auth: dict = Depends(get_auth)) -> dict:
        with engine.connect() as conn:
            plano = conn.execute(
                text("SELECT plano FROM contas WHERE conta_id = :cid"),
                {"cid": auth["conta_id"]},
            ).scalar()
        if nome not in recursos_do_plano(plano):
            raise HTTPException(403, f"Recurso '{nome}' nao esta incluido no seu plano")
        return auth

    return _checar


def normalizar_role(valor: str | None, padrao: str = "vendedor") -> str:
    """Aceita só as três funções conhecidas; qualquer outra coisa vira o padrão.
    Mantém o formato já usado no banco (string minúscula), sem enum novo."""
    r = (valor or "").strip().lower()
    return r if r in ROLES_VALIDAS else padrao


def get_auth(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    """Identidade completa do usuário autenticado: além do email, traz a conta
    (assinatura) e a função (gerente/supervisor/vendedor). É a base do controle
    de acesso:
    - vendedor: enxerga apenas a própria carteira;
    - supervisor: a própria carteira + a dos vendedores atribuídos a ele;
    - gerente: enxerga tudo da conta dele."""
    global _schema_multiusuario_pronto
    email = get_current_user(credentials)
    with engine.begin() as conn:
        if not _schema_multiusuario_pronto:
            garantir_multiusuario(conn)
            garantir_seguranca(conn)
            _schema_multiusuario_pronto = True
        row = conn.execute(
            text("SELECT usuario_id, email, conta_id, role, supervisor_id FROM usuarios WHERE email = :e"),
            {"e": email},
        ).fetchone()
    if not row:
        raise HTTPException(401, "Usuário não encontrado")
    role = normalizar_role(row.role)
    return {
        "email": row.email,
        "usuario_id": str(row.usuario_id),
        "conta_id": str(row.conta_id) if row.conta_id else None,
        "role": role,
        "is_gerente": role == "gerente",
        "is_supervisor": role == "supervisor",
        "supervisor_id": str(row.supervisor_id) if row.supervisor_id else None,
    }


def exigir_gerente(auth: dict = Depends(get_auth)) -> dict:
    """Dependência para rotas restritas ao gerente (ADM da assinatura).
    Supervisor NÃO herda estas permissões — ele acompanha, não administra."""
    if not auth["is_gerente"]:
        raise HTTPException(403, "Acesso restrito ao gerente da conta")
    return auth


def exigir_gestor(auth: dict = Depends(get_auth)) -> dict:
    """Rotas de acompanhamento de equipe: gerente vê a conta inteira, supervisor
    vê só o próprio ramo. O escopo dos dados é aplicado dentro de cada rota."""
    if not (auth["is_gerente"] or auth["is_supervisor"]):
        raise HTTPException(403, "Acesso restrito a gerentes e supervisores")
    return auth


def escopo_vendedores(conn, auth: dict):
    """IDs de usuário cujos dados o autenticado pode enxergar.

    ESCOLHA DE MODELO: a camada "Equipe" do organograma é representada pelo
    próprio vínculo `usuarios.supervisor_id` — a equipe de um supervisor é o
    conjunto de vendedores que apontam para ele. Não foi criada tabela
    `equipes` porque o projeto não tinha esse conceito e cada vendedor pertence
    a um único supervisor; uma tabela extra só duplicaria a mesma relação.

    Devolve None quando não há restrição por dono (gerente vê a conta inteira).
    """
    if auth["is_gerente"]:
        return None
    if auth["is_supervisor"]:
        subordinados = conn.execute(
            text("SELECT usuario_id FROM usuarios WHERE conta_id = :cid AND supervisor_id = :sid"),
            {"cid": auth["conta_id"], "sid": auth["usuario_id"]},
        ).fetchall()
        return [auth["usuario_id"]] + [str(r.usuario_id) for r in subordinados]
    return [auth["usuario_id"]]


def filtro_escopo(conn, auth: dict, coluna: str = "vendedor_id", prefixo: str = ""):
    """Monta o par (trecho SQL, params) que aplica o escopo de carteira.

    Uso: `where = f"WHERE conta_id = :cid {trecho}"`. Para gerente o trecho é
    vazio; para supervisor/vendedor vira um `AND ... = ANY(:vids)`."""
    ids = escopo_vendedores(conn, auth)
    if ids is None:
        return "", {"cid": auth["conta_id"]}
    col = f"{prefixo}{coluna}" if prefixo else coluna
    return f"AND {col} = ANY(CAST(:vids AS uuid[]))", {"cid": auth["conta_id"], "vids": ids}


def escopo_emails(conn, auth: dict):
    """Mesma ideia de escopo_vendedores, mas em emails — usado pelas tabelas que
    identificam o dono por `usuario_email` (eventos/atividades)."""
    ids = escopo_vendedores(conn, auth)
    if ids is None:
        return None
    rows = conn.execute(
        text("SELECT email FROM usuarios WHERE usuario_id = ANY(CAST(:ids AS uuid[]))"), {"ids": ids}
    ).fetchall()
    return [r.email for r in rows] or [auth["email"]]


def checar_acesso_empresa(conn, empresa_id: str, auth: dict):
    """Garante que a empresa pertence à conta do usuário e, para vendedores e
    supervisores, que ela está no escopo dele. Levanta 404 (não revela
    existência fora do escopo)."""
    row = conn.execute(
        text("SELECT conta_id, vendedor_id FROM empresas WHERE empresa_id = :id"),
        {"id": empresa_id},
    ).fetchone()
    if not row:
        raise HTTPException(404, "Empresa não encontrada")
    if auth["conta_id"] and str(row.conta_id) != auth["conta_id"]:
        raise HTTPException(404, "Empresa não encontrada")
    ids = escopo_vendedores(conn, auth)
    if ids is not None and str(row.vendedor_id) not in ids:
        raise HTTPException(404, "Empresa não encontrada")
    return row


# =========================
# EMAIL (Resend)
# =========================
# Remetente de TODO email que o sistema manda (convite de usuário, ativação de
# conta e orçamento). O padrão `onboarding@resend.dev` é o endereço de SANDBOX
# do Resend: com ele a API só entrega para o email dono da conta Resend — mandar
# convite para qualquer outro destinatário volta 403
# ("You can only send testing emails to your own email address").
# Para mandar email para a equipe ou para clientes é preciso verificar um domínio no Resend
# (Domains → Add Domain, publicar os registros DNS) e apontar RESEND_FROM para
# um endereço desse domínio, ex.: "ProspectaGeo <nao-responda@seudominio.com.br>".
REMETENTE_EMAIL = os.getenv("RESEND_FROM") or "onboarding@resend.dev"


def link_ativacao(token: str) -> str:
    return f"{FRONTEND_ORIGINS[0]}/ativar?token={token}"


def motivo_falha_email(erro: Exception) -> str:
    """Transforma o erro cru do Resend em algo que o usuário consiga agir.

    O 403 do remetente de sandbox é o caso comum e o mais enganoso: a mensagem
    original fala de "testing emails" e não deixa claro que o problema é o
    remetente, não o destinatário."""
    motivo = str(erro)
    if "own email address" in motivo or "testing emails" in motivo:
        return (
            f"O remetente {REMETENTE_EMAIL} é o endereço de teste do Resend e só entrega "
            "para o email dono da conta Resend. Verifique um domínio no Resend e defina a "
            "variável RESEND_FROM para liberar o envio a outros destinatários."
        )
    if "domain is not verified" in motivo or "not verified" in motivo:
        return (
            f"O domínio do remetente {REMETENTE_EMAIL} não está verificado no Resend. "
            "Conclua a verificação em Domains → Add Domain."
        )
    return motivo


def endereco_de_resposta(conn, usuario_email: str) -> str:
    """Para onde a resposta do cliente deve voltar.

    Tem de ser a caixa que esta sendo observada pelo watch do Gmail/Outlook --
    e o e-mail de login nao e necessariamente essa. Sem reply_to, a resposta
    volta para REMETENTE_EMAIL (o remetente do Resend), que ninguem le: a
    notificacao de resposta fica impossivel, nao apenas quebrada.
    """
    row = conn.execute(
        text("""
            SELECT email_address FROM email_subscriptions
            WHERE usuario_email = :email
              AND provider IN ('gmail', 'outlook')
              AND email_address IS NOT NULL
            ORDER BY atualizado_em DESC
            LIMIT 1
        """),
        {"email": usuario_email},
    ).fetchone()
    return row.email_address if row and row.email_address else usuario_email


async def enviar_email(destino: str, token: str) -> tuple[bool, Optional[str]]:
    """Envia o convite de ativação.

    Devolve (enviado, motivo_da_falha) em vez de levantar exceção: o usuário já
    foi gravado quando esta função roda, e deixar o erro subir devolvia 500 para
    uma operação que, do ponto de vista do banco, deu certo — o gerente via
    "não foi possível criar o usuário", tentava de novo e batia em
    "Email já cadastrado". Quem chama decide o que fazer com a falha."""
    link = link_ativacao(token)
    if not resend.api_key:
        return False, "RESEND_API_KEY não está configurada no servidor."
    try:
        resend.Emails.send(
            {
                "from": REMETENTE_EMAIL,
                "to": destino,
                "subject": "Ative sua conta 🚀",
                "html": f"<p>Olá!</p><p>Clique no link abaixo para criar sua senha:</p><p><a href='{link}'>{link}</a></p>",
            }
        )
        return True, None
    except Exception as e:
        print(f"❌ Falha ao enviar convite para {destino} (from={REMETENTE_EMAIL}): {e}")
        return False, motivo_falha_email(e)


# =========================
# MODELOS
# =========================
class UsuarioCreate(BaseModel):
    nome: str
    email: EmailStr
    telefone: str | None = None
    role: str | None = None           # 'vendedor' (padrão) | 'supervisor' | 'gerente'
    supervisor_id: str | None = None  # só faz sentido quando role = 'vendedor'


class UsuarioGerenciar(BaseModel):
    ativo: bool | None = None
    role: str | None = None           # 'vendedor' | 'supervisor' | 'gerente'
    # Chega como string (atribuir) ou None. Para DESVINCULAR sem excluir o
    # usuário, o cliente manda `limpar_supervisor: true` — necessário porque
    # None é indistinguível de "campo não enviado".
    supervisor_id: str | None = None
    limpar_supervisor: bool = False


class ContaSignup(BaseModel):
    empresa_nome: str       # nome da conta/empresa que assina
    nome: str               # nome do gerente (ADM)
    email: EmailStr
    telefone: str | None = None


class UsuarioUpdate(BaseModel):
    nome: str | None = None
    telefone: str | None = None
    cargo: str | None = None
    empresa_nome: str | None = None
    bio: str | None = None


class EmpresaCreate(BaseModel):
    nome: str
    segmento: str | None = None
    porte: str | None = None
    cidade: str | None = None
    endereco: str | None = None
    numero: str | None = None
    cep: str | None = None
    bairro: str | None = None
    regiao: str | None = None
    observacoes: str | None = None
    cnpj: str | None = None
    site: str | None = None
    linkedin_empresa: str | None = None
    responsavel_principal: str | None = None
    status: str | None = None
    origem_lead: str | None = None
    ultima_interacao: datetime | None = None
    proxima_acao: str | None = None
    data_proxima_acao: date | None = None
    motivo_perdido: str | None = None
    temperatura: str | None = None
    logo_url: str | None = None
    # snapshot do Google Places (vindos da tela de busca/prefill)
    google_place_id: str | None = None
    latitude: float | None = None
    longitude: float | None = None
    google_rating: float | None = None
    google_rating_count: int | None = None
    business_status: str | None = None
    google_synced_at: datetime | None = None


class EmpresaUpdate(BaseModel):
    nome: str | None = None
    logo_url: str | None = None
    segmento: str | None = None
    porte: str | None = None
    cidade: str | None = None
    endereco: str | None = None
    numero: str | None = None
    cep: str | None = None
    bairro: str | None = None
    regiao: str | None = None
    observacoes: str | None = None
    cnpj: str | None = None
    site: str | None = None
    linkedin_empresa: str | None = None
    responsavel_principal: str | None = None
    status: str | None = None
    status_cadastro: str | None = None
    origem_lead: str | None = None
    ultima_interacao: datetime | None = None
    proxima_acao: str | None = None
    data_proxima_acao: date | None = None
    motivo_perdido: str | None = None
    temperatura: str | None = None


class SegmentoCreate(BaseModel):
    nome: str


class EventoCreate(BaseModel):
    titulo: str
    tipo: str
    data: date
    hora_inicio: str
    hora_fim: Optional[str] = None
    empresa_id: Optional[str] = None
    empresa_nome: Optional[str] = None
    descricao: Optional[str] = None
    email_convidado: Optional[str] = None


class EventoUpdate(BaseModel):
    titulo: str | None = None
    tipo: str | None = None
    data: date | None = None
    hora_inicio: str | None = None
    hora_fim: str | None = None
    empresa_id: str | None = None
    empresa_nome: str | None = None
    descricao: str | None = None
    email_convidado: str | None = None


class AtivarConta(BaseModel):
    token: str
    senha: str


class Login(BaseModel):
    email: EmailStr
    senha: str
    mfa_code: Optional[str] = None


class Token(BaseModel):
    access_token: str
    token_type: str


class MFAAtivar(BaseModel):
    code: str


class MFADesativar(BaseModel):
    senha: str


class ContatoUpdate(BaseModel):
    nome: str | None = None
    funcao: str | None = None
    email: str | None = None
    celular: str | None = None
    whatsapp: str | None = None
    linkedin: str | None = None
    observacoes: str | None = None
    prioridade: str | None = None
    nivel_influencia: str | None = None
    decisor: bool | None = None
    canal_preferido: str | None = None
    data_ultimo_contato: date | None = None


class ReuniaoOutlook(BaseModel):
    titulo: str
    descricao: Optional[str] = None
    data: date
    hora_inicio: str
    hora_fim: str
    email_convidado: Optional[str] = None
    emails_convidados: Optional[list[str]] = None


# Teto do circulo aceito pelo locationBias da Places API. Acima disso a busca
# passa a usar retangulo (ver _bias_localizacao).
RAIO_CIRCULO_MAX_M = 50_000
# Teto nosso, so para barrar valor absurdo digitado no campo livre. 2.000 km
# cobre o Brasil inteiro a partir de qualquer ponto dele.
RAIO_MAX_M = 2_000_000


class PlacesSearchRequest(BaseModel):
    query: str
    lat: float | None = None
    lng: float | None = None
    radius: int = 15000


class PlacesTetoUpdate(BaseModel):
    ligado: bool


class RascunhoCreate(BaseModel):
    google_place_id: str | None = None
    nome: str
    endereco_completo: str | None = None
    cidade: str | None = None
    latitude: float | None = None
    longitude: float | None = None
    telefone_empresa: str | None = None
    site: str | None = None
    google_rating: float | None = None
    google_rating_count: int | None = None
    business_status: str | None = None
    segmento: str | None = None


class ReuniaoGoogle(BaseModel):
    titulo: str
    descricao: Optional[str] = None
    data: date
    hora_inicio: str
    hora_fim: str
    email_convidado: Optional[str] = None
    emails_convidados: Optional[list[str]] = None


class EquipamentoCreate(BaseModel):
    nome: str
    codigo: Optional[str] = None      # SKU — identificador único dentro do catálogo
    descricao: Optional[str] = None
    preco_base: float = 0
    quantidade: Optional[int] = 0
    tipo: Optional[str] = None        # "equipamento" (padrão) ou "servico"


class EquipamentoUpdate(BaseModel):
    nome: Optional[str] = None
    codigo: Optional[str] = None
    descricao: Optional[str] = None
    preco_base: Optional[float] = None
    quantidade: Optional[int] = None
    ativo: Optional[bool] = None
    tipo: Optional[str] = None


class ObservacaoCreate(BaseModel):
    texto: str
    # Marcador livre (Comercial, Financeiro, Logistica...) -- a lista fica no
    # front de proposito: e vocabulario de equipe, muda sem migracao.
    marcador: Optional[str] = None


class OrcamentoItemIn(BaseModel):
    equipamento_id: Optional[str] = None
    descricao: str
    quantidade: int = 1
    preco_unitario: float = 0
    # So faz sentido em item AVULSO: quando ha equipamento_id, o tipo e o do
    # catalogo e gravar copia aqui criaria duas verdades que divergem no dia em
    # que alguem reclassificar o item. Sem esta coluna, servico avulso caia no
    # grafico de equipamentos -- era o buraco do item 1.3.
    tipo: Optional[str] = None


class OrcamentoCreate(BaseModel):
    empresa_id: str
    titulo: Optional[str] = None
    observacoes: Optional[str] = None
    itens: list[OrcamentoItemIn] = []


class OrcamentoUpdate(BaseModel):
    titulo: Optional[str] = None
    observacoes: Optional[str] = None
    itens: Optional[list[OrcamentoItemIn]] = None


class OrcamentoStatusUpdate(BaseModel):
    status: str
    motivo_recusa: Optional[str] = None


# =========================
# SEGMENTOS (helpers)
# =========================
def normalizar_texto(valor: str) -> str:
    sem_acentos = unicodedata.normalize("NFD", valor.lower())
    sem_acentos = "".join(ch for ch in sem_acentos if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", sem_acentos).strip()


def limpar_segmento(nome: str) -> str:
    return re.sub(r"\s+", " ", nome.strip())[:120]


def segmento_valido(nome: str) -> bool:
    nome_limpo = limpar_segmento(nome)
    normalizado = normalizar_texto(nome_limpo)
    if len(normalizado) < 3 or not re.search(r"[a-z]", normalizado):
        return False
    segmentos_base = {normalizar_texto(segmento) for segmento in SEGMENTOS_PADRAO}
    if normalizado in segmentos_base:
        return True
    palavras = set(re.findall(r"[a-z0-9]+", normalizado))
    return bool(palavras & PALAVRAS_CHAVE_SEGMENTO)


@uma_vez
def garantir_tabela_segmentos(conn):
    conn.execute(
        text(
            """
        CREATE TABLE IF NOT EXISTS segmentos (
            segmento_id uuid PRIMARY KEY,
            nome character varying(120) NOT NULL,
            nome_normalizado character varying(120) UNIQUE NOT NULL,
            criado_em timestamp without time zone DEFAULT CURRENT_TIMESTAMP
        )
    """
        )
    )
    for segmento in SEGMENTOS_PADRAO:
        nome = limpar_segmento(segmento)
        conn.execute(
            text(
                """
                INSERT INTO segmentos (segmento_id, nome, nome_normalizado)
                VALUES (:id, :nome, :nome_normalizado)
                ON CONFLICT (nome_normalizado) DO NOTHING
            """
            ),
            {"id": str(uuid.uuid4()), "nome": nome, "nome_normalizado": normalizar_texto(nome)},
        )


def salvar_segmento(conn, nome: str) -> str:
    nome_limpo = limpar_segmento(nome)
    if not segmento_valido(nome_limpo):
        raise HTTPException(400, "Segmento nao reconhecido.")
    garantir_tabela_segmentos(conn)
    conn.execute(
        text(
            """
            INSERT INTO segmentos (segmento_id, nome, nome_normalizado)
            VALUES (:id, :nome, :nome_normalizado)
            ON CONFLICT (nome_normalizado) DO UPDATE SET nome = EXCLUDED.nome
        """
        ),
        {"id": str(uuid.uuid4()), "nome": nome_limpo, "nome_normalizado": normalizar_texto(nome_limpo)},
    )
    return nome_limpo


@uma_vez
def garantir_campos_pipeline(conn):
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS data_proxima_acao date"))
    conn.execute(
        text(
            "ALTER TABLE empresas ADD COLUMN IF NOT EXISTS status_atualizado_em timestamp without time zone DEFAULT CURRENT_TIMESTAMP"
        )
    )
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS motivo_perdido text"))
    # Logo da empresa. Guardada como data URL (o form reduz a imagem para 256px
    # antes de enviar) porque o projeto nao tem bucket de arquivos.
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS logo_url text"))
    # Numero do imovel separado da rua. Juntos no mesmo campo, o Nominatim nao
    # resolve o endereco; separados, da para montar o "212 Rua X" que ele espera.
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS numero text"))
    # Data de entrada da empresa na base. Sem ela nao da para dizer quantos
    # clientes existiam num mes passado -- status_atualizado_em so conta a
    # ultima mudanca de status, e o historico so tem quem chegou a mudar.
    # Sem DEFAULT de proposito: preencher as linhas antigas com a data da
    # migration diria que todas nasceram hoje.
    conn.execute(
        text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS criado_em timestamp without time zone")
    )
    # Backfill unico: para quem ja mudou de status alguma vez, a primeira
    # entrada do historico e a melhor aproximacao que existe. Quem nunca mudou
    # fica NULL, e o grafico simplesmente nao conta essa empresa no passado.
    conn.execute(
        text(
            """
        UPDATE empresas e SET criado_em = h.primeiro
        FROM (
            SELECT empresa_id, MIN(alterado_em) AS primeiro
            FROM empresa_status_historico GROUP BY empresa_id
        ) h
        WHERE h.empresa_id = e.empresa_id AND e.criado_em IS NULL
    """
        )
    )
    conn.execute(
        text(
            """
        CREATE TABLE IF NOT EXISTS empresa_status_historico (
            historico_id uuid PRIMARY KEY,
            empresa_id uuid NOT NULL,
            status_anterior character varying(50),
            status_novo character varying(50) NOT NULL,
            observacao text,
            alterado_em timestamp without time zone DEFAULT CURRENT_TIMESTAMP
        )
    """
        )
    )


@uma_vez
def garantir_colunas_places(conn):
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS google_place_id TEXT"))
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS latitude DOUBLE PRECISION"))
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS longitude DOUBLE PRECISION"))
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS google_rating DOUBLE PRECISION"))
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS google_rating_count INTEGER"))
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS endereco_completo TEXT"))
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS telefone_empresa TEXT"))
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS business_status TEXT"))
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS google_synced_at TIMESTAMP WITHOUT TIME ZONE"))
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS status_cadastro TEXT DEFAULT 'ativo'"))


@uma_vez
def garantir_tabelas_places_cache(conn):
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS places_cache (
            id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            query TEXT NOT NULL,
            lat_grid DOUBLE PRECISION NOT NULL,
            lng_grid DOUBLE PRECISION NOT NULL,
            results JSONB NOT NULL,
            search_count INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(query, lat_grid, lng_grid)
        )
    """))
    # O raio faz parte da identidade do resultado: a mesma query no mesmo ponto
    # com 20 km e com 300 km sao buscas diferentes. Sem ele na chave, escolher
    # um raio maior devolvia o resultado estreito gravado antes -- e em silencio,
    # porque cache servido nao da erro.
    conn.execute(text(
        "ALTER TABLE places_cache ADD COLUMN IF NOT EXISTS raio_grid INTEGER DEFAULT 15"))
    conn.execute(text(
        "UPDATE places_cache SET raio_grid = 15 WHERE raio_grid IS NULL"))
    # A UNIQUE antiga (sem raio) sai; entra um indice com as quatro colunas. O
    # ON CONFLICT do INSERT depende dele existir. E seguro rodar sobre dados:
    # a chave nova e mais especifica que a antiga, entao nao ha como colidir.
    conn.execute(text(
        "ALTER TABLE places_cache DROP CONSTRAINT IF EXISTS places_cache_query_lat_grid_lng_grid_key"))
    conn.execute(text(
        """CREATE UNIQUE INDEX IF NOT EXISTS places_cache_chave
           ON places_cache (query, lat_grid, lng_grid, raio_grid)"""))
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS places_ranking (
            id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            query TEXT NOT NULL,
            lat_grid DOUBLE PRECISION NOT NULL,
            lng_grid DOUBLE PRECISION NOT NULL,
            results JSONB NOT NULL,
            search_count INTEGER NOT NULL,
            rank_position INTEGER NOT NULL,
            month VARCHAR(7) NOT NULL,
            saved_date TIMESTAMP DEFAULT NOW()
        )
    """))


@uma_vez
def garantir_tabela_places_uso(conn):
    """Contador de chamadas PAGAS ao Google Places, por usuario e por mes.

    `chave` e o usuario_id de quem buscou: cada um tem o proprio balde, entao
    ninguem gasta a cota do colega e da para ver quem consumiu o que. A coluna e
    TEXT (nao UUID) porque usuario sem usuario_id cai no proprio e-mail.
    `mes` e 'AAAA-MM', mesmo formato de places_ranking.month.
    Nao confundir com places_cache.search_count, que e popularidade do termo
    para o ranking e conta tambem os acertos de cache (que nao custam nada).
    As linhas do mes anterior ficam: sao o historico de consumo por vendedor."""
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS places_uso_mensal (
            chave TEXT NOT NULL,
            mes VARCHAR(7) NOT NULL,
            chamadas INTEGER NOT NULL DEFAULT 0,
            atualizado_em TIMESTAMP NOT NULL DEFAULT NOW(),
            PRIMARY KEY (chave, mes)
        )
    """))


@uma_vez
def garantir_tabela_places_teto(conn):
    """Interruptor do teto mensal, por conta (assinatura).

    Ausencia de linha = teto LIGADO. O default seguro e o limite valendo: se
    esta tabela sumisse, a conta volta a ser protegida, nao o contrario."""
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS places_teto_conta (
            conta_id TEXT PRIMARY KEY,
            ligado BOOLEAN NOT NULL DEFAULT TRUE,
            alterado_em TIMESTAMP NOT NULL DEFAULT NOW(),
            alterado_por TEXT
        )
    """))


@uma_vez
def garantir_colunas_oauth(conn):
    conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS outlook_access_token text"))
    conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS outlook_refresh_token text"))
    conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS google_access_token text"))
    conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS google_refresh_token text"))


@uma_vez
def garantir_multiusuario(conn):
    """Schema do modelo multiusuário (assinatura → gerente + vendedores):
    - tabela `contas` (a assinatura, paga pelo ADM/gerente);
    - `usuarios.conta_id` + `usuarios.role` ('gerente' | 'vendedor');
    - `empresas.conta_id` + `empresas.vendedor_id` (dono da carteira);
    - `eventos.conta_id` (para o gerente ver a agenda de todos).
    Inclui a migração do pool antigo (dados sem conta) para uma conta inicial."""
    conn.execute(
        text(
            """
        CREATE TABLE IF NOT EXISTS contas (
            conta_id uuid PRIMARY KEY,
            nome text NOT NULL,
            criado_em timestamp without time zone DEFAULT CURRENT_TIMESTAMP
        )
    """
        )
    )
    # Plano da assinatura. Default 'completo' de proposito: a cobranca por
    # modulo ainda nao existe, e nascer com tudo ligado significa que ligar o
    # gating nao tira nada de ninguem hoje. O dia de cobrar e mudar o default
    # e o plano das contas -- nao mexer em tela.
    conn.execute(text("ALTER TABLE contas ADD COLUMN IF NOT EXISTS plano text DEFAULT 'completo'"))
    conn.execute(text("UPDATE contas SET plano = 'completo' WHERE plano IS NULL"))
    conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS conta_id uuid"))
    conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS role text DEFAULT 'vendedor'"))
    # Hierarquia Gerente → Supervisor → Vendedor. O vínculo é uma auto-referência
    # em `usuarios`: cada vendedor aponta para no máximo um supervisor. Não há
    # tabela `equipes` — a "equipe" é o conjunto de vendedores de um supervisor
    # (ver ESCOLHA DE MODELO no topo de escopo_vendedores).
    conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS supervisor_id uuid"))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_usuarios_supervisor ON usuarios(supervisor_id)"))
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS conta_id uuid"))
    conn.execute(text("ALTER TABLE empresas ADD COLUMN IF NOT EXISTS vendedor_id uuid"))
    conn.execute(text("ALTER TABLE eventos ADD COLUMN IF NOT EXISTS conta_id uuid"))

    # ---- Migração do pool antigo (idempotente) ----
    orfaos_usuario = conn.execute(text("SELECT 1 FROM usuarios WHERE conta_id IS NULL LIMIT 1")).fetchone()
    orfaos_empresa = conn.execute(text("SELECT 1 FROM empresas WHERE conta_id IS NULL LIMIT 1")).fetchone()
    if not (orfaos_usuario or orfaos_empresa):
        return

    conta_id = conn.execute(text("SELECT conta_id FROM contas ORDER BY criado_em ASC LIMIT 1")).scalar()
    if not conta_id:
        conta_id = str(uuid.uuid4())
        conn.execute(
            text("INSERT INTO contas (conta_id, nome) VALUES (:id, :nome)"),
            {"id": conta_id, "nome": "Conta Principal"},
        )

    # Vincula usuários órfãos à conta inicial e normaliza role
    conn.execute(text("UPDATE usuarios SET conta_id = :cid WHERE conta_id IS NULL"), {"cid": conta_id})
    conn.execute(text("UPDATE usuarios SET role = 'vendedor' WHERE role IS NULL"))

    # O usuário mais antigo da conta vira gerente, se ainda não houver um
    tem_gerente = conn.execute(
        text("SELECT 1 FROM usuarios WHERE conta_id = :cid AND role = 'gerente' LIMIT 1"),
        {"cid": conta_id},
    ).fetchone()
    if not tem_gerente:
        conn.execute(
            text(
                """
                UPDATE usuarios SET role = 'gerente'
                WHERE usuario_id = (
                    SELECT usuario_id FROM usuarios WHERE conta_id = :cid
                    ORDER BY data_criacao ASC NULLS LAST LIMIT 1
                )
            """
            ),
            {"cid": conta_id},
        )

    # Vincula empresas órfãs à conta e tenta inferir o dono pelo responsavel_principal
    conn.execute(text("UPDATE empresas SET conta_id = :cid WHERE conta_id IS NULL"), {"cid": conta_id})
    conn.execute(
        text(
            """
            UPDATE empresas e SET vendedor_id = u.usuario_id
            FROM usuarios u
            WHERE e.vendedor_id IS NULL
              AND e.responsavel_principal IS NOT NULL
              AND lower(u.email) = lower(e.responsavel_principal)
        """
        )
    )
    # Eventos herdam a conta do dono (usuario_email)
    conn.execute(
        text(
            """
            UPDATE eventos ev SET conta_id = u.conta_id
            FROM usuarios u
            WHERE ev.conta_id IS NULL AND lower(u.email) = lower(ev.usuario_email)
        """
        )
    )


@uma_vez
def garantir_tabela_notificacoes(conn):
    conn.execute(
        text(
            """
        CREATE TABLE IF NOT EXISTS notificacoes (
            notificacao_id UUID PRIMARY KEY,
            usuario_email TEXT NOT NULL,
            tipo TEXT NOT NULL,
            titulo TEXT NOT NULL,
            mensagem TEXT NOT NULL,
            empresa_id UUID NULL,
            empresa_nome TEXT NULL,
            platform TEXT NULL,
            meta JSONB DEFAULT '{}'::jsonb,
            lida BOOLEAN DEFAULT FALSE,
            criado_em TIMESTAMP DEFAULT NOW()
        )
    """
        )
    )
    conn.execute(text("ALTER TABLE notificacoes ADD COLUMN IF NOT EXISTS platform VARCHAR(30)"))
    conn.execute(text("ALTER TABLE notificacoes ADD COLUMN IF NOT EXISTS meta JSONB DEFAULT '{}'"))
    conn.execute(
        text(
            """
        CREATE TABLE IF NOT EXISTS email_subscriptions (
            sub_id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            usuario_email TEXT NOT NULL,
            provider VARCHAR(20) NOT NULL,
            subscription_id TEXT,
            email_address TEXT,
            history_id BIGINT,
            expires_at TIMESTAMP,
            access_token TEXT,
            refresh_token TEXT,
            criado_em TIMESTAMP DEFAULT NOW(),
            atualizado_em TIMESTAMP DEFAULT NOW()
        )
    """
        )
    )
    conn.execute(text("ALTER TABLE eventos ADD COLUMN IF NOT EXISTS outlook_event_id TEXT"))
    conn.execute(text("ALTER TABLE eventos ADD COLUMN IF NOT EXISTS google_event_id TEXT"))
    conn.execute(text("ALTER TABLE eventos ADD COLUMN IF NOT EXISTS email_convidado TEXT"))
    conn.execute(text("ALTER TABLE eventos ADD COLUMN IF NOT EXISTS status_resposta TEXT DEFAULT 'pendente'"))
    # resourceId do canal de push do Google Calendar — precisa ser devolvido
    # junto do channel id para encerrar o canal antigo em channels/stop.
    conn.execute(text("ALTER TABLE email_subscriptions ADD COLUMN IF NOT EXISTS resource_id TEXT"))


# Status possíveis de um orçamento, na ordem do fluxo.
ORCAMENTO_STATUS = ["rascunho", "enviado", "em_negociacao", "aprovado", "recusado"]

# Os dois catalogos de venda. Moram na mesma tabela, separados por `tipo`.
TIPOS_CATALOGO = ["equipamento", "servico"]


def _tipo_catalogo(valor, padrao="equipamento"):
    """Normaliza o `tipo` vindo de query/body/form. Recusa valor desconhecido em
    vez de cair no padrao: filtrar por um tipo que nao existe devolveria o
    catalogo inteiro e o usuario veria servico na aba de equipamento."""
    if valor is None or valor == "":
        return padrao
    limpo = str(valor).strip().lower()
    if limpo not in TIPOS_CATALOGO:
        raise HTTPException(400, f"Tipo invalido: use {' ou '.join(TIPOS_CATALOGO)}")
    return limpo


@uma_vez
def garantir_vendas(conn):
    """Catálogo de equipamentos + orçamentos e seus itens.

    Escopo por conta (assinatura), igual a empresas: o vendedor enxerga o que é
    dele, o gerente enxerga tudo da conta."""
    conn.execute(
        text(
            """
        CREATE TABLE IF NOT EXISTS equipamentos (
            equipamento_id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            conta_id UUID NOT NULL,
            nome TEXT NOT NULL,
            descricao TEXT,
            preco_base NUMERIC(12,2) DEFAULT 0,
            ativo BOOLEAN DEFAULT TRUE,
            criado_em TIMESTAMP DEFAULT NOW()
        )
    """
        )
    )
    conn.execute(
        text(
            """
        CREATE TABLE IF NOT EXISTS orcamentos (
            orcamento_id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            conta_id UUID NOT NULL,
            empresa_id UUID NOT NULL,
            vendedor_id UUID,
            titulo TEXT,
            observacoes TEXT,
            status VARCHAR(20) NOT NULL DEFAULT 'rascunho',
            total NUMERIC(12,2) DEFAULT 0,
            data_envio TIMESTAMP,
            data_decisao TIMESTAMP,
            motivo_recusa TEXT,
            criado_em TIMESTAMP DEFAULT NOW(),
            atualizado_em TIMESTAMP DEFAULT NOW()
        )
    """
        )
    )
    conn.execute(
        text(
            """
        CREATE TABLE IF NOT EXISTS orcamento_itens (
            item_id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            orcamento_id UUID NOT NULL REFERENCES orcamentos(orcamento_id) ON DELETE CASCADE,
            equipamento_id UUID,
            descricao TEXT NOT NULL,
            quantidade INTEGER NOT NULL DEFAULT 1,
            preco_unitario NUMERIC(12,2) NOT NULL DEFAULT 0
        )
    """
        )
    )
    # Campos de estoque/catálogo usados pela importação de Excel. `codigo` é o
    # SKU: quando presente, é ele (e não o nome) que define se a linha importada
    # cria um item novo ou atualiza um existente.
    conn.execute(text("ALTER TABLE equipamentos ADD COLUMN IF NOT EXISTS codigo TEXT"))
    conn.execute(text("ALTER TABLE equipamentos ADD COLUMN IF NOT EXISTS quantidade INTEGER DEFAULT 0"))
    # `tipo` separa os dois catalogos que a tela de vendas mostra em abas.
    # Nao existe tabela `servicos`: servico e equipamento tem exatamente os
    # mesmos campos e o mesmo papel (virar item de orcamento), e a unica
    # diferenca e que servico nao tem estoque. Com um discriminador, servico
    # entra em orcamento, nas metricas por item e na ficha da empresa sem
    # alterar `orcamento_itens` nem nenhuma consulta existente. Decisao
    # consciente, igual a de nao criar tabela `equipes`.
    conn.execute(text(
        "ALTER TABLE equipamentos ADD COLUMN IF NOT EXISTS tipo TEXT DEFAULT 'equipamento'"))
    # Mesmo discriminador para o item AVULSO, que nao tem linha no catalogo.
    # Sem default: NULL aqui significa "veja o tipo do catalogo pelo
    # equipamento_id", e so o avulso preenche. Aditivo e idempotente, no mesmo
    # padrao das colunas acima -- nenhuma linha existente e reescrita.
    conn.execute(text(
        "ALTER TABLE orcamento_itens ADD COLUMN IF NOT EXISTS tipo TEXT"))
    # Linha nascida antes da coluna existir fica com NULL, e NULL nao casa com
    # `tipo = 'equipamento'` no filtro — sumiria da aba de equipamentos.
    conn.execute(text(
        "UPDATE equipamentos SET tipo = 'equipamento' WHERE tipo IS NULL"))
    conn.execute(
        text(
            """CREATE UNIQUE INDEX IF NOT EXISTS idx_equipamentos_codigo
               ON equipamentos(conta_id, lower(codigo)) WHERE codigo IS NOT NULL"""
        )
    )
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_orcamentos_conta ON orcamentos(conta_id)"))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_orcamentos_empresa ON orcamentos(empresa_id)"))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_orc_itens_orcamento ON orcamento_itens(orcamento_id)"))
    conn.execute(text("CREATE INDEX IF NOT EXISTS idx_equipamentos_conta ON equipamentos(conta_id)"))


@uma_vez
def garantir_observacoes(conn):
    """Feed de observacoes da empresa: uma linha por anotacao, com autor e data.

    A coluna `empresas.observacoes` continua existindo e nao muda de dono -- ela
    e o texto do cadastro, editado no formulario da empresa. Isto aqui e outra
    coisa: o que a equipe foi anotando ao longo do relacionamento, que num campo
    unico so daria para acumular apagando o que veio antes.
    """
    conn.execute(
        text(
            """
        CREATE TABLE IF NOT EXISTS empresa_observacoes (
            observacao_id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            empresa_id UUID NOT NULL REFERENCES empresas(empresa_id) ON DELETE CASCADE,
            autor_id UUID,
            texto TEXT NOT NULL,
            marcador TEXT,
            criado_em TIMESTAMP DEFAULT NOW()
        )
    """
        )
    )
    conn.execute(
        text("CREATE INDEX IF NOT EXISTS idx_empresa_obs ON empresa_observacoes(empresa_id, criado_em DESC)")
    )


def aplicar_migracoes():
    """Roda as migracoes na subida, numa transacao propria.

    Assim a primeira requisicao ja encontra o schema pronto e nenhuma rota paga
    a conta. Falha aqui nao derruba o processo: nada e marcado como feito, e o
    caminho antigo -- cada rota aplicando a sua -- volta a valer sozinho.
    """
    try:
        with engine.begin() as conn:
            garantir_seguranca(conn)
            garantir_tabela_segmentos(conn)
            garantir_campos_pipeline(conn)
            garantir_colunas_places(conn)
            garantir_tabelas_places_cache(conn)
            garantir_tabela_places_uso(conn)
            garantir_tabela_places_teto(conn)
            garantir_colunas_oauth(conn)
            garantir_multiusuario(conn)
            garantir_tabela_notificacoes(conn)
            garantir_vendas(conn)
            garantir_observacoes(conn)
        print("✅ migracoes aplicadas na subida:", len(_MIGRACOES_FEITAS))
    except Exception as e:  # noqa: BLE001 - qualquer falha cai no plano antigo
        print("⚠️ migracoes na subida falharam, cada rota aplica a sua:", e)


aplicar_migracoes()


# =========================
# JOB: RASCUNHOS EXPIRÁVEIS
# =========================
def verificar_rascunhos_expirados():
    print("⏰ JOB: verificando rascunhos expirados...")
    try:
        with engine.begin() as conn:
            garantir_tabela_notificacoes(conn)

            agora = datetime.utcnow()
            limite_aviso = agora - timedelta(days=25)
            limite_exclusao = agora - timedelta(days=30)

            # Rascunhos com 25-29 dias → aviso.
            # Notificação é isolada por usuário: vai SÓ para o dono (vendedor) do
            # rascunho. O gerente não recebe (evita volume alto); o que o gerente
            # vê é decidido à parte.
            avisos = conn.execute(
                text(
                    """
                SELECT e.empresa_id, e.nome, e.status_atualizado_em, u.email
                FROM empresas e
                JOIN usuarios u
                  ON u.ativo = TRUE
                 AND u.usuario_id = e.vendedor_id
                WHERE e.status = 'Rascunho'
                  AND e.status_atualizado_em <= :limite_aviso
                  AND e.status_atualizado_em > :limite_exclusao
            """
                ),
                {"limite_aviso": limite_aviso, "limite_exclusao": limite_exclusao},
            ).fetchall()

            for r in avisos:
                dias_restantes = 30 - int((agora - r.status_atualizado_em).days)
                existe = conn.execute(
                    text(
                        """
                    SELECT 1 FROM notificacoes
                    WHERE empresa_id = :eid AND tipo = 'rascunho_aviso'
                      AND usuario_email = :email
                      AND criado_em >= NOW() - INTERVAL '23 hours'
                """
                    ),
                    {"eid": r.empresa_id, "email": r.email},
                ).fetchone()
                if not existe:
                    conn.execute(
                        text(
                            """
                        INSERT INTO notificacoes
                            (notificacao_id, usuario_email, tipo, titulo, mensagem, empresa_id, empresa_nome)
                        VALUES (:id, :email, 'rascunho_aviso', :titulo, :mensagem, :eid, :enome)
                    """
                        ),
                        {
                            "id": str(uuid.uuid4()),
                            "email": r.email,
                            "titulo": f"Rascunho expira em {dias_restantes} dia{'s' if dias_restantes != 1 else ''}",
                            "mensagem": f"O rascunho '{r.nome}' será excluído automaticamente em {dias_restantes} dia{'s' if dias_restantes != 1 else ''}. Complete o cadastro para não perder.",
                            "eid": r.empresa_id,
                            "enome": r.nome,
                        },
                    )
                    print(f"📢 Aviso gerado para rascunho: {r.nome}")

            # Rascunhos com 30+ dias → excluir. Notifica SÓ o dono (vendedor).
            expirados = conn.execute(
                text(
                    """
                SELECT e.empresa_id, e.nome, u.email
                FROM empresas e
                JOIN usuarios u
                  ON u.ativo = TRUE
                 AND u.usuario_id = e.vendedor_id
                WHERE e.status = 'Rascunho'
                  AND e.status_atualizado_em <= :limite_exclusao
            """
                ),
                {"limite_exclusao": limite_exclusao},
            ).fetchall()

            empresas_excluidas = set()
            for r in expirados:
                conn.execute(
                    text(
                        """
                    INSERT INTO notificacoes
                        (notificacao_id, usuario_email, tipo, titulo, mensagem, empresa_id, empresa_nome)
                    VALUES (:id, :email, 'rascunho_excluido', :titulo, :mensagem, :eid, :enome)
                """
                    ),
                    {
                        "id": str(uuid.uuid4()),
                        "email": r.email,
                        "titulo": "Rascunho excluído automaticamente",
                        "mensagem": f"O rascunho '{r.nome}' foi excluído por inatividade após 30 dias. Cadastre novamente se necessário.",
                        "eid": r.empresa_id,
                        "enome": r.nome,
                    },
                )
                if r.empresa_id not in empresas_excluidas:
                    conn.execute(text("DELETE FROM contatos WHERE empresa_id = :id"), {"id": r.empresa_id})
                    conn.execute(text("DELETE FROM empresa_status_historico WHERE empresa_id = :id"), {"id": r.empresa_id})
                    conn.execute(text("DELETE FROM empresas WHERE empresa_id = :id"), {"id": r.empresa_id})
                    empresas_excluidas.add(r.empresa_id)
                    print(f"🗑️ Rascunho excluído: {r.nome}")

        print("✅ JOB: verificação de rascunhos concluída")
    except Exception as e:
        print(f"🔴 JOB ERRO: {str(e)}")


# =========================
# WEBHOOK HELPERS
# =========================

# Padrões de remetentes automáticos a ignorar
BLOCKED_SENDER_PATTERNS = [
    "noreply",
    "no-reply",
    "no_reply",
    "donotreply",
    "do-not-reply",
    "mailer-daemon",
    "postmaster",
    "bounce",
    "bounces",
    "notifications@",
    "notify@",
    "alert@",
    "alerts@",
    "system@",
    "auto@",
    "automated@",
    "autoresponder",
    "support@",
    "helpdesk@",
    "feedback@",
    "unsubscribe",
    "newsletter",
    "news@",
    "info@noreply",
    "microsoft@",
    "google@",
    "amazonses",
    "sendgrid",
    "mailchimp",
    "hubspot",
    "salesforce",
]


def is_automated_sender(email: str) -> bool:
    """Retorna True se o remetente parecer automático/noreply."""
    email_lower = email.lower()
    return any(pattern in email_lower for pattern in BLOCKED_SENDER_PATTERNS)


def find_company_by_sender(conn, sender_email: str, conta_id=None):
    # Quando conta_id é informado, restringe a busca às empresas daquela conta —
    # evita casar com empresa de outro tenant e vazar o nome em notificação.
    results = conn.execute(
        text(
            """
        SELECT
            c.empresa_id,
            c.contato_id,
            e.nome as empresa_nome,
            c.data_ultimo_contato,
            e.ultima_interacao
        FROM contatos c
        JOIN empresas e
            ON e.empresa_id = c.empresa_id
        WHERE LOWER(c.email) = LOWER(:email)
          AND (:conta_id IS NULL OR e.conta_id = :conta_id)
        ORDER BY
            c.decisor DESC NULLS LAST,
            c.data_ultimo_contato DESC NULLS LAST,
            e.ultima_interacao DESC NULLS LAST
        LIMIT 1
    """
        ),
        {"email": sender_email.strip(), "conta_id": conta_id},
    ).fetchone()

    if results:
        return (
            results._mapping["empresa_id"],
            results._mapping["contato_id"],
            results._mapping["empresa_nome"],
        )

    return None, None, None


def create_interaction_notification(
    conn,
    usuario_email: str,
    empresa_id,
    empresa_nome: str,
    platform: str,
    sender_name: str,
    sender_email: str,
    subject: str,
    conversation_id: str = "",
):
    cutoff = datetime.utcnow() - timedelta(minutes=1)
    existe = conn.execute(
        text(
            """
        SELECT 1 FROM notificacoes
        WHERE empresa_id = :eid AND tipo = 'email_interaction' AND platform = :platform
          AND meta->>'sender_email' = :semail
          AND criado_em >= :cutoff
    """
        ),
        {"eid": str(empresa_id), "platform": platform, "semail": sender_email, "cutoff": cutoff},
    ).fetchone()
    if existe:
        return

    label = "Gmail" if platform == "gmail" else "Outlook"
    conn.execute(
        text(
            """
        INSERT INTO notificacoes
            (notificacao_id, usuario_email, tipo, titulo, mensagem,
             empresa_id, empresa_nome, platform, meta, lida, criado_em)
        VALUES
            (:id, :email, 'email_interaction', :titulo, :mensagem,
             :eid, :enome, :platform, CAST(:meta AS JSONB), FALSE, NOW())
    """
        ),
        {
            "id": str(uuid.uuid4()),
            "email": usuario_email,
            "titulo": empresa_nome,
            "mensagem": f"Nova interação via {label}",
            "eid": str(empresa_id),
            "enome": empresa_nome,
            "platform": platform,
            "meta": json.dumps(
                {
                    "sender_email": sender_email,
                    "sender_name": sender_name,
                    "subject": subject,
                    "conversation_id": conversation_id,
                }
            ),
        },
    )


# =========================
# GMAIL WATCH
# =========================
def google_com_retry(usuario_email: str, access_token: str, refresh_token: str, fazer):
    """Executa fazer(token); se o Google devolver 401, renova e repete uma vez.

    Access token do Google dura 1 hora. Todo job que reusa um token guardado no
    banco falha depois disso -- e falhar aqui e silencioso, porque o job so
    imprime e volta. Era exatamente assim que o watch do Gmail morria: durava os
    7 dias do watch inicial e nenhuma renovacao depois disso conseguia passar.
    """
    resp = fazer(access_token)
    if resp.status_code == 401 and refresh_token:
        novo = asyncio.run(_refresh_google_token(refresh_token, usuario_email))
        if novo:
            return novo, fazer(novo)
    return access_token, resp


def setup_gmail_watch(usuario_email: str, access_token: str, refresh_token: str, gmail_address: str = ""):
    try:
        # Sem o endereco da caixa nao da para montar a URL do watch nem para o
        # webhook achar a assinatura depois. Quem chama nem sempre o tem.
        if not gmail_address:
            access_token, prof = google_com_retry(
                usuario_email, access_token, refresh_token,
                lambda t: http_requests.get(
                    "https://gmail.googleapis.com/gmail/v1/users/me/profile",
                    headers={"Authorization": f"Bearer {t}"},
                    timeout=15,
                ),
            )
            if prof.ok:
                gmail_address = prof.json().get("emailAddress", "")
            if not gmail_address:
                print(f"[Gmail Watch] endereco da caixa desconhecido para {usuario_email}: {prof.text}")
                return

        access_token, res = google_com_retry(
            usuario_email, access_token, refresh_token,
            lambda t: http_requests.post(
                f"https://gmail.googleapis.com/gmail/v1/users/{gmail_address}/watch",
                headers={"Authorization": f"Bearer {t}", "Content-Type": "application/json"},
                json={"topicName": GMAIL_PUBSUB_TOPIC, "labelIds": ["INBOX"]},
                timeout=15,
            ),
        )
        if not res.ok:
            print(f"[Gmail Watch] Erro ({res.status_code}) para {usuario_email}: {res.text}")
            return
        data = res.json()
        history_id = int(data.get("historyId", 0))
        expires_ms = int(data.get("expiration", 0))
        expires_at = datetime.utcfromtimestamp(expires_ms / 1000) if expires_ms else None
        with engine.begin() as conn:
            garantir_tabela_notificacoes(conn)
            existing = conn.execute(
                text(
                    """
                SELECT sub_id FROM email_subscriptions
                WHERE usuario_email = :email AND provider = 'gmail'
            """
                ),
                {"email": usuario_email},
            ).fetchone()
            if existing:
                conn.execute(
                    text(
                        """
                    UPDATE email_subscriptions
                    SET history_id=:hid, expires_at=:exp, access_token=:at,
                        refresh_token=:rt, email_address=:addr, atualizado_em=NOW()
                    WHERE usuario_email=:email AND provider='gmail'
                """
                    ),
                    {
                        "hid": history_id,
                        "exp": expires_at,
                        "at": access_token,
                        "rt": refresh_token,
                        "addr": gmail_address,
                        "email": usuario_email,
                    },
                )
            else:
                conn.execute(
                    text(
                        """
                    INSERT INTO email_subscriptions
                        (sub_id, usuario_email, provider, email_address, history_id,
                         expires_at, access_token, refresh_token)
                    VALUES (:id, :email, 'gmail', :addr, :hid, :exp, :at, :rt)
                """
                    ),
                    {
                        "id": str(uuid.uuid4()),
                        "email": usuario_email,
                        "addr": gmail_address,
                        "hid": history_id,
                        "exp": expires_at,
                        "at": access_token,
                        "rt": refresh_token,
                    },
                )
        print(f"[Gmail Watch] OK para {gmail_address}, historyId={history_id}")
    except Exception as e:
        print(f"[Gmail Watch] Exceção: {e}")


# =========================
# OUTLOOK EMAIL SUBSCRIPTION
# =========================
def outlook_com_retry(usuario_email: str, access_token: str, refresh_token: str, fazer):
    """Executa fazer(token); se o Graph devolver 401, renova e repete uma vez.

    Mesmo motivo do google_com_retry: token da Microsoft dura cerca de 1 hora,
    e todo job que reusa um token guardado no banco falha depois disso.
    """
    resp = fazer(access_token)
    if resp.status_code == 401 and refresh_token:
        novo = asyncio.run(_refresh_outlook_token(refresh_token, usuario_email))
        if novo:
            return novo, fazer(novo)
    return access_token, resp


def setup_outlook_subscription(usuario_email: str, access_token: str, refresh_token: str):
    try:
        expires_at = datetime.utcnow() + timedelta(minutes=4000)
        access_token, res = outlook_com_retry(
            usuario_email, access_token, refresh_token,
            lambda t: http_requests.post(
                "https://graph.microsoft.com/v1.0/subscriptions",
                headers={"Authorization": f"Bearer {t}", "Content-Type": "application/json"},
                json={
                    "changeType": "created",
                    "notificationUrl": f"{BACKEND_URL}/webhooks/outlook",
                    "resource": "me/mailFolders('Inbox')/messages",
                    "expirationDateTime": expires_at.strftime("%Y-%m-%dT%H:%M:%S.0000000Z"),
                    "clientState": OUTLOOK_WEBHOOK_SECRET,
                },
                timeout=15,
            ),
        )
        if not res.ok:
            print(f"[Outlook Sub] Erro ({res.status_code}) para {usuario_email}: {res.text}")
            return
        sub_id = res.json().get("id")
        me_res = http_requests.get(
            "https://graph.microsoft.com/v1.0/me",
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=10,
        )
        email_address = me_res.json().get("mail") or me_res.json().get("userPrincipalName", "")
        with engine.begin() as conn:
            garantir_tabela_notificacoes(conn)
            existing = conn.execute(
                text(
                    """
                SELECT sub_id FROM email_subscriptions
                WHERE usuario_email = :email AND provider = 'outlook'
            """
                ),
                {"email": usuario_email},
            ).fetchone()
            if existing:
                conn.execute(
                    text(
                        """
                    UPDATE email_subscriptions
                    SET subscription_id=:sid, expires_at=:exp, access_token=:at,
                        refresh_token=:rt, email_address=:addr, atualizado_em=NOW()
                    WHERE usuario_email=:email AND provider='outlook'
                """
                    ),
                    {
                        "sid": sub_id,
                        "exp": expires_at,
                        "at": access_token,
                        "rt": refresh_token,
                        "addr": email_address,
                        "email": usuario_email,
                    },
                )
            else:
                conn.execute(
                    text(
                        """
                    INSERT INTO email_subscriptions
                        (sub_id, usuario_email, provider, subscription_id,
                         email_address, expires_at, access_token, refresh_token)
                    VALUES (:id, :email, 'outlook', :sid, :addr, :exp, :at, :rt)
                """
                    ),
                    {
                        "id": str(uuid.uuid4()),
                        "email": usuario_email,
                        "sid": sub_id,
                        "addr": email_address,
                        "exp": expires_at,
                        "at": access_token,
                        "rt": refresh_token,
                    },
                )
        print(f"[Outlook Sub] OK para {email_address}, id={sub_id}")
    except Exception as e:
        print(f"[Outlook Sub] Exceção: {e}")


# =========================
# OUTLOOK CALENDAR SUBSCRIPTION
# =========================
def setup_outlook_calendar_subscription(usuario_email: str, access_token: str, refresh_token: str):
    try:
        expires_at = datetime.utcnow() + timedelta(minutes=4000)
        access_token, res = outlook_com_retry(
            usuario_email, access_token, refresh_token,
            lambda t: http_requests.post(
                "https://graph.microsoft.com/v1.0/subscriptions",
                headers={"Authorization": f"Bearer {t}", "Content-Type": "application/json"},
                json={
                    "changeType": "updated",
                    "notificationUrl": f"{BACKEND_URL}/webhooks/outlook-calendar",
                    "resource": "me/events",
                    "expirationDateTime": expires_at.strftime("%Y-%m-%dT%H:%M:%S.0000000Z"),
                    "clientState": OUTLOOK_WEBHOOK_SECRET,
                },
                timeout=15,
            ),
        )
        if not res.ok:
            print(f"[Outlook Cal Sub] Erro ({res.status_code}) para {usuario_email}: {res.text}")
            return
        sub_id = res.json().get("id")
        with engine.begin() as conn:
            garantir_tabela_notificacoes(conn)
            existing = conn.execute(
                text(
                    """
                SELECT sub_id FROM email_subscriptions
                WHERE usuario_email = :email AND provider = 'outlook_calendar'
            """
                ),
                {"email": usuario_email},
            ).fetchone()
            if existing:
                conn.execute(
                    text(
                        """
                    UPDATE email_subscriptions
                    SET subscription_id=:sid, expires_at=:exp, access_token=:at,
                        refresh_token=:rt, atualizado_em=NOW()
                    WHERE usuario_email=:email AND provider='outlook_calendar'
                """
                    ),
                    {"sid": sub_id, "exp": expires_at, "at": access_token, "rt": refresh_token, "email": usuario_email},
                )
            else:
                conn.execute(
                    text(
                        """
                    INSERT INTO email_subscriptions
                        (sub_id, usuario_email, provider, subscription_id,
                         expires_at, access_token, refresh_token)
                    VALUES (:id, :email, 'outlook_calendar', :sid, :exp, :at, :rt)
                """
                    ),
                    {
                        "id": str(uuid.uuid4()),
                        "email": usuario_email,
                        "sid": sub_id,
                        "exp": expires_at,
                        "at": access_token,
                        "rt": refresh_token,
                    },
                )
        print(f"[Outlook Cal Sub] OK para {usuario_email}, id={sub_id}")
    except Exception as e:
        print(f"[Outlook Cal Sub] Exceção: {e}")


# =========================
# WEBHOOKS
# =========================

# Conta de servico que assina o push autenticado do Pub/Sub, e a audience
# configurada na assinatura. Enquanto GMAIL_PUSH_SA estiver vazio a verificacao
# fica desligada -- ligar e um passo de configuracao, nao um deploy que derruba
# um webhook que ja esta funcionando.
GMAIL_PUSH_SA = os.getenv("GMAIL_PUSH_SA", "")
GMAIL_PUSH_AUDIENCE = os.getenv("GMAIL_PUSH_AUDIENCE", "")

_google_jwks = None


def _jwks_google():
    """Cliente das chaves publicas do Google, criado no primeiro uso.

    Construir no import faria a API inteira nao subir por causa de um recurso
    que so o webhook usa -- caro demais para o risco.
    """
    global _google_jwks
    if _google_jwks is None:
        _google_jwks = jwt.PyJWKClient("https://www.googleapis.com/oauth2/v3/certs")
    return _google_jwks


def push_autenticado(request: Request) -> bool:
    """Confere o JWT que o Pub/Sub assina quando a assinatura usa autenticacao.

    Sem isso o webhook aceita qualquer POST de quem souber a URL, e o corpo da
    mensagem manda em qual caixa o codigo vai mexer. Com a verificacao ligada,
    so passa requisicao assinada pela conta de servico esperada.
    """
    if not GMAIL_PUSH_SA:
        return True

    cabecalho = request.headers.get("authorization", "")
    if not cabecalho.lower().startswith("bearer "):
        print("[PUBSUB] recusado: push sem Authorization")
        return False

    token = cabecalho.split(" ", 1)[1]
    try:
        chave = _jwks_google().get_signing_key_from_jwt(token).key
        claims = jwt.decode(
            token,
            chave,
            algorithms=["RS256"],
            audience=GMAIL_PUSH_AUDIENCE or None,
            options={"verify_aud": bool(GMAIL_PUSH_AUDIENCE)},
        )
    except Exception as e:
        print(f"[PUBSUB] recusado: JWT invalido ({e})")
        return False

    if claims.get("email") != GMAIL_PUSH_SA:
        print(f"[PUBSUB] recusado: conta de servico inesperada ({claims.get('email')})")
        return False
    if not claims.get("email_verified", False):
        print("[PUBSUB] recusado: email da conta de servico nao verificado")
        return False
    return True


@app.post("/webhooks/gmail", include_in_schema=False)
async def gmail_webhook(request: Request):
    # 401 de proposito: o Pub/Sub reentrega o que foi recusado, entao uma
    # rejeicao indevida se corrige sozinha quando a config for arrumada.
    if not push_autenticado(request):
        raise HTTPException(401, "Push nao autenticado")

    try:
        body = await request.json()
    except Exception:
        return {"ok": True}

    msg_data = body.get("message", {}).get("data", "")
    if not msg_data:
        return {"ok": True}

    try:
        decoded = json.loads(base64.b64decode(msg_data).decode("utf-8"))
        gmail_addr = decoded.get("emailAddress", "")
        new_hist = int(decoded.get("historyId", 0))
    except Exception as e:
        print(f"[Gmail Webhook] Decode erro: {e}")
        return {"ok": True}

    with engine.begin() as conn:
        garantir_tabela_notificacoes(conn)
        sub = conn.execute(
            text(
                """
            SELECT * FROM email_subscriptions
            WHERE provider='gmail' AND email_address=:addr
        """
            ),
            {"addr": gmail_addr},
        ).fetchone()
        if not sub:
            return {"ok": True}

        sub = dict(sub._mapping)
        old_hist = sub.get("history_id") or new_hist
        access_token = sub.get("access_token", "")
        refresh_token = sub.get("refresh_token", "")
        usuario_email = sub.get("usuario_email", "")

        if new_hist <= old_hist:
            return {"ok": True}

        conn.execute(
            text(
                """
            UPDATE email_subscriptions SET history_id=:hid, atualizado_em=NOW()
            WHERE provider='gmail' AND email_address=:addr
        """
            ),
            {"hid": new_hist, "addr": gmail_addr},
        )

    hist_res = http_requests.get(
        f"https://gmail.googleapis.com/gmail/v1/users/{gmail_addr}/history"
        f"?startHistoryId={old_hist}&historyTypes=messageAdded",
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=15,
    )

    if hist_res.status_code == 401 and refresh_token:
        new_access = await _refresh_google_token(refresh_token, usuario_email)
        if new_access:
            access_token = new_access
            hist_res = http_requests.get(
                f"https://gmail.googleapis.com/gmail/v1/users/{gmail_addr}/history"
                f"?startHistoryId={old_hist}&historyTypes=messageAdded",
                headers={"Authorization": f"Bearer {access_token}"},
                timeout=15,
            )

    print("[GMAIL] HISTORY RAW:")
    print(hist_res.text[:5000])

    if not hist_res.ok:
        return {"ok": True}

    for record in hist_res.json().get("history", []):
        print("[GMAIL] RECORD:", record)

        entries = []

        entries.extend(record.get("messagesAdded", []))

        for msg in record.get("messages", []):
            entries.append({"message": msg})

        for entry in entries:

            msg_id = entry.get("message", {}).get("id")

            if not msg_id:
                continue

            # começa aqui
            msg_res = http_requests.get(
                f"https://gmail.googleapis.com/gmail/v1/users/{gmail_addr}/messages/{msg_id}"
                "?format=full",
                headers={
                    "Authorization":
                        f"Bearer {access_token}"
                },
                timeout=10,
            )

            print(
                "[GMAIL] msg status:",
                msg_res.status_code
            )

            # token expirado → refresh automático
            if msg_res.status_code == 401:

                print(
                    "[GMAIL] token expirado, renovando..."
                )

                refresh_res = http_requests.post(
                    "https://oauth2.googleapis.com/token",
                    data={
                        "client_id":
                            GOOGLE_CLIENT_ID,
                        "client_secret":
                            GOOGLE_CLIENT_SECRET,
                        "refresh_token":
                            refresh_token,
                        "grant_type":
                            "refresh_token",
                    },
                    timeout=10,
                )

                print(
                    "[GMAIL] refresh status:",
                    refresh_res.status_code
                )

                if refresh_res.ok:

                    refresh_json = refresh_res.json()

                    access_token = refresh_json.get(
                        "access_token",
                        access_token
                    )

                    with engine.begin() as conn:
                        conn.execute(
                            text("""
                                UPDATE email_subscriptions
                                SET
                                    access_token = :token,
                                    atualizado_em = NOW()
                                WHERE usuario_email = :uemail
                                AND provider = 'gmail'
                            """),
                            {
                                "token": access_token,
                                "uemail": usuario_email,
                            },
                        )

                    print(
                        "[GMAIL] token renovado"
                    )

                    # retry request
                    msg_res = http_requests.get(
                        f"https://gmail.googleapis.com/gmail/v1/users/{gmail_addr}/messages/{msg_id}"
                        "?format=full",
                        headers={
                            "Authorization":
                                f"Bearer {access_token}"
                        },
                        timeout=10,
                    )

                    print(
                        "[GMAIL] retry status:",
                        msg_res.status_code
                    )

            if not msg_res.ok:
                print(
                    "[GMAIL] erro ao buscar mensagem:",
                    msg_res.text
                )
                continue
            
            
            msg_json = msg_res.json()

            headers_map = {
                h["name"].lower(): h["value"]
                for h in msg_json.get("payload", {}).get("headers", [])
            }

            print("[GMAIL] headers:", headers_map)

            thread_id = msg_json.get("threadId", "")

            from_raw = headers_map.get("from", "")
            subject = headers_map.get("subject", "")
            in_reply = headers_map.get("in-reply-to", "")

            subject_clean = (subject or "").strip().lower()

            reply_prefixes = (
            "re:",
            "res:",
            "aw:",
            "fw:",
            "fwd:",
            "aceito:",
            "accepted:",
            "recusado:",
            "declined:",
            "talvez:",
            "tentative:"
        )

            is_reply = (
                bool(in_reply)
                or subject_clean.startswith(reply_prefixes)
            )

            print("[GMAIL] subject:", subject)
            print("[GMAIL] in_reply:", in_reply)
            print("[GMAIL] is_reply:", is_reply)

            if not is_reply:
                print("[GMAIL] ignorado - não é reply")
                continue

            match = re.match(
                r"^(.*?)\s*<(.+?)>$",
                from_raw.strip()
            )

            sender_name = (
                match.group(1).strip().strip('"')
                if match else ""
            )

            sender_email = (
                match.group(2).strip()
                if match else from_raw.strip()
            )

            print("[GMAIL] sender_email:", sender_email)

            is_calendar_response = subject.lower().startswith(
                (
                    "aceito:",
                    "accepted:",
                    "recusado:",
                    "declined:",
                    "talvez:",
                    "tentative:"
                )
            )

            if (
                sender_email.lower() == gmail_addr.lower()
                and not is_calendar_response
            ):
                print("[GMAIL] ignorado - meu próprio email")
                continue

            if is_automated_sender(sender_email) and not is_calendar_response:
                print("[GMAIL] ignorado - remetente automático")
                continue

            with engine.begin() as conn:
                
                if is_calendar_response:
                    print("[GMAIL] calendar response detectada")

                    calendar_ref = ""
                    for header_value in (in_reply, headers_map.get("references", "")):
                        match_calendar_ref = re.search(r"calendar-([a-f0-9-]+)@google\.com", header_value or "", re.I)
                        if match_calendar_ref:
                            calendar_ref = match_calendar_ref.group(1)
                            break

                    evento = conn.execute(
                        text("""
                            SELECT evento_id, empresa_id, empresa_nome, titulo, google_event_id
                            FROM eventos
                            WHERE usuario_email = :email
                              AND LOWER(COALESCE(email_convidado, '')) = :sender_email
                              AND (
                                  google_event_id = :calendar_ref
                                  OR :calendar_ref = ''
                                  OR google_event_id IS NULL
                              )
                            ORDER BY criado_em DESC
                            LIMIT 1
                        """),
                        {
                            "email": usuario_email,
                            "sender_email": sender_email.lower(),
                            "calendar_ref": calendar_ref,
                        },
                    ).fetchone()

                    if evento:
                        empresa_id   = evento.empresa_id
                        empresa_nome = evento.empresa_nome
                        titulo_evento = evento.titulo or subject
                        evento_id_db  = evento.evento_id

                        # Determina tipo e status a partir do assunto
                        if subject_clean.startswith(("aceito:", "accepted:")):
                            notif_tipo  = "calendar_accepted"
                            verbo       = "aceitou"
                            novo_status = "aceito"
                        elif subject_clean.startswith(("recusado:", "declined:", "recusou:")):
                            notif_tipo  = "calendar_declined"
                            verbo       = "recusou"
                            novo_status = "negado"
                        elif subject_clean.startswith(("talvez:", "tentative:")):
                            notif_tipo  = "calendar_tentative"
                            verbo       = "disse talvez para"
                            novo_status = "talvez"
                        else:
                            notif_tipo  = "calendar_accepted"
                            verbo       = "respondeu"
                            novo_status = "aceito"

                        print(f"[GMAIL] notif tipo={notif_tipo} status={novo_status}")

                        # Evita duplicata nos últimos 5 min
                        existe = conn.execute(
                            text("""
                                SELECT 1 FROM notificacoes
                                WHERE empresa_id = :eid
                                  AND tipo = :tipo
                                  AND meta->>'sender_email' = :semail
                                  AND criado_em >= NOW() - INTERVAL '5 minutes'
                            """),
                            {"eid": str(empresa_id), "tipo": notif_tipo, "semail": sender_email},
                        ).fetchone()

                        if not existe:
                            conn.execute(
                                text("""
                                    INSERT INTO notificacoes
                                        (notificacao_id, usuario_email, tipo, titulo, mensagem,
                                         empresa_id, empresa_nome, platform, meta, lida, criado_em)
                                    VALUES
                                        (:id, :uemail, :tipo, :titulo, :mensagem,
                                         :eid, :enome, 'gmail', CAST(:meta AS JSONB), FALSE, NOW())
                                """),
                                {
                                    "id":      str(uuid.uuid4()),
                                    "uemail":  usuario_email,
                                    "tipo":    notif_tipo,
                                    "titulo":  f"{empresa_nome} {verbo} a call",
                                    "mensagem": f"{sender_name or sender_email} {verbo} o convite para '{titulo_evento}'.",
                                    "eid":     str(empresa_id),
                                    "enome":   empresa_nome,
                                    "meta":    json.dumps({
                                        "sender_email":    sender_email,
                                        "sender_name":     sender_name,
                                        "subject":         subject,
                                        "conversation_id": thread_id,
                                    }),
                                },
                            )

                        # Atualiza status_resposta no evento
                        conn.execute(
                            text("UPDATE eventos SET status_resposta = :status WHERE evento_id = :eid"),
                            {"status": novo_status, "eid": str(evento_id_db)},
                        )
                        print(f"[GMAIL] status_resposta atualizado: {novo_status}")

                    continue

                sender_email = sender_email.strip().lower()

                empresa = conn.execute(
                    text("""
                        SELECT
                            e.empresa_id,
                            e.nome,
                            c.email
                        FROM contatos c
                        JOIN empresas e
                            ON e.empresa_id = c.empresa_id
                        WHERE LOWER(TRIM(c.email)) = LOWER(TRIM(:email))
                        LIMIT 1
                    """),
                    {"email": sender_email},
                ).fetchone()

                print("[GMAIL] procurando email:", repr(sender_email))

                teste_contatos = conn.execute(
                    text("""
                        SELECT email
                        FROM contatos
                        WHERE email IS NOT NULL
                        LIMIT 20
                    """)
                ).fetchall()

                print(
                    "[GMAIL] emails no banco:",
                    [x.email for x in teste_contatos]
                )

                if not empresa:
                    calendar_ref = ""
                    for header_value in (in_reply, headers_map.get("references", "")):
                        match_calendar_ref = re.search(r"calendar-([a-f0-9-]+)@google\.com", header_value or "", re.I)
                        if match_calendar_ref:
                            calendar_ref = match_calendar_ref.group(1)
                            break

                    empresa = conn.execute(
                        text("""
                            SELECT
                                empresa_id,
                                empresa_nome AS nome
                            FROM eventos
                            WHERE usuario_email = :email
                              AND LOWER(COALESCE(email_convidado, '')) = :sender_email
                              AND (
                                  google_event_id = :calendar_ref
                                  OR :calendar_ref = ''
                                  OR google_event_id IS NULL
                              )
                            ORDER BY criado_em DESC
                            LIMIT 1
                        """),
                        {
                            "email": usuario_email,
                            "sender_email": sender_email,
                            "calendar_ref": calendar_ref,
                        },
                    ).fetchone()

                empresa_id = None
                empresa_nome = None

                if empresa:
                    empresa_id = str(empresa.empresa_id)
                    empresa_nome = empresa.nome

                if not empresa_id:

                    evento = conn.execute(
                        text("""
                            SELECT
                                empresa_id,
                                empresa_nome
                            FROM eventos
                            WHERE LOWER(email_convidado) = LOWER(:email)
                            AND google_event_id IS NOT NULL
                            AND conta_id = (SELECT conta_id FROM usuarios WHERE email = :uemail)
                            ORDER BY criado_em DESC
                            LIMIT 1
                        """),
                        {"email": sender_email, "uemail": usuario_email},
                    ).fetchone()

                    if evento:
                        empresa_id = str(evento.empresa_id)
                        empresa_nome = evento.empresa_nome

                        print(
                            "[GMAIL] empresa encontrada por evento:",
                            empresa_nome
                        )

                print(
                    "[GMAIL] empresa encontrada:",
                    empresa_id,
                    empresa_nome
                )

                if empresa_id:
                    print("[GMAIL] criando notificação")

                    create_interaction_notification(
                        conn,
                        usuario_email,
                        empresa_id,
                        empresa_nome,
                        "gmail",
                        sender_name,
                        sender_email,
                        subject,
                        thread_id,
                    )


@app.api_route("/webhooks/outlook", methods=["GET", "POST"], include_in_schema=False)
async def outlook_webhook(request: Request):
    validation_token = request.query_params.get("validationToken")
    if validation_token:
        from fastapi.responses import PlainTextResponse

        return PlainTextResponse(content=validation_token, status_code=200)

    try:
        body = await request.json()
    except Exception:
        return {"ok": True}

    for notif in body.get("value", []):
        if notif.get("clientState") != OUTLOOK_WEBHOOK_SECRET:
            continue
        sub_id = notif.get("subscriptionId")
        msg_id = notif.get("resourceData", {}).get("id")
        if not msg_id:
            continue

        with engine.begin() as conn:
            garantir_tabela_notificacoes(conn)
            sub = conn.execute(
                text(
                    """
                SELECT * FROM email_subscriptions
                WHERE provider='outlook' AND subscription_id=:sid
            """
                ),
                {"sid": sub_id},
            ).fetchone()
            if not sub:
                continue
            sub = dict(sub._mapping)

        access_token = sub.get("access_token", "")
        usuario_email = sub.get("usuario_email", "")
        own_email = sub.get("email_address", "")

        # Busca mensagem com cabeçalhos de resposta para detectar replies reais
        msg_res = http_requests.get(
            f"https://graph.microsoft.com/v1.0/me/messages/{msg_id}"
            "?$select=from,subject,conversationId,internetMessageHeaders",
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=10,
        )
        if not msg_res.ok:
            continue

        msg_data = msg_res.json()
        from_obj = msg_data.get("from", {}).get("emailAddress", {})
        sender_email = from_obj.get("address", "")
        sender_name = from_obj.get("name", "")
        subject = msg_data.get("subject", "")
        conversation_id = msg_data.get("conversationId", "")

        # Verifica se é uma resposta real via cabeçalhos Internet padrão
        internet_headers = {
            h.get("name", "").lower(): h.get("value", "")
            for h in msg_data.get("internetMessageHeaders", [])
        }

        in_reply_to = internet_headers.get("in-reply-to", "")

        subject_clean = (subject or "").strip().lower()

        # Detecção de replies e respostas de calendário.
        # Outlook usa assuntos localizados para aceite/recusa de convite.
        reply_prefixes = (
            "re:",
            "res:",
            "aw:",
            "fw:",
            "fwd:",
        )
        calendar_response_prefixes = (
            "aceito:",
            "accepted:",
            "recusado:",
            "declined:",
            "talvez:",
            "tentative:",
        )

        is_reply = (
            bool(in_reply_to)
            or subject_clean.startswith(reply_prefixes)
        )

        # calendário fica por conta do webhook calendar
        if not is_reply:
            print(
                f"[Outlook Webhook] Ignorado (não é reply): {subject}"
            )
            continue
            print(f"[Outlook Webhook] Ignorado (não é reply/calendário): {subject}")
            continue

        if not sender_email:
            continue

        if sender_email.lower() == own_email.lower():
            continue

        if is_automated_sender(sender_email):
            print(
                f"[Outlook Webhook] Ignorado (remetente automático): {sender_email}"
            )
            continue

        with engine.begin() as conn:

            conta_dono = conn.execute(
                text("SELECT conta_id FROM usuarios WHERE email = :e"),
                {"e": usuario_email},
            ).scalar()
            empresa_id, _, empresa_nome = find_company_by_sender(conn, sender_email, conta_dono)
            if empresa_id:
                create_interaction_notification(
                    conn,
                    usuario_email,
                    empresa_id,
                    empresa_nome,
                    "outlook",
                    sender_name,
                    sender_email,
                    subject,
                    conversation_id,
                )

    return {"ok": True}


@app.api_route("/webhooks/outlook-calendar", methods=["GET", "POST"], include_in_schema=False)
async def outlook_calendar_webhook(request: Request):
    validation_token = request.query_params.get("validationToken")
    if validation_token:
        from fastapi.responses import PlainTextResponse

        return PlainTextResponse(
            content=validation_token,
            status_code=200
        )

    try:
        body = await request.json()
    except Exception:
        return {"ok": True}

    for notif in body.get("value", []):

        print("[OUTLOOK CALENDAR] notif:", notif)

        if notif.get("clientState") != OUTLOOK_WEBHOOK_SECRET:
            print("[OUTLOOK CALENDAR] clientState inválido")
            continue

        sub_id = notif.get("subscriptionId")
        event_id = notif.get("resourceData", {}).get("id")

        print("[OUTLOOK CALENDAR] sub_id:", sub_id)
        print("[OUTLOOK CALENDAR] event_id:", event_id)

        if not event_id:
            print("[OUTLOOK CALENDAR] sem event_id")
            continue

        with engine.begin() as conn:
            garantir_tabela_notificacoes(conn)

            sub = conn.execute(
                text("""
                    SELECT *
                    FROM email_subscriptions
                    WHERE provider='outlook_calendar'
                      AND subscription_id=:sid
                """),
                {"sid": sub_id},
            ).fetchone()

            print("[OUTLOOK CALENDAR] sub:", sub)

            if not sub:
                print("[OUTLOOK CALENDAR] subscription não encontrada")
                continue

            sub = dict(sub._mapping)

        access_token = sub.get("access_token", "")
        usuario_email = sub.get("usuario_email", "")

        event_res = http_requests.get(
            f"https://graph.microsoft.com/v1.0/me/events/{event_id}"
            "?$select=subject,attendees",
            headers={
                "Authorization": f"Bearer {access_token}"
            },
            timeout=10,
        )

        print(
            "[OUTLOOK CALENDAR] event_res status:",
            event_res.status_code
        )

        # token expirado → tenta refresh automático
        if event_res.status_code == 401:

            print(
                "[OUTLOOK CALENDAR] token expirado, renovando..."
            )

            refresh_res = http_requests.post(
                "https://login.microsoftonline.com/common/oauth2/v2.0/token",
                data={
                    "client_id": OUTLOOK_CLIENT_ID,
                    "client_secret": OUTLOOK_CLIENT_SECRET,
                    "grant_type": "refresh_token",
                    "refresh_token": sub.get(
                        "refresh_token",
                        ""
                    ),
                    "scope":
                        "offline_access "
                        "https://graph.microsoft.com/.default",
                },
                timeout=10,
            )

            print(
                "[OUTLOOK CALENDAR] refresh status:",
                refresh_res.status_code
            )

            if refresh_res.ok:

                refresh_json = refresh_res.json()

                access_token = refresh_json.get(
                    "access_token",
                    access_token
                )

                new_refresh_token = refresh_json.get(
                    "refresh_token",
                    sub.get("refresh_token")
                )

                with engine.begin() as conn:
                    conn.execute(
                        text("""
                            UPDATE email_subscriptions
                            SET
                                access_token = :atoken,
                                refresh_token = :rtoken,
                                atualizado_em = NOW()
                            WHERE subscription_id = :sid
                        """),
                        {
                            "atoken": access_token,
                            "rtoken": new_refresh_token,
                            "sid": sub_id,
                        },
                    )

                print(
                    "[OUTLOOK CALENDAR] token renovado"
                )

                # tenta novamente
                event_res = http_requests.get(
                    f"https://graph.microsoft.com/v1.0/me/events/{event_id}"
                    "?$select=subject,attendees",
                    headers={
                        "Authorization":
                            f"Bearer {access_token}"
                    },
                    timeout=10,
                )

                print(
                    "[OUTLOOK CALENDAR] retry status:",
                    event_res.status_code
                )

        if not event_res.ok:
            print(
                "[OUTLOOK CALENDAR] erro graph:",
                event_res.text
            )
            continue

        event_data = event_res.json()



        print("[OUTLOOK CALENDAR] event_data:")
        print(json.dumps(event_data, indent=2))

        subject = event_data.get("subject", "")
        attendees = event_data.get("attendees", [])

        with engine.begin() as conn:

            evento = conn.execute(
                text("""
                    SELECT
                        evento_id,
                        empresa_id,
                        empresa_nome,
                        titulo,
                        outlook_event_id
                    FROM eventos
                    WHERE usuario_email = :uemail
                      AND outlook_event_id = :event_id
                    ORDER BY criado_em DESC
                    LIMIT 1
                """),
                {
                    "uemail":   usuario_email,
                    "event_id": event_id,
                },
            ).fetchone()

            print(
                "[OUTLOOK CALENDAR] evento encontrado:",
                evento
            )

            if not evento:
                print(
                    "[OUTLOOK CALENDAR] nenhum evento encontrado"
                )
                continue

            evento = dict(evento._mapping)

            empresa_id = evento.get("empresa_id")
            empresa_nome = evento.get("empresa_nome")
            titulo_evento = (
                evento.get("titulo")
                or subject
            )

            print(
                "[OUTLOOK CALENDAR] empresa:",
                empresa_nome
            )

            if not empresa_id or not empresa_nome:
                print(
                    "[OUTLOOK CALENDAR] empresa inválida"
                )
                continue

            response_map = {
                "accepted": (
                    "calendar_accepted",
                    "aceitou"
                ),
                "declined": (
                    "calendar_declined",
                    "recusou"
                ),
                "tentativelyAccepted": (
                    "calendar_tentative",
                    "disse talvez para"
                ),
            }

            for attendee in attendees:

                response = attendee.get(
                    "status",
                    {}
                ).get("response", "")

                email_addr = attendee.get(
                    "emailAddress",
                    {}
                ).get("address", "")

                name = attendee.get(
                    "emailAddress",
                    {}
                ).get("name", email_addr)

                print(
                    "[OUTLOOK CALENDAR] attendee:",
                    name,
                    email_addr,
                    response
                )

                if response not in response_map:
                    print(
                        "[OUTLOOK CALENDAR] response ignorada:",
                        response
                    )
                    continue

                notif_tipo, verbo = response_map[
                    response
                ]

                existe = conn.execute(
                    text("""
                        SELECT 1
                        FROM notificacoes
                        WHERE empresa_id = :eid
                          AND tipo = :tipo
                          AND meta->>'attendee_email' = :aemail
                          AND criado_em >= NOW()
                          - INTERVAL '5 minutes'
                    """),
                    {
                        "eid": str(empresa_id),
                        "tipo": notif_tipo,
                        "aemail": email_addr,
                    },
                ).fetchone()

                if existe:
                    print(
                        "[OUTLOOK CALENDAR] já existe"
                    )
                    continue

                print(
                    "[OUTLOOK CALENDAR] criando notificação"
                )

                conn.execute(
                    text("""
                        INSERT INTO notificacoes
                            (
                                notificacao_id,
                                usuario_email,
                                tipo,
                                titulo,
                                mensagem,
                                empresa_id,
                                empresa_nome,
                                platform,
                                meta,
                                lida,
                                criado_em
                            )
                        VALUES
                            (
                                :id,
                                :uemail,
                                :tipo,
                                :titulo,
                                :mensagem,
                                :eid,
                                :enome,
                                'outlook',
                                CAST(:meta AS JSONB),
                                FALSE,
                                NOW()
                            )
                    """),
                    {
                        "id": str(uuid.uuid4()),
                        "uemail": usuario_email,
                        "tipo": notif_tipo,
                        "titulo":
                            f"{empresa_nome} {verbo} a call",
                        "mensagem":
                            f"{name} {verbo} "
                            f"o convite para "
                            f"'{titulo_evento}'.",
                        "eid": str(empresa_id),
                        "enome": empresa_nome,
                        "meta": json.dumps(
                            {
                                "attendee_email":
                                    email_addr,
                                "attendee_name":
                                    name,
                                "outlook_event_id":
                                    event_id,
                                "event_subject":
                                    subject,
                            }
                        ),
                    },
                )

                # Atualiza status_resposta no evento
                status_map = {
                    "accepted":           "aceito",
                    "declined":           "negado",
                    "tentativelyAccepted":"talvez",
                }
                novo_status = status_map.get(response)
                if novo_status:
                    conn.execute(
                        text("""
                            UPDATE eventos
                            SET status_resposta = :status
                            WHERE evento_id = :eid
                        """),
                        {
                            "status": novo_status,
                            "eid":    str(evento.get("evento_id")),
                        },
                    )
                    print(f"[OUTLOOK CALENDAR] status_resposta atualizado: {novo_status}")

    return {"ok": True}


# =========================
# RENOVAÇÃO DE SUBSCRIPTIONS
# =========================
def renovar_gmail_watches():
    """Renova o watch do Gmail antes de expirar, e abre para quem ainda nao tem.

    Dois detalhes que a versao anterior errava:

    - O token vinha da propria assinatura, gravado quando o watch foi criado.
      Como dura 1 hora, toda renovacao posterior batia em 401. Agora vem de
      usuarios, que e onde o refresh mantem o token atual.
    - O SELECT so via quem ja tinha assinatura, entao quem conectou o Google
      antes de o watch existir -- ou perdeu a assinatura por qualquer motivo --
      nunca mais ganhava uma. O LEFT JOIN cobre esse caso.
    """
    with engine.begin() as conn:
        garantir_tabela_notificacoes(conn)
        subs = conn.execute(
            text(
                """
            SELECT u.email AS usuario_email,
                   u.google_access_token, u.google_refresh_token,
                   s.email_address
            FROM usuarios u
            LEFT JOIN email_subscriptions s
              ON s.usuario_email = u.email AND s.provider = 'gmail'
            WHERE u.google_access_token IS NOT NULL
              AND (s.sub_id IS NULL
                   OR s.expires_at IS NULL
                   OR s.expires_at <= NOW() + INTERVAL '36 hours')
        """
            )
        ).fetchall()
    for sub in subs:
        s = dict(sub._mapping)
        setup_gmail_watch(
            s["usuario_email"],
            s.get("google_access_token") or "",
            s.get("google_refresh_token") or "",
            s.get("email_address") or "",
        )


def renovar_outlook_subscriptions():
    """Renova (ou recria) as assinaturas do Graph antes de expirarem.

    Tres defeitos na versao anterior, e o terceiro era o pior:

    - o token vinha da propria assinatura, gravado quando ela foi criada. Token
      da Microsoft dura cerca de 1 hora, entao todo PATCH batia em 401;
    - o resultado do PATCH nunca era conferido;
    - e o expires_at era gravado **de qualquer jeito**. O banco passava a
      afirmar que a assinatura valia por mais ~3 dias enquanto o Graph ja a
      tinha deixado expirar. Mentira que se auto-renovava a cada ciclo e
      escondia a falha para sempre -- inclusive do painel de integracoes.

    Assinatura vencida o Graph nao ressuscita: quando o PATCH falha, cria outra.
    """
    with engine.connect() as conn:
        usuarios = conn.execute(
            text("""
                SELECT email, outlook_access_token, outlook_refresh_token
                FROM usuarios WHERE outlook_access_token IS NOT NULL
            """)
        ).fetchall()
        existentes = conn.execute(
            text("""
                SELECT usuario_email, provider, subscription_id, expires_at
                FROM email_subscriptions
                WHERE provider IN ('outlook', 'outlook_calendar')
            """)
        ).fetchall()

    por_usuario: dict = {}
    for e in existentes:
        por_usuario.setdefault(e.usuario_email, {})[e.provider] = e

    limite = datetime.utcnow() + timedelta(hours=12)
    criadores = {
        "outlook": setup_outlook_subscription,
        "outlook_calendar": setup_outlook_calendar_subscription,
    }

    for u in usuarios:
        atuais = por_usuario.get(u.email, {})
        token = u.outlook_access_token or ""
        refresh = u.outlook_refresh_token or ""

        for provider, criar in criadores.items():
            atual = atuais.get(provider)

            if atual and atual.subscription_id and atual.expires_at and atual.expires_at > limite:
                # O banco pode estar mentindo: a versao antiga gravava validade
                # nova mesmo quando o PATCH falhava. Uma linha dessas seria
                # pulada aqui para sempre, entao confirma no Graph antes.
                sid_atual = atual.subscription_id
                token, resp = outlook_com_retry(
                    u.email, token, refresh,
                    lambda t: http_requests.get(
                        f"https://graph.microsoft.com/v1.0/subscriptions/{sid_atual}",
                        headers={"Authorization": f"Bearer {t}"},
                        timeout=15,
                    ),
                )
                if resp.ok:
                    continue
                print(f"[Outlook Sub] {provider} nao existe mais no Graph ({resp.status_code}) para {u.email}; recriando")
                criar(u.email, token, refresh)
                continue

            renovada = False
            if atual and atual.subscription_id:
                nova_exp = datetime.utcnow() + timedelta(minutes=4000)
                sid = atual.subscription_id
                token, resp = outlook_com_retry(
                    u.email, token, refresh,
                    lambda t: http_requests.patch(
                        f"https://graph.microsoft.com/v1.0/subscriptions/{sid}",
                        headers={"Authorization": f"Bearer {t}", "Content-Type": "application/json"},
                        json={"expirationDateTime": nova_exp.strftime("%Y-%m-%dT%H:%M:%S.0000000Z")},
                        timeout=15,
                    ),
                )
                renovada = resp.ok
                if renovada:
                    with engine.begin() as conn:
                        conn.execute(
                            text("""
                                UPDATE email_subscriptions
                                SET expires_at = :exp, atualizado_em = NOW()
                                WHERE subscription_id = :sid
                            """),
                            {"exp": nova_exp, "sid": sid},
                        )
                    print(f"[Outlook Sub] {provider} renovada para {u.email}")
                else:
                    print(f"[Outlook Sub] PATCH {provider} falhou ({resp.status_code}) para {u.email}: {resp.text}")

            if not renovada:
                criar(u.email, token, refresh)


# =========================
# RANKING MENSAL
# =========================
def gerar_ranking_mensal():
    print("📊 Gerando ranking mensal do Places...")
    mes_atual = datetime.utcnow().strftime("%Y-%m")
    try:
        with engine.begin() as conn:
            garantir_tabelas_places_cache(conn)

            # Snapshot de popularidade — APENAS analítico (termo + nº de buscas).
            # Não copiamos conteúdo de places para cá, então nada aqui vence o
            # limite de cache de 30 dias dos Termos do Google.
            top10 = conn.execute(text("""
                SELECT query, lat_grid, lng_grid, search_count
                FROM places_cache
                ORDER BY search_count DESC
                LIMIT 10
            """)).fetchall()

            conn.execute(text("DELETE FROM places_ranking"))
            for i, row in enumerate(top10, 1):
                conn.execute(text("""
                    INSERT INTO places_ranking
                        (id, query, lat_grid, lng_grid, results, search_count, rank_position, month, saved_date)
                    VALUES (:id, :q, :lat, :lng, '[]'::jsonb, :count, :pos, :month, NOW())
                """), {
                    "id": str(uuid.uuid4()),
                    "q": row.query,
                    "lat": row.lat_grid,
                    "lng": row.lng_grid,
                    "count": row.search_count,
                    "pos": i,
                    "month": mes_atual,
                })

            # Limpeza compatível com os Termos: remove só o conteúdo já expirado
            # (> 30 dias). O cache ainda fresco continua válido e evita rebuscas
            # pagas no Google.
            conn.execute(text(
                "DELETE FROM places_cache WHERE updated_at < NOW() - INTERVAL '30 days'"
            ))

        print(f"✅ Ranking mensal gerado: {mes_atual} — {len(top10)} termos.")
    except Exception as e:
        print(f"🔴 Erro ao gerar ranking: {e}")


# =========================
# JOB: RETENÇÃO LGPD
# =========================
# Meses sem interação após os quais um lead vira candidato a anonimização/exclusão.
LGPD_RETENCAO_MESES = int(os.getenv("LGPD_RETENCAO_MESES", "18"))
# Por segurança, o job só REPORTA por padrão. Exclusão automática exige opt-in
# explícito (LGPD_RETENCAO_AUTO=true), pois apaga dados de produção.
LGPD_RETENCAO_AUTO = os.getenv("LGPD_RETENCAO_AUTO", "false").lower() == "true"


def retencao_lgpd():
    """Identifica leads sem interação há mais de LGPD_RETENCAO_MESES meses.
    Registra em auditoria; só exclui se LGPD_RETENCAO_AUTO estiver ligado."""
    print("⏰ JOB: revisão de retenção LGPD...")
    try:
        with engine.begin() as conn:
            garantir_seguranca(conn)
            limite = datetime.utcnow() - timedelta(days=LGPD_RETENCAO_MESES * 30)
            candidatos = conn.execute(
                text(
                    """
                    SELECT empresa_id, conta_id, nome
                    FROM empresas
                    WHERE COALESCE(ultima_interacao, status_atualizado_em, data_criacao) < :limite
                      AND status NOT IN ('Cliente', 'Fechado', 'Ganho')
                    """
                ),
                {"limite": limite},
            ).fetchall()
            if not candidatos:
                print("✅ Retenção LGPD: nenhum lead elegível.")
                return
            print(f"📋 Retenção LGPD: {len(candidatos)} lead(s) sem interação há >{LGPD_RETENCAO_MESES} meses.")
            registrar_auditoria(
                acao="RETENCAO_LGPD_REVISAO", recurso="empresas",
                quantidade=len(candidatos),
                meta={"auto": LGPD_RETENCAO_AUTO, "meses": LGPD_RETENCAO_MESES},
                conn=conn,
            )
            if LGPD_RETENCAO_AUTO:
                ids = [str(c.empresa_id) for c in candidatos]
                conn.execute(
                    text("DELETE FROM empresas WHERE empresa_id = ANY(:ids)"),
                    {"ids": ids},
                )
                registrar_auditoria(
                    acao="RETENCAO_LGPD_EXCLUSAO", recurso="empresas",
                    quantidade=len(ids), conn=conn,
                )
                print(f"🗑️ Retenção LGPD: {len(ids)} lead(s) excluído(s) (auto).")
    except Exception as e:
        print(f"🔴 Erro no job de retenção LGPD: {e}")


# =========================
# RESPOSTAS DE CONVITE — GOOGLE
# =========================
def verificar_respostas_google(usuario_email: Optional[str] = None):
    """Lê a resposta do convidado direto do Google Calendar.

    Com `usuario_email`, checa só os eventos daquele usuário — é assim que o
    webhook de push responde a uma mudança sem varrer a base inteira.

    O caminho antigo era indireto: esperava o e-mail "Aceito: ..." que o Google
    manda ao organizador e deduzia a resposta pelo assunto (ver o webhook do
    Gmail, mais acima). Esse e-mail só é enviado se a conta tiver "Respostas a
    eventos" ligado nas notificações do Calendar — o padrão é não enviar, e sem
    ele a resposta nunca chegava ao CRM.

    O responseStatus do próprio evento é autoritativo e está sempre lá. É o
    mesmo caminho que o Outlook já usa via Graph em outlook_calendar_webhook.
    """
    status_map = {
        "accepted":  ("calendar_accepted",  "aceitou",           "aceito"),
        "declined":  ("calendar_declined",  "recusou",           "negado"),
        "tentative": ("calendar_tentative", "disse talvez para", "talvez"),
    }

    with engine.connect() as conn:
        pendentes = conn.execute(
            text("""
            SELECT e.evento_id, e.usuario_email, e.empresa_id, e.empresa_nome,
                   e.titulo, e.google_event_id,
                   u.google_access_token, u.google_refresh_token
            FROM eventos e
            JOIN usuarios u ON u.email = e.usuario_email
            WHERE e.google_event_id IS NOT NULL
              AND COALESCE(e.status_resposta, 'pendente') = 'pendente'
              AND u.google_access_token IS NOT NULL
              AND e.data >= CURRENT_DATE - INTERVAL '1 day'
              AND (CAST(:uemail AS TEXT) IS NULL OR e.usuario_email = :uemail)
        """),
            {"uemail": usuario_email},
        ).fetchall()

    for row in pendentes:
        ev = dict(row._mapping)
        access_token = ev.get("google_access_token")

        def _buscar(token):
            return http_requests.get(
                "https://www.googleapis.com/calendar/v3/calendars/primary/events/"
                f"{ev['google_event_id']}",
                headers={"Authorization": f"Bearer {token}"},
                timeout=15,
            )

        try:
            resp = _buscar(access_token)
            # 401 aqui é token vencido, não evento inexistente: renova e repete.
            if resp.status_code == 401 and ev.get("google_refresh_token"):
                access_token = asyncio.run(
                    _refresh_google_token(ev["google_refresh_token"], ev["usuario_email"])
                )
                if not access_token:
                    continue
                resp = _buscar(access_token)
            if resp.status_code != 200:
                print(f"[GOOGLE RSVP] evento {ev['google_event_id']}: {resp.status_code}")
                continue
            attendees = resp.json().get("attendees", []) or []
        except Exception as e:
            print(f"[GOOGLE RSVP] erro ao consultar evento: {e}")
            continue

        for att in attendees:
            # O organizador aparece na lista como convidado de si mesmo.
            if att.get("self") or att.get("organizer"):
                continue
            mapped = status_map.get(att.get("responseStatus", ""))
            if not mapped:
                continue  # needsAction: ainda não respondeu
            notif_tipo, verbo, novo_status = mapped

            email_conv = att.get("email", "")
            nome_conv  = att.get("displayName") or email_conv
            titulo_evento = ev.get("titulo") or "reunião"

            with engine.begin() as conn:
                existe = conn.execute(
                    text("""
                        SELECT 1 FROM notificacoes
                        WHERE usuario_email = :uemail
                          AND tipo = :tipo
                          AND meta->>'google_event_id' = :gid
                          AND meta->>'attendee_email' = :aemail
                    """),
                    {
                        "uemail": ev["usuario_email"],
                        "tipo":   notif_tipo,
                        "gid":    ev["google_event_id"],
                        "aemail": email_conv,
                    },
                ).fetchone()

                if not existe and ev.get("empresa_id") and ev.get("empresa_nome"):
                    conn.execute(
                        text("""
                            INSERT INTO notificacoes
                                (notificacao_id, usuario_email, tipo, titulo, mensagem,
                                 empresa_id, empresa_nome, platform, meta, lida, criado_em)
                            VALUES
                                (:id, :uemail, :tipo, :titulo, :mensagem,
                                 :eid, :enome, 'google', CAST(:meta AS JSONB), FALSE, NOW())
                        """),
                        {
                            "id":      str(uuid.uuid4()),
                            "uemail":  ev["usuario_email"],
                            "tipo":    notif_tipo,
                            "titulo":  f"{ev['empresa_nome']} {verbo} a call",
                            "mensagem": f"{nome_conv} {verbo} o convite para '{titulo_evento}'.",
                            "eid":     str(ev["empresa_id"]),
                            "enome":   ev["empresa_nome"],
                            "meta":    json.dumps({
                                "sender_email":    email_conv,
                                "sender_name":     nome_conv,
                                "attendee_email":  email_conv,
                                "attendee_name":   nome_conv,
                                "google_event_id": ev["google_event_id"],
                                "event_subject":   titulo_evento,
                            }),
                        },
                    )

                conn.execute(
                    text("UPDATE eventos SET status_resposta = :status WHERE evento_id = :eid"),
                    {"status": novo_status, "eid": str(ev["evento_id"])},
                )
                print(f"[GOOGLE RSVP] {email_conv} -> {novo_status} (evento {ev['evento_id']})")
            break  # o convite do CRM tem um convidado; o primeiro que respondeu decide


# =========================
# PUSH DO GOOGLE CALENDAR
# =========================
GOOGLE_CAL_WATCH_URL = (
    "https://www.googleapis.com/calendar/v3/calendars/primary/events/watch"
)
GOOGLE_CAL_STOP_URL = "https://www.googleapis.com/calendar/v3/channels/stop"


def setup_google_calendar_watch(usuario_email: str, access_token: str, refresh_token: str):
    """Registra um canal de push do Google Calendar para o usuario.

    O Google avisa o webhook a cada mudanca na agenda -- inclusive quando um
    convidado responde ao convite. Sem isso a resposta so aparecia no ciclo
    seguinte do poller. O canal expira sozinho (a data vem em `expiration`),
    e renovar_google_calendar_watches o recria antes do prazo.

    O endereco so precisa ser HTTPS com certificado valido; nao exige dominio
    verificado.
    """
    # engine.begin(), nao connect(): a coluna resource_id nasce aqui, num ALTER
    # dentro de garantir_tabela_notificacoes, e DDL precisa de commit. Ler a
    # coluna antes de garantir que ela existe quebrava o job inteiro em banco
    # que ainda nao tinha passado por essa migracao.
    with engine.begin() as conn:
        garantir_tabela_notificacoes(conn)
        antigo = conn.execute(
            text("""
                SELECT subscription_id, resource_id FROM email_subscriptions
                WHERE usuario_email = :email AND provider = 'google_calendar'
            """),
            {"email": usuario_email},
        ).fetchone()

    def _com_token(token, fn):
        return google_com_retry(usuario_email, token, refresh_token, fn)

    # Encerra o canal anterior: sem isso os dois ficam vivos durante a
    # sobreposicao e o webhook recebe a mesma mudanca duas vezes.
    if antigo and antigo.subscription_id and antigo.resource_id:
        try:
            _com_token(access_token, lambda t: http_requests.post(
                GOOGLE_CAL_STOP_URL,
                json={"id": antigo.subscription_id, "resourceId": antigo.resource_id},
                headers={"Authorization": f"Bearer {t}"},
                timeout=15,
            ))
        except Exception as e:
            print(f"[GOOGLE CALENDAR] falha ao encerrar canal antigo: {e}")

    canal_id = str(uuid.uuid4())
    try:
        access_token, res = _com_token(access_token, lambda t: http_requests.post(
            GOOGLE_CAL_WATCH_URL,
            json={
                "id": canal_id,
                "type": "web_hook",
                "address": f"{BACKEND_URL}/webhooks/google-calendar",
            },
            headers={"Authorization": f"Bearer {t}"},
            timeout=20,
        ))
    except Exception as e:
        print(f"[GOOGLE CALENDAR] erro ao abrir canal: {e}")
        return

    if res.status_code not in (200, 201):
        print(f"[GOOGLE CALENDAR] watch falhou ({res.status_code}): {res.text}")
        return

    data = res.json()
    exp_ms = data.get("expiration")
    expires_at = datetime.utcfromtimestamp(int(exp_ms) / 1000) if exp_ms else None

    with engine.begin() as conn:
        garantir_tabela_notificacoes(conn)
        if antigo:
            conn.execute(
                text("""
                    UPDATE email_subscriptions
                    SET subscription_id = :cid, resource_id = :rid, expires_at = :exp,
                        access_token = :at, refresh_token = :rt, atualizado_em = NOW()
                    WHERE usuario_email = :email AND provider = 'google_calendar'
                """),
                {"cid": canal_id, "rid": data.get("resourceId"), "exp": expires_at,
                 "at": access_token, "rt": refresh_token, "email": usuario_email},
            )
        else:
            conn.execute(
                text("""
                    INSERT INTO email_subscriptions
                        (sub_id, usuario_email, provider, subscription_id, resource_id,
                         email_address, expires_at, access_token, refresh_token)
                    VALUES
                        (:sid, :email, 'google_calendar', :cid, :rid,
                         :email, :exp, :at, :rt)
                """),
                {"sid": str(uuid.uuid4()), "email": usuario_email, "cid": canal_id,
                 "rid": data.get("resourceId"), "exp": expires_at,
                 "at": access_token, "rt": refresh_token},
            )
    print(f"[GOOGLE CALENDAR] canal OK para {usuario_email}, expira {expires_at}")


def renovar_google_calendar_watches():
    """Abre canal para quem ainda nao tem e renova os que estao perto de expirar.

    O LEFT JOIN cobre quem conectou o Google antes de o push existir: essa
    turma nunca passaria pelo callback do OAuth de novo.
    """
    with engine.connect() as conn:
        pendentes = conn.execute(
            text("""
            SELECT u.email AS usuario_email, u.google_access_token, u.google_refresh_token
            FROM usuarios u
            LEFT JOIN email_subscriptions s
              ON s.usuario_email = u.email AND s.provider = 'google_calendar'
            WHERE u.google_access_token IS NOT NULL
              AND (s.sub_id IS NULL
                   OR s.expires_at IS NULL
                   OR s.expires_at <= NOW() + INTERVAL '24 hours')
        """)
        ).fetchall()
    for row in pendentes:
        u = dict(row._mapping)
        setup_google_calendar_watch(
            u["usuario_email"],
            u.get("google_access_token") or "",
            u.get("google_refresh_token") or "",
        )


@app.post("/webhooks/google-calendar", include_in_schema=False)
def google_calendar_webhook(request: Request):
    """Ping do Google avisando que a agenda mudou.

    O corpo vem vazio de proposito -- a notificacao nao diz o que mudou, so que
    mudou. Quem le a mudanca e verificar_respostas_google, recortado para o
    usuario do canal. Rota sincrona porque essa checagem e bloqueante: assim o
    FastAPI a joga no threadpool em vez de travar o event loop.
    """
    canal_id       = request.headers.get("x-goog-channel-id", "")
    resource_id    = request.headers.get("x-goog-resource-id", "")
    resource_state = request.headers.get("x-goog-resource-state", "")

    # A primeira mensagem de todo canal e so o handshake.
    if resource_state == "sync":
        print(f"[GOOGLE CALENDAR] handshake do canal {canal_id}")
        return {"ok": True}

    with engine.connect() as conn:
        sub = conn.execute(
            text("""
                SELECT usuario_email FROM email_subscriptions
                WHERE provider = 'google_calendar'
                  AND subscription_id = :cid
                  AND (resource_id IS NULL OR resource_id = :rid)
            """),
            {"cid": canal_id, "rid": resource_id},
        ).fetchone()

    if not sub:
        print(f"[GOOGLE CALENDAR] canal desconhecido: {canal_id}")
        return {"ok": True}

    verificar_respostas_google(usuario_email=sub.usuario_email)
    return {"ok": True}


# =========================
# SCHEDULER
# =========================
scheduler = BackgroundScheduler()
scheduler.add_job(verificar_rascunhos_expirados, "cron", hour=8, minute=0)
scheduler.add_job(
    renovar_gmail_watches, "interval", hours=6, id="renew_gmail",
    next_run_time=datetime.now() + timedelta(minutes=2),
)
scheduler.add_job(
    renovar_outlook_subscriptions, "interval", hours=6, id="renew_outlook",
    next_run_time=datetime.now() + timedelta(minutes=3),
)
scheduler.add_job(verificar_respostas_google, "interval", minutes=5, id="rsvp_google")
# Rede de seguranca do push: abre canal para quem ainda nao tem e renova os
# que vao expirar. O next_run_time cobre quem ja estava com o Google conectado.
scheduler.add_job(
    renovar_google_calendar_watches, "interval", hours=6, id="renew_gcal",
    next_run_time=datetime.now() + timedelta(minutes=1),
)
scheduler.add_job(gerar_ranking_mensal, "cron", day="last", hour=23, minute=30)
scheduler.add_job(retencao_lgpd, "cron", hour=4, minute=0, id="retencao_lgpd")
scheduler.start()
print("⏰ Scheduler iniciado — verificação diária às 8h UTC")


# =========================
# ROTAS BÁSICAS
# =========================
@app.get("/")
def home():
    return {"msg": "API rodando 🚀"}


@app.post("/admin/verificar-rascunhos")
def trigger_verificar_rascunhos(email: str = Depends(get_current_user)):
    """Dispara a verificação de rascunhos expirados.

    Passou a exigir token: estava aberta para a internet inteira e nada no
    frontend a chamava.
    """
    verificar_rascunhos_expirados()
    return {"msg": "Verificação executada"}


@app.post("/admin/verificar-respostas-google")
def trigger_verificar_respostas_google(email: str = Depends(get_current_user)):
    """Roda agora a checagem que o scheduler faz a cada 2 min.

    Serve para conferir uma resposta recém-enviada sem esperar o próximo ciclo.
    Diferente de /admin/verificar-rascunhos, exige autenticação."""
    verificar_respostas_google()
    return {"msg": "Verificação executada"}


@app.get("/admin/google-calendar-watch")
def status_google_calendar_watch(email: str = Depends(get_current_user)):
    """Diz se o canal de push existe e até quando vale.

    Sem isso, só o log do Railway responde essa pergunta — e foi essa cegueira
    que deixou passar o job quebrando em silêncio. Não devolve o id do canal.
    """
    with engine.begin() as conn:
        garantir_tabela_notificacoes(conn)
        row = conn.execute(
            text("""
                SELECT subscription_id, resource_id, expires_at, atualizado_em
                FROM email_subscriptions
                WHERE usuario_email = :email AND provider = 'google_calendar'
            """),
            {"email": email},
        ).fetchone()
    if not row or not row.subscription_id:
        return {"canal": False, "msg": "Nenhum canal de push registrado."}
    return {
        "canal": True,
        "expira_em": row.expires_at.isoformat() if row.expires_at else None,
        "atualizado_em": row.atualizado_em.isoformat() if row.atualizado_em else None,
    }


@app.get("/admin/integracoes-equipe")
def status_integracoes_equipe(auth: dict = Depends(exigir_gerente)):
    """Estado dos canais de cada pessoa da conta.

    O painel individual so responde pela propria caixa. Um vendedor cujo watch
    morreu fica semanas sem receber notificacao e ninguem descobre -- foi
    exatamente esse o modo de falha desta rodada, so que na caixa do dono.

    Restrito ao gerente: e a unica visao que atravessa usuarios.
    """
    with engine.begin() as conn:
        garantir_tabela_notificacoes(conn)
        linhas = conn.execute(
            text("""
                SELECT u.nome, u.email, u.role,
                       (u.google_access_token IS NOT NULL) AS google_conectado,
                       s.provider, s.subscription_id, s.email_address,
                       s.expires_at, s.atualizado_em
                FROM usuarios u
                LEFT JOIN email_subscriptions s ON s.usuario_email = u.email
                WHERE u.conta_id = :cid
                ORDER BY u.nome
            """),
            {"cid": auth["conta_id"]},
        ).fetchall()

    pessoas: dict = {}
    for r in linhas:
        pessoa = pessoas.setdefault(
            r.email,
            {
                "nome": r.nome,
                "email": r.email,
                "funcao": ROTULO_FUNCAO[normalizar_role(r.role)],
                "google_conectado": bool(r.google_conectado),
                "canais": {},
            },
        )
        if r.provider:
            pessoa["canais"][r.provider] = {
                "ativo": bool(r.subscription_id),
                "caixa": r.email_address,
                "expira_em": r.expires_at.isoformat() if r.expires_at else None,
                "atualizado_em": r.atualizado_em.isoformat() if r.atualizado_em else None,
            }

    return {"pessoas": list(pessoas.values())}


@app.get("/admin/integracoes-status")
def status_integracoes(email: str = Depends(get_current_user)):
    """Estado dos canais que alimentam as notificacoes deste usuario.

    Existe pela mesma razao do endpoint acima: watch que falha nao aparece em
    lugar nenhum da aplicacao -- so no log. Sem isso, "a notificacao nao chega"
    e indistinguivel de "o canal nunca subiu".
    """
    with engine.begin() as conn:
        garantir_tabela_notificacoes(conn)
        rows = conn.execute(
            text("""
                SELECT provider, subscription_id, email_address,
                       expires_at, atualizado_em
                FROM email_subscriptions
                WHERE usuario_email = :email
            """),
            {"email": email},
        ).fetchall()

    canais = {
        r.provider: {
            "ativo": bool(r.subscription_id),
            "caixa": r.email_address,
            "expira_em": r.expires_at.isoformat() if r.expires_at else None,
            "atualizado_em": r.atualizado_em.isoformat() if r.atualizado_em else None,
        }
        for r in rows
    }
    for provedor in ("gmail", "outlook", "google_calendar"):
        canais.setdefault(provedor, {"ativo": False})

    return {
        "canais": canais,
        # O default de GMAIL_PUBSUB_TOPIC e um placeholder. Se ele sobreviveu,
        # o watch do Gmail falha e resposta de cliente nunca vira notificacao.
        "gmail_pubsub_configurado": "SEU_PROJECT_ID" not in GMAIL_PUBSUB_TOPIC,
        "remetente_sandbox": REMETENTE_EMAIL == "onboarding@resend.dev",
    }


# =========================
# NOTIFICAÇÕES
# =========================
@app.get("/notificacoes")
def listar_notificacoes(empresa_id: Optional[str] = None, email: str = Depends(get_current_user)):
    with engine.begin() as conn:
        garantir_tabela_notificacoes(conn)
        if empresa_id:
            result = conn.execute(
                text(
                    """
                SELECT * FROM notificacoes
                WHERE usuario_email = :email AND empresa_id = :eid
                ORDER BY criado_em DESC LIMIT 50
            """
                ),
                {"email": email, "eid": empresa_id},
            )
        else:
            result = conn.execute(
                text(
                    """
                SELECT * FROM notificacoes
                WHERE usuario_email = :email
                ORDER BY criado_em DESC LIMIT 50
            """
                ),
                {"email": email},
            )
        return [dict(row._mapping) for row in result]


@app.get("/notificacoes/nao-lidas")
def contar_nao_lidas(email: str = Depends(get_current_user)):
    with engine.begin() as conn:
        garantir_tabela_notificacoes(conn)
        result = conn.execute(
            text(
                """
            SELECT COUNT(*) as total FROM notificacoes
            WHERE usuario_email = :email AND lida = FALSE
        """
            ),
            {"email": email},
        ).fetchone()
        return {"total": result._mapping["total"]}


@app.put("/notificacoes/{notificacao_id}/ler")
def marcar_lida(notificacao_id: str, email: str = Depends(get_current_user)):
    with engine.begin() as conn:
        conn.execute(
            text(
                """
            UPDATE notificacoes SET lida = TRUE
            WHERE notificacao_id = :id AND usuario_email = :email
        """
            ),
            {"id": notificacao_id, "email": email},
        )
    return {"msg": "Notificação marcada como lida"}


@app.put("/notificacoes/ler-todas")
def marcar_todas_lidas(email: str = Depends(get_current_user)):
    with engine.begin() as conn:
        conn.execute(
            text(
                """
            UPDATE notificacoes SET lida = TRUE
            WHERE usuario_email = :email AND lida = FALSE
        """
            ),
            {"email": email},
        )
    return {"msg": "Todas as notificações marcadas como lidas"}


@app.delete("/notificacoes/{notificacao_id}")
def deletar_notificacao(notificacao_id: str, email: str = Depends(get_current_user)):
    with engine.begin() as conn:
        conn.execute(
            text(
                """
            DELETE FROM notificacoes
            WHERE notificacao_id = :id AND usuario_email = :email
        """
            ),
            {"id": notificacao_id, "email": email},
        )
    return {"msg": "Notificação removida"}


# =========================
# MEU PERFIL
# =========================
@app.get("/me")
def get_me(auth: dict = Depends(get_auth)):
    with engine.connect() as conn:
        usuario = conn.execute(
            text(
                """
                SELECT u.usuario_id, u.nome, u.email, u.telefone, u.cargo, u.empresa_nome, u.bio,
                       u.data_criacao, u.role, u.conta_id, u.supervisor_id, ct.nome AS conta_nome,
                       ct.plano AS plano,
                       s.nome AS supervisor_nome,
                       COALESCE(u.mfa_ativado, FALSE) AS mfa_ativado
                FROM usuarios u
                LEFT JOIN contas ct ON ct.conta_id = u.conta_id
                LEFT JOIN usuarios s ON s.usuario_id = u.supervisor_id
                WHERE u.email = :email
            """
            ),
            {"email": auth["email"]},
        ).fetchone()
    if not usuario:
        raise HTTPException(404, "Usuário não encontrado")
    dados = dict(usuario._mapping)
    role = normalizar_role(dados.get("role"))
    dados["role"] = role
    dados["is_gerente"] = role == "gerente"
    dados["is_supervisor"] = role == "supervisor"
    # Rótulo pronto para a interface — o card do usuário mostra a função, não o cargo.
    dados["funcao"] = ROTULO_FUNCAO[role]
    # Recursos da assinatura, prontos para a UI. Lista e nao plano cru: a tela
    # pergunta "posso mostrar insights?", nao "que pacote e esse?" -- assim
    # renomear plano ou remontar pacote nao toca em nenhum componente.
    dados["plano"] = dados.get("plano") or PLANO_PADRAO
    dados["recursos"] = recursos_do_plano(dados["plano"])
    return dados


@app.put("/me")
def update_me(dados: UsuarioUpdate, email: str = Depends(get_current_user)):
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                UPDATE usuarios SET
                    nome = COALESCE(:nome, nome),
                    telefone = COALESCE(:telefone, telefone),
                    cargo = COALESCE(:cargo, cargo),
                    empresa_nome = COALESCE(:empresa_nome, empresa_nome),
                    bio = COALESCE(:bio, bio)
                WHERE email = :email
            """
            ),
            {
                "nome": dados.nome,
                "telefone": dados.telefone,
                "cargo": dados.cargo,
                "empresa_nome": dados.empresa_nome,
                "bio": dados.bio,
                "email": email,
            },
        )
    return {"msg": "Perfil atualizado com sucesso 🚀"}


# =========================
# MICROSOFT OAUTH
# =========================
@app.get("/auth/outlook/login")
def outlook_login():
    url = (
        f"https://login.microsoftonline.com/{MICROSOFT_TENANT_ID}/oauth2/v2.0/authorize"
        f"?client_id={MICROSOFT_CLIENT_ID}"
        f"&response_type=code"
        f"&redirect_uri={MICROSOFT_REDIRECT_URI}"
        f"&response_mode=query"
        f"&scope=openid%20profile%20email%20User.Read%20Mail.Read%20Mail.Send%20Calendars.ReadWrite%20offline_access"
    )
    return {"auth_url": url}


@app.get("/auth/outlook/callback")
async def outlook_callback(code: str, email: str = Depends(get_current_user)):
    async with httpx.AsyncClient() as client:
        response = await client.post(
            f"https://login.microsoftonline.com/{MICROSOFT_TENANT_ID}/oauth2/v2.0/token",
            data={
                "client_id": MICROSOFT_CLIENT_ID,
                "client_secret": MICROSOFT_CLIENT_SECRET,
                "code": code,
                "redirect_uri": MICROSOFT_REDIRECT_URI,
                "grant_type": "authorization_code",
            },
        )
    tokens = response.json()
    if "access_token" not in tokens:
        raise HTTPException(400, f"Erro: {tokens.get('error_description', 'Erro desconhecido')}")
    with engine.begin() as conn:
        garantir_colunas_oauth(conn)
        conn.execute(
            text("UPDATE usuarios SET outlook_access_token = :a, outlook_refresh_token = :r WHERE email = :e"),
            {"a": tokens.get("access_token"), "r": tokens.get("refresh_token"), "e": email},
        )
    import threading

    threading.Thread(
        target=setup_outlook_subscription,
        args=(email, tokens.get("access_token"), tokens.get("refresh_token", "")),
        daemon=True,
    ).start()
    threading.Thread(
        target=setup_outlook_calendar_subscription,
        args=(email, tokens.get("access_token"), tokens.get("refresh_token", "")),
        daemon=True,
    ).start()
    return {"msg": "Outlook conectado com sucesso 🚀"}


@app.get("/auth/outlook/status")
def outlook_status(email: str = Depends(get_current_user)):
    with engine.connect() as conn:
        result = conn.execute(
            text("SELECT outlook_access_token FROM usuarios WHERE email = :email"),
            {"email": email},
        ).fetchone()
    if not result:
        raise HTTPException(404, "Usuário não encontrado")
    return {"conectado": result._mapping.get("outlook_access_token") is not None}


@app.delete("/auth/outlook/disconnect")
def outlook_disconnect(email: str = Depends(get_current_user)):
    with engine.begin() as conn:
        conn.execute(
            text("UPDATE usuarios SET outlook_access_token = NULL, outlook_refresh_token = NULL WHERE email = :email"),
            {"email": email},
        )
    return {"msg": "Outlook desconectado com sucesso"}


# =========================
# GOOGLE OAUTH
# =========================
@app.get("/auth/google/login")
def google_login():
    url = (
        f"https://accounts.google.com/o/oauth2/v2/auth?client_id={GOOGLE_CLIENT_ID}"
        f"&response_type=code&redirect_uri={GOOGLE_REDIRECT_URI}"
        f"&scope=https://www.googleapis.com/auth/gmail.send%20"
        f"https://www.googleapis.com/auth/gmail.readonly%20"
        f"https://www.googleapis.com/auth/calendar.events%20"
        f"https://www.googleapis.com/auth/calendar"
        f"&access_type=offline&prompt=consent"
    )
    return {"auth_url": url}


@app.get("/auth/google/callback")
async def google_callback(code: str, email: str = Depends(get_current_user)):
    async with httpx.AsyncClient() as client:
        response = await client.post(
            "https://oauth2.googleapis.com/token",
            data={
                "client_id": GOOGLE_CLIENT_ID,
                "client_secret": GOOGLE_CLIENT_SECRET,
                "code": code,
                "redirect_uri": GOOGLE_REDIRECT_URI,
                "grant_type": "authorization_code",
            },
        )
    tokens = response.json()
    if "access_token" not in tokens:
        raise HTTPException(400, f"Erro Google: {tokens.get('error_description', tokens)}")
    with engine.begin() as conn:
        garantir_colunas_oauth(conn)
        conn.execute(
            text("UPDATE usuarios SET google_access_token = :a, google_refresh_token = :r WHERE email = :e"),
            {"a": tokens.get("access_token"), "r": tokens.get("refresh_token"), "e": email},
        )
    async with httpx.AsyncClient() as client:
        userinfo_res = await client.get(
            "https://www.googleapis.com/oauth2/v2/userinfo",
            headers={"Authorization": f"Bearer {tokens.get('access_token')}"},
        )
    gmail_address = userinfo_res.json().get("email", email)
    import threading

    threading.Thread(
        target=setup_gmail_watch,
        args=(email, tokens.get("access_token"), tokens.get("refresh_token", ""), gmail_address),
        daemon=True,
    ).start()
    threading.Thread(
        target=setup_google_calendar_watch,
        args=(email, tokens.get("access_token"), tokens.get("refresh_token", "")),
        daemon=True,
    ).start()
    return {"msg": "Google conectado com sucesso 🚀"}


@app.get("/auth/google/status")
def google_status(email: str = Depends(get_current_user)):
    with engine.connect() as conn:
        result = conn.execute(
            text("SELECT google_access_token FROM usuarios WHERE email = :email"),
            {"email": email},
        ).fetchone()
    if not result:
        raise HTTPException(404, "Usuário não encontrado")
    return {"conectado": result._mapping.get("google_access_token") is not None}


@app.delete("/auth/google/disconnect")
def google_disconnect(email: str = Depends(get_current_user)):
    with engine.begin() as conn:
        conn.execute(
            text("UPDATE usuarios SET google_access_token = NULL, google_refresh_token = NULL WHERE email = :email"),
            {"email": email},
        )
    return {"msg": "Google desconectado com sucesso"}


# =========================
# TOKEN REFRESH HELPERS
# =========================
async def _refresh_outlook_token(refresh_token: str, email: str) -> str:
    async with httpx.AsyncClient() as client:
        response = await client.post(
            f"https://login.microsoftonline.com/{MICROSOFT_TENANT_ID}/oauth2/v2.0/token",
            data={
                "client_id": MICROSOFT_CLIENT_ID,
                "client_secret": MICROSOFT_CLIENT_SECRET,
                "refresh_token": refresh_token,
                "grant_type": "refresh_token",
                "scope": "openid profile email User.Read Mail.Read Mail.Send Calendars.ReadWrite offline_access",
            },
        )
    tokens = response.json()
    new_access = tokens.get("access_token")
    if new_access:
        with engine.begin() as conn:
            conn.execute(
                text("UPDATE usuarios SET outlook_access_token = :a, outlook_refresh_token = :r WHERE email = :e"),
                {"a": new_access, "r": tokens.get("refresh_token", refresh_token), "e": email},
            )
    return new_access


async def _refresh_google_token(refresh_token: str, email: str) -> str:
    async with httpx.AsyncClient() as client:
        response = await client.post(
            "https://oauth2.googleapis.com/token",
            data={
                "client_id": GOOGLE_CLIENT_ID,
                "client_secret": GOOGLE_CLIENT_SECRET,
                "refresh_token": refresh_token,
                "grant_type": "refresh_token",
            },
        )
    tokens = response.json()
    new_access = tokens.get("access_token")
    if new_access:
        with engine.begin() as conn:
            conn.execute(
                text("UPDATE usuarios SET google_access_token = :a WHERE email = :e"),
                {"a": new_access, "e": email},
            )
    return new_access


# =========================
# REUNIÃO OUTLOOK
# =========================
@app.post("/eventos/{evento_id}/agendar-outlook")
async def agendar_reuniao_outlook(evento_id: str, reuniao: ReuniaoOutlook, email: str = Depends(get_current_user)):
    try:
        with engine.connect() as conn:
            usuario = conn.execute(
                text("SELECT outlook_access_token, outlook_refresh_token FROM usuarios WHERE email = :email"),
                {"email": email},
            ).fetchone()
        if not usuario or not usuario._mapping.get("outlook_access_token"):
            raise HTTPException(400, "Outlook não conectado.")
        access_token = usuario._mapping["outlook_access_token"]
        refresh_token = usuario._mapping.get("outlook_refresh_token")
        data_str = reuniao.data.isoformat()
        evento_graph = {
            "subject": reuniao.titulo,
            "body": {"contentType": "HTML", "content": reuniao.descricao or ""},
            "start": {"dateTime": f"{data_str}T{reuniao.hora_inicio}:00", "timeZone": "America/Sao_Paulo"},
            "end": {"dateTime": f"{data_str}T{reuniao.hora_fim}:00", "timeZone": "America/Sao_Paulo"},
        }
        todos_emails = reuniao.emails_convidados or ([reuniao.email_convidado] if reuniao.email_convidado else [])
        todos_emails = [e for e in todos_emails if e and e.strip()]
        if todos_emails:
            evento_graph["attendees"] = [{"emailAddress": {"address": e.strip()}, "type": "required"} for e in todos_emails]
        headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}
        async with httpx.AsyncClient() as client:
            response = await client.post("https://graph.microsoft.com/v1.0/me/events", json=evento_graph, headers=headers)
        if response.status_code == 401 and refresh_token:
            access_token = await _refresh_outlook_token(refresh_token, email)
            if not access_token:
                raise HTTPException(401, "Token expirado. Reconecte o Outlook.")
            headers["Authorization"] = f"Bearer {access_token}"
            async with httpx.AsyncClient() as client:
                response = await client.post(
                    "https://graph.microsoft.com/v1.0/me/events",
                    json=evento_graph,
                    headers=headers,
                )
        if response.status_code not in (200, 201):
            raise HTTPException(500, f"Erro Outlook: {response.text}")
        outlook_event = response.json()
        with engine.begin() as conn:
            conn.execute(
                text(
                    """
                UPDATE eventos SET outlook_event_id = :oid
                WHERE evento_id = :id AND usuario_email = :email
            """
                ),
                {"oid": outlook_event.get("id"), "id": evento_id, "email": email},
            )
        return {
            "msg": "Reunião criada no Outlook Calendar 🚀",
            "outlook_event_id": outlook_event.get("id"),
            "link": outlook_event.get("webLink"),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# =========================
# REUNIÃO GOOGLE
# =========================
@app.post("/eventos/{evento_id}/agendar-google")
async def agendar_reuniao_google(evento_id: str, reuniao: ReuniaoGoogle, email: str = Depends(get_current_user)):
    try:
        with engine.connect() as conn:
            usuario = conn.execute(
                text("SELECT google_access_token, google_refresh_token FROM usuarios WHERE email = :email"),
                {"email": email},
            ).fetchone()
        if not usuario or not usuario._mapping.get("google_access_token"):
            raise HTTPException(400, "Google Calendar não conectado.")
        access_token = usuario._mapping["google_access_token"]
        refresh_token = usuario._mapping.get("google_refresh_token")
        data_str = reuniao.data.isoformat()
        evento_google = {
            "summary": reuniao.titulo,
            "description": reuniao.descricao or "",
            "start": {"dateTime": f"{data_str}T{reuniao.hora_inicio}:00", "timeZone": "America/Sao_Paulo"},
            "end": {"dateTime": f"{data_str}T{reuniao.hora_fim}:00", "timeZone": "America/Sao_Paulo"},
        }
        todos_emails_g = reuniao.emails_convidados or ([reuniao.email_convidado] if reuniao.email_convidado else [])
        todos_emails_g = [e for e in todos_emails_g if e and e.strip()]
        if todos_emails_g:
            evento_google["attendees"] = [{"email": e.strip()} for e in todos_emails_g]
        headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://www.googleapis.com/calendar/v3/calendars/primary/events?sendUpdates=all",
                json=evento_google,
                headers=headers,
            )
        if response.status_code == 401 and refresh_token:
            access_token = await _refresh_google_token(refresh_token, email)
            if not access_token:
                raise HTTPException(401, "Token expirado. Reconecte o Google.")
            headers["Authorization"] = f"Bearer {access_token}"
            async with httpx.AsyncClient() as client:
                response = await client.post(
                    "https://www.googleapis.com/calendar/v3/calendars/primary/events?sendUpdates=all",
                    json=evento_google,
                    headers=headers,
                )
        if response.status_code not in (200, 201):
            raise HTTPException(500, f"Erro Google Calendar: {response.text}")
        
        google_event = response.json()

        print("[GOOGLE EVENT]")
        print(json.dumps(google_event, indent=2))

        with engine.begin() as conn:
            conn.execute(
                text(
                    """
                UPDATE eventos
                SET google_event_id = :gid,
                    email_convidado = COALESCE(:email_convidado, email_convidado)
                WHERE evento_id = :id
                AND usuario_email = :email
            """
                ),
                {
                    "gid": google_event.get("id"),
                    "email_convidado": (todos_emails_g[0] if todos_emails_g else reuniao.email_convidado),
                    "id": evento_id,
                    "email": email,
                },
            )


        return {
            "msg": "Reunião criada no Google Calendar 🚀",
            "google_event_id": google_event.get("id"),
            "link": google_event.get("htmlLink"),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# =========================
# EVENTOS
# =========================
@app.get("/eventos")
def listar_eventos(auth: dict = Depends(get_auth)):
    with engine.connect() as conn:
        # Gerente enxerga a agenda de toda a conta; supervisor, a da equipe dele;
        # vendedor, só a sua.
        emails = escopo_emails(conn, auth)
        if emails is None:
            result = conn.execute(
                text("SELECT * FROM eventos WHERE conta_id = :cid ORDER BY data, hora_inicio"),
                {"cid": auth["conta_id"]},
            )
        else:
            result = conn.execute(
                text("SELECT * FROM eventos WHERE usuario_email = ANY(:emails) ORDER BY data, hora_inicio"),
                {"emails": emails},
            )
        return [dict(row._mapping) for row in result]


@app.post("/eventos", status_code=201)
def criar_evento(evento: EventoCreate, auth: dict = Depends(get_auth)):
    evento_id = str(uuid.uuid4())
    with engine.begin() as conn:
        garantir_tabela_notificacoes(conn)
        conn.execute(
            text(
                """
            INSERT INTO eventos (evento_id, titulo, tipo, data, hora_inicio, hora_fim,
                empresa_id, empresa_nome, descricao, email_convidado, usuario_email, conta_id, criado_em)
            VALUES (:id, :titulo, :tipo, :data, :hora_inicio, :hora_fim,
                :empresa_id, :empresa_nome, :descricao, :email_convidado, :email, :conta_id, NOW())
        """
            ),
            {
                "id": evento_id,
                "titulo": evento.titulo,
                "tipo": evento.tipo,
                "data": evento.data,
                "hora_inicio": evento.hora_inicio,
                "hora_fim": evento.hora_fim,
                "empresa_id": evento.empresa_id,
                "empresa_nome": evento.empresa_nome,
                "descricao": evento.descricao,
                "email_convidado": evento.email_convidado,
                "email": auth["email"],
                "conta_id": auth["conta_id"],
            },
        )
    return {"msg": "Evento criado com sucesso 🚀", "id": evento_id}


@app.put("/eventos/{evento_id}")
def atualizar_evento(evento_id: str, evento: EventoUpdate, email: str = Depends(get_current_user)):
    with engine.begin() as conn:
        result = conn.execute(
            text("SELECT evento_id FROM eventos WHERE evento_id = :id AND usuario_email = :email"),
            {"id": evento_id, "email": email},
        ).fetchone()
        if not result:
            raise HTTPException(404, "Evento não encontrado")
        conn.execute(
            text(
                """
            UPDATE eventos SET titulo=COALESCE(:titulo,titulo), tipo=COALESCE(:tipo,tipo),
                data=COALESCE(:data,data), hora_inicio=COALESCE(:hora_inicio,hora_inicio),
                hora_fim=COALESCE(:hora_fim,hora_fim), empresa_id=COALESCE(:empresa_id,empresa_id),
                empresa_nome=COALESCE(:empresa_nome,empresa_nome), descricao=COALESCE(:descricao,descricao),
                email_convidado=COALESCE(:email_convidado,email_convidado)
            WHERE evento_id=:id AND usuario_email=:email
        """
            ),
            {
                "titulo": evento.titulo,
                "tipo": evento.tipo,
                "data": evento.data,
                "hora_inicio": evento.hora_inicio,
                "hora_fim": evento.hora_fim,
                "empresa_id": evento.empresa_id,
                "empresa_nome": evento.empresa_nome,
                "descricao": evento.descricao,
                "email_convidado": evento.email_convidado,
                "id": evento_id,
                "email": email,
            },
        )
    return {"msg": "Evento atualizado com sucesso 🚀"}


@app.delete("/eventos/{evento_id}")
def deletar_evento(evento_id: str, email: str = Depends(get_current_user)):
    with engine.begin() as conn:
        result = conn.execute(
            text("DELETE FROM eventos WHERE evento_id=:id AND usuario_email=:email RETURNING evento_id"),
            {"id": evento_id, "email": email},
        ).fetchone()
    if not result:
        raise HTTPException(404, "Evento não encontrado")
    return {"msg": "Evento deletado com sucesso"}


@app.get("/empresas/{empresa_id}/atividades")
def listar_atividades_empresa(empresa_id: str, auth: dict = Depends(get_auth)):
    with engine.connect() as conn:
        checar_acesso_empresa(conn, empresa_id, auth)
        # Gerente vê todas as atividades da empresa; supervisor, as da equipe;
        # vendedor, só as suas.
        emails = escopo_emails(conn, auth)
        if emails is None:
            escopo = ""
            params = {"empresa_id": empresa_id}
        else:
            escopo = "AND usuario_email = ANY(:emails)"
            params = {"empresa_id": empresa_id, "emails": emails}
        result = conn.execute(
            text(f"""
                SELECT evento_id, titulo, tipo, data, hora_inicio, hora_fim,
                       empresa_id, empresa_nome, email_convidado, status_resposta, criado_em
                FROM eventos
                WHERE empresa_id = :empresa_id
                  {escopo}
                ORDER BY data DESC, hora_inicio DESC
            """),
            params,
        )
        rows = []
        for row in result:
            r = dict(row._mapping)
            data = r.get("data")
            hora = r.get("hora_inicio")
            r["data_hora"] = f"{data}T{hora}" if data and hora else (str(data) if data else None)
            rows.append(r)
        return rows


@app.put("/eventos/{evento_id}/status")
def atualizar_status_evento(evento_id: str, body: dict, email: str = Depends(get_current_user)):
    status = body.get("status_resposta")
    if status not in ("aceito", "negado", "talvez", "novo_horario", "pendente"):
        raise HTTPException(400, "Status inválido")
    with engine.begin() as conn:
        result = conn.execute(
            text("UPDATE eventos SET status_resposta=:status WHERE evento_id=:id AND usuario_email=:email RETURNING evento_id"),
            {"status": status, "id": evento_id, "email": email},
        ).fetchone()
    if not result:
        raise HTTPException(404, "Evento não encontrado")
    return {"msg": "Status atualizado"}


# =========================
# SEGMENTOS
# =========================
@app.get("/segmentos")
def listar_segmentos():
    with engine.begin() as conn:
        garantir_tabela_segmentos(conn)
        result = conn.execute(text("SELECT nome FROM segmentos ORDER BY nome"))
        return {"segmentos": [row._mapping["nome"] for row in result]}


@app.post("/segmentos", status_code=201)
def criar_segmento(segmento: SegmentoCreate):
    with engine.begin() as conn:
        nome = salvar_segmento(conn, segmento.nome)
    return {"nome": nome, "validado": True}


# =========================
# EMPRESAS
# =========================
def _normalizar_query(q: str) -> str:
    # Normaliza acentos/caixa/espaços para maximizar acertos no cache
    # (ex.: "Metalúrgicas ", "metalurgicas" -> "metalurgicas") e reduzir
    # chamadas pagas à API do Google.
    q = unicodedata.normalize("NFKD", q or "").encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", q).strip().lower()


def _recalcular_cadastradas(results_list: list) -> list:
    place_ids = [r["place_id"] for r in results_list if r.get("place_id")]
    if not place_ids:
        return results_list
    with engine.connect() as conn:
        rows = conn.execute(
            text("SELECT google_place_id FROM empresas WHERE google_place_id = ANY(:ids)"),
            {"ids": place_ids},
        )
        ja_cadastradas = {r[0] for r in rows}
    for r in results_list:
        r["ja_cadastrada"] = r.get("place_id") in ja_cadastradas
    return results_list


# =========================================================================
# TETO MENSAL DE CHAMADAS PAGAS AO GOOGLE PLACES
# =========================================================================
def _fuso_cota_places() -> timezone:
    return timezone(timedelta(hours=PLACES_FUSO_HORAS))


def _mes_cota_places() -> str:
    """Competencia do balde, 'AAAA-MM', no fuso configurado."""
    return datetime.now(_fuso_cota_places()).strftime("%Y-%m")


def _reset_cota_places() -> datetime:
    """Instante em UTC em que o balde zera: meia-noite do dia 1 do mes seguinte."""
    agora = datetime.now(_fuso_cota_places())
    # Dia 28 + 4 dias cai sempre no mes seguinte, em qualquer mes do calendario.
    proximo = (agora.replace(day=28) + timedelta(days=4)).replace(
        day=1, hour=0, minute=0, second=0, microsecond=0
    )
    return proximo.astimezone(timezone.utc)


def _chave_cota_places(auth: dict) -> str:
    """Balde individual: o teto e por usuario, nao por assinatura."""
    return f"usuario:{auth['usuario_id']}" if auth.get("usuario_id") else f"email:{auth['email']}"


def _chave_conta_places(auth: dict) -> str:
    """O consumo e por usuario, mas o interruptor e da conta: quem decide
    desligar o teto e o gerente, e a decisao vale para a assinatura inteira."""
    return f"conta:{auth['conta_id']}" if auth.get("conta_id") else f"email:{auth['email']}"


def _teto_ligado(conta_chave: str) -> bool:
    with engine.begin() as conn:
        garantir_tabela_places_teto(conn)
        valor = conn.execute(
            text("SELECT ligado FROM places_teto_conta WHERE conta_id = :c"),
            {"c": conta_chave},
        ).scalar()
    return True if valor is None else bool(valor)


def _detalhe_cota_places(usadas: int) -> dict:
    """Corpo do 429 do NOSSO teto. O front desabilita o campo ate `reset_em` e
    mostra `mensagem` -- por isso a mensagem tem de ser verdadeira: o contador
    zera na virada do mes do fuso configurado, nao "amanha" por suposicao."""
    if PLACES_LIMITE_MENSAL <= 0:
        msg = ("Busca nova esta desativada (limite mensal configurado como 0). "
               "O que ja esta em cache continua abrindo.")
    else:
        msg = (f"Voce atingiu seu limite de {PLACES_LIMITE_MENSAL} buscas novas no mes. "
               "Buscas ja feitas antes continuam abrindo pelo cache, sem custo.")
    return {
        "erro": "limite_mensal_places",
        "escopo": "usuario",
        "usadas": usadas,
        "limite": PLACES_LIMITE_MENSAL,
        "reset_em": _reset_cota_places().isoformat(),
        "mensagem": msg,
    }


def reservar_chamada_places(auth: dict) -> int:
    """Reserva UMA chamada paga antes de chamar o Google. Levanta 429 no estouro.

    Cobra primeiro e estorna se a chamada nao acontecer (`devolver_chamada_places`).
    O caminho inverso -- contar depois -- deixaria duas requisicoes simultaneas
    passarem pelo mesmo ultimo credito. Aqui o `ON CONFLICT ... WHERE` resolve a
    corrida dentro do proprio UPDATE: quem perde nao recebe linha de volta.

    Com o teto desligado pelo gerente, continua CONTANDO e para de BARRAR: o
    numero de /places/cota segue valendo (e como se ve o consumo real durante o
    teste), e o unico limite que resta e o do Google Cloud."""
    chave = _chave_cota_places(auth)
    if not _teto_ligado(_chave_conta_places(auth)):
        with engine.begin() as conn:
            garantir_tabela_places_uso(conn)
            return conn.execute(
                text("""
                    INSERT INTO places_uso_mensal (chave, mes, chamadas, atualizado_em)
                    VALUES (:chave, :mes, 1, NOW())
                    ON CONFLICT (chave, mes) DO UPDATE
                       SET chamadas = places_uso_mensal.chamadas + 1, atualizado_em = NOW()
                    RETURNING chamadas
                """),
                {"chave": chave, "mes": _mes_cota_places()},
            ).scalar()
    if PLACES_LIMITE_MENSAL <= 0:
        raise HTTPException(429, detail=_detalhe_cota_places(0))
    with engine.begin() as conn:
        garantir_tabela_places_uso(conn)
        usadas = conn.execute(
            text("""
                INSERT INTO places_uso_mensal (chave, mes, chamadas, atualizado_em)
                VALUES (:chave, :mes, 1, NOW())
                ON CONFLICT (chave, mes) DO UPDATE
                   SET chamadas = places_uso_mensal.chamadas + 1, atualizado_em = NOW()
                 WHERE places_uso_mensal.chamadas < :limite
                RETURNING chamadas
            """),
            {"chave": chave, "mes": _mes_cota_places(), "limite": PLACES_LIMITE_MENSAL},
        ).scalar()
    if usadas is None:
        raise HTTPException(429, detail=_detalhe_cota_places(PLACES_LIMITE_MENSAL))
    return usadas


def devolver_chamada_places(chave: str) -> None:
    """Estorna a reserva quando a chamada nao chegou a ser faturada.

    O Google cobra a resposta 200; timeout, 5xx e recusa por cota nao geram
    cobranca, entao tambem nao podem consumir o teto do mes."""
    try:
        with engine.begin() as conn:
            conn.execute(
                text("""
                    UPDATE places_uso_mensal
                       SET chamadas = GREATEST(chamadas - 1, 0), atualizado_em = NOW()
                     WHERE chave = :chave AND mes = :mes
                """),
                {"chave": chave, "mes": _mes_cota_places()},
            )
    except Exception as e:  # noqa: BLE001 - estorno nunca derruba a requisicao
        print("⚠️ falha ao estornar cota do places:", e)


def ler_cota_places(auth: dict) -> dict:
    ligado = _teto_ligado(_chave_conta_places(auth))
    with engine.begin() as conn:
        garantir_tabela_places_uso(conn)
        usadas = conn.execute(
            text("SELECT chamadas FROM places_uso_mensal WHERE chave = :c AND mes = :m"),
            {"c": _chave_cota_places(auth), "m": _mes_cota_places()},
        ).scalar() or 0
    bloqueado = ligado and usadas >= PLACES_LIMITE_MENSAL
    return {
        "usadas": usadas,
        "limite": PLACES_LIMITE_MENSAL,
        "restantes": max(PLACES_LIMITE_MENSAL - usadas, 0) if ligado else None,
        "bloqueado": bloqueado,
        "teto_ligado": ligado,
        # So o gerente alterna: o vendedor nao levanta o proprio teto.
        "pode_alternar": auth.get("is_gerente", False),
        "reset_em": _reset_cota_places().isoformat(),
        "mensagem": _detalhe_cota_places(usadas)["mensagem"] if bloqueado else None,
    }


@app.get("/places/cota")
def cota_places(auth: dict = Depends(get_auth)):
    """Estado do teto do mes. O front le isto na montagem em vez de adivinhar
    pelo localStorage -- a verdade sobre a cota mora no servidor."""
    return ler_cota_places(auth)


@app.post("/places/cota/teto")
def alternar_teto_places(req: PlacesTetoUpdate, auth: dict = Depends(exigir_gerente)):
    """Liga/desliga o teto mensal da conta (o botao de teste do gerente).

    Desligado, nenhuma busca e barrada pelo nosso lado -- o unico limite que
    sobra e o do Google Cloud. Por isso a rota e do gerente, guarda quem mexeu
    e a hora, e o padrao continua sendo LIGADO."""
    with engine.begin() as conn:
        garantir_tabela_places_teto(conn)
        conn.execute(
            text("""
                INSERT INTO places_teto_conta (conta_id, ligado, alterado_em, alterado_por)
                VALUES (:c, :ligado, NOW(), :quem)
                ON CONFLICT (conta_id) DO UPDATE
                   SET ligado = :ligado, alterado_em = NOW(), alterado_por = :quem
            """),
            {"c": _chave_conta_places(auth), "ligado": req.ligado, "quem": auth["email"]},
        )
    return ler_cota_places(auth)


def _bias_localizacao(lat: float, lng: float, raio_m: int) -> dict:
    """Area de busca para o `locationBias` do searchText.

    Ate 50 km usa circulo, que e o que a API aceita e o que descreve melhor um
    raio. Acima disso a API recusa o circulo, entao vira um retangulo que
    ENVOLVE o circulo pedido -- area um pouco maior nos cantos, o que para
    prospeccao e aceitavel (rede mais larga), e nao ha alternativa de circulo
    grande na API."""
    if raio_m <= RAIO_CIRCULO_MAX_M:
        return {"circle": {"center": {"latitude": lat, "longitude": lng}, "radius": float(raio_m)}}
    graus_lat = raio_m / 111_320.0
    # Longitude encurta com a latitude; o cos evita um retangulo estreito demais
    # longe do equador. O piso impede divisao por ~zero perto dos polos.
    graus_lng = raio_m / (111_320.0 * max(math.cos(math.radians(lat)), 0.01))
    return {"rectangle": {
        "low": {"latitude": max(lat - graus_lat, -89.9), "longitude": max(lng - graus_lng, -179.9)},
        "high": {"latitude": min(lat + graus_lat, 89.9), "longitude": min(lng + graus_lng, 179.9)},
    }}


@app.post("/places/search")
async def search_places(req: PlacesSearchRequest, auth: dict = Depends(get_auth)):
    """Busca empresas no Google Places dentro de um raio.

    `radius` vem em metros e faz parte da identidade do resultado: a mesma query
    no mesmo ponto com raios diferentes sao buscas diferentes, e o cache
    (30 dias) e chaveado tambem por ele. Ate 50 km a area e um circulo; acima
    disso a Places API recusa o circulo e usamos um retangulo que o envolve,
    entao entram tambem os cantos. O retorno e limitado a 20 empresas
    independentemente do raio -- raio maior amplia a area, nao a quantidade."""
    usuario_email = auth["email"]
    if not GOOGLE_PLACES_API_KEY:
        raise HTTPException(503, "Google Places API não configurada")

    lat = req.lat or -15.7801
    lng = req.lng or -47.9292
    lat_grid = round(lat, 1)
    lng_grid = round(lng, 1)
    # Raio em km inteiros, dentro dos limites. Arredondar agrupa buscas quase
    # iguais (99,6 km e 100 km sao a mesma coisa) sem deixar o cache servir um
    # raio visivelmente diferente do pedido.
    raio_m = max(1_000, min(int(req.radius or 15000), RAIO_MAX_M))
    raio_grid = round(raio_m / 1000)
    query_norm = _normalizar_query(req.query)

    # Bloqueia queries vazias/curtas demais para não gastar chamada paga.
    if len(query_norm) < 2:
        return []

    # Cache de conteúdo — válido por 30 dias (limite dos Termos do Google Maps
    # Platform). Após esse prazo o conteúdo expira e é rebuscado. O place_id é o
    # único dado de Places guardado permanentemente (na tabela 'empresas').
    # IMPORTANTE: no acerto de cache só incrementamos o contador de popularidade;
    # NÃO mexemos em updated_at, senão o conteúdo nunca expiraria (violaria a
    # regra dos 30 dias).
    with engine.begin() as conn:
        garantir_tabelas_places_cache(conn)
        cached = conn.execute(
            text("""
                SELECT id, results FROM places_cache
                WHERE query=:q AND lat_grid=:lat AND lng_grid=:lng
                AND COALESCE(raio_grid, 15)=:raio
                AND updated_at >= NOW() - INTERVAL '30 days'
            """),
            {"q": query_norm, "lat": lat_grid, "lng": lng_grid, "raio": raio_grid},
        ).fetchone()
        if cached:
            conn.execute(
                text("UPDATE places_cache SET search_count=search_count+1 WHERE id=:id"),
                {"id": cached.id},
            )
            results = cached.results if isinstance(cached.results, list) else json.loads(cached.results)
            return _recalcular_cadastradas(results)

    # Cache miss: daqui para frente a chamada e paga. O teto entra ANTES dela.
    chave_cota = _chave_cota_places(auth)
    reservar_chamada_places(auth)

    # 3. Chama Google Places API
    payload = {
        "textQuery": req.query,
        "locationBias": _bias_localizacao(lat, lng, raio_m),
        "languageCode": "pt-BR",
        "regionCode": "BR",
        "maxResultCount": 20,
    }
    api_headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": GOOGLE_PLACES_API_KEY,
        # Field mask enxuto: mantém apenas campos da faixa Pro (endereço/local/nome).
        # Removidos rating/userRatingCount (Atmosphere) e telefone/site/horários
        # (Enterprise) para a busca cair da faixa Enterprise para a Pro, mais barata.
        # addressComponents fornece cidade, bairro, cep e rua; os demais campos Pro
        # (id/displayName/formattedAddress/location) não adicionam custo.
        "X-Goog-FieldMask": "places.id,places.displayName,places.formattedAddress,places.addressComponents,places.location,places.businessStatus,places.primaryTypeDisplayName",
    }
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.post("https://places.googleapis.com/v1/places:searchText", json=payload, headers=api_headers)
    except httpx.HTTPError as e:
        devolver_chamada_places(chave_cota)
        raise HTTPException(502, f"Google Places indisponivel: {e}")

    # So resposta 200 e faturada; qualquer outra devolve o credito reservado.
    if resp.status_code != 200:
        devolver_chamada_places(chave_cota)

    if resp.status_code == 429 or (
        resp.status_code != 200 and
        any(k in resp.text.upper() for k in ("RESOURCE_EXHAUSTED", "QUOTA", "RATE_LIMIT"))
    ):
        with engine.begin() as conn:
            garantir_tabela_notificacoes(conn)
            existe = conn.execute(
                text("SELECT 1 FROM notificacoes WHERE usuario_email=:e AND tipo='quota_exceeded' AND criado_em >= NOW() - INTERVAL '1 hour'"),
                {"e": usuario_email},
            ).fetchone()
            if not existe:
                conn.execute(
                    text("INSERT INTO notificacoes (notificacao_id, usuario_email, tipo, titulo, mensagem, lida, criado_em) VALUES (:id, :e, 'quota_exceeded', :titulo, :msg, FALSE, NOW())"),
                    {"id": str(uuid.uuid4()), "e": usuario_email, "titulo": "Google recusou a busca por cota", "msg": f"O Google recusou a busca por limite de cota. Pode ser o limite por minuto ou o teto da conta no Google Cloud — a tela libera nova tentativa em {PLACES_ESPERA_GOOGLE_MIN} min."},
                )
        raise HTTPException(429, detail={
            "erro": "cota_google",
            "escopo": "google",
            "usadas": None,
            "limite": None,
            "reset_em": (datetime.now(timezone.utc) + timedelta(minutes=PLACES_ESPERA_GOOGLE_MIN)).isoformat(),
            # Aqui quem recusou foi o Google, e nao da para afirmar quando volta:
            # o 429 dele cobre tanto limite por minuto quanto teto da conta.
            "mensagem": ("O Google recusou a busca por cota — pode ser o limite por minuto ou o teto "
                         f"da conta no Google Cloud. Nova tentativa liberada em {PLACES_ESPERA_GOOGLE_MIN} min."),
        })

    if resp.status_code != 200:
        raise HTTPException(502, f"Google Places erro: {resp.text}")

    data = resp.json()
    places = data.get("places", [])
    place_ids = [p["id"] for p in places if "id" in p]
    ja_cadastradas = set()
    if place_ids:
        with engine.connect() as conn:
            garantir_colunas_places(conn)
            rows = conn.execute(text("SELECT google_place_id FROM empresas WHERE google_place_id = ANY(:ids)"), {"ids": place_ids})
            ja_cadastradas = {r[0] for r in rows}

    result = []
    for p in places:
        loc = p.get("location", {})
        nome_obj = p.get("displayName", {})
        tipo_obj = p.get("primaryTypeDisplayName", {})
        address_components = p.get("addressComponents", [])
        cidade = ""
        bairro = ""
        cep = ""
        rua = ""
        numero = ""
        for comp in address_components:
            types = comp.get("types", [])
            valor = comp.get("longText", "")
            if "locality" in types and not cidade:
                cidade = valor
            elif "administrative_area_level_2" in types and not cidade:
                cidade = valor
            elif ("sublocality_level_1" in types or "sublocality" in types) and not bairro:
                bairro = valor
            elif "postal_code" in types:
                cep = valor.replace("-", "").replace(" ", "")
            elif "route" in types:
                rua = valor
            elif "street_number" in types:
                numero = valor
        endereco_rua = f"{rua}, {numero}".strip(", ") if rua else ""
        result.append({
            "place_id": p.get("id"),
            "nome": nome_obj.get("text", ""),
            "endereco": p.get("formattedAddress", ""),
            "endereco_rua": endereco_rua,
            "cidade": cidade,
            "bairro": bairro,
            "cep": cep,
            "lat": loc.get("latitude"),
            "lng": loc.get("longitude"),
            "business_status": p.get("businessStatus"),
            "tipo": tipo_obj.get("text") if tipo_obj else None,
            "ja_cadastrada": p.get("id") in ja_cadastradas,
        })

    # 4. Salva no cache
    with engine.begin() as conn:
        garantir_tabelas_places_cache(conn)
        conn.execute(
            text("""
                INSERT INTO places_cache (id, query, lat_grid, lng_grid, raio_grid, results, search_count, created_at, updated_at)
                VALUES (:id, :q, :lat, :lng, :raio, CAST(:results AS JSONB), 1, NOW(), NOW())
                ON CONFLICT (query, lat_grid, lng_grid, raio_grid)
                DO UPDATE SET results=CAST(:results AS JSONB), search_count=places_cache.search_count+1, updated_at=NOW()
            """),
            {"id": str(uuid.uuid4()), "q": query_norm, "lat": lat_grid, "lng": lng_grid,
             "raio": raio_grid, "results": json.dumps(result)},
        )

    return result


@app.post("/places/generate-ranking")
def generate_ranking(usuario_email: str = Depends(get_current_user)):
    gerar_ranking_mensal()
    return {"msg": "Ranking gerado com sucesso"}


@app.get("/places/top10")
def top10_places(usuario_email: str = Depends(get_current_user)):
    with engine.begin() as conn:
        garantir_tabelas_places_cache(conn)
        rows = conn.execute(
            text("SELECT rank_position, query, search_count, month FROM places_ranking ORDER BY rank_position ASC")
        ).fetchall()
    # Apenas termos mais buscados (analítico). Não devolvemos conteúdo de places
    # aqui para respeitar o limite de cache de 30 dias do Google.
    return [
        {
            "posicao": row.rank_position,
            "query": row.query,
            "total_buscas": row.search_count,
            "mes": row.month,
        }
        for row in rows
    ]


@app.get("/empresas/rascunhos")
def listar_rascunhos(auth: dict = Depends(get_auth)):
    with engine.connect() as conn:
        garantir_colunas_places(conn)
        # Vendedor: só os seus. Supervisor: os da equipe dele. Gerente: a conta toda.
        trecho, params = filtro_escopo(conn, auth)
        escopo = f"AND conta_id = :cid {trecho}"
        rows = conn.execute(
            text(
                "SELECT * FROM empresas WHERE status_cadastro = 'rascunho'"
                f" {escopo}"
                " ORDER BY status_atualizado_em DESC NULLS LAST"
            ),
            params,
        )
        return [dict(r._mapping) for r in rows]


@app.post("/empresas/rascunho", status_code=201)
def criar_rascunho(rascunho: RascunhoCreate, auth: dict = Depends(get_auth)):
    """Cria a empresa como RASCUNHO -- fora do funil ate alguem completar.

    `status` nasce 'Rascunho', nao 'Lead'. As duas colunas existem e parecem
    redundantes, mas quem manda na interface e `status`: toda a tela decide "isto
    e rascunho?" com `status === 'Rascunho'` e nenhuma linha do frontend le
    `status_cadastro`. Nascendo como 'Lead', a empresa entrava direto no funil,
    contava na conversao e na base -- sem ninguem ter olhado para ela. Isso
    importa mais agora que da para criar 20 de uma vez pela busca.

    `status_cadastro` continua 'rascunho' porque as contagens do backend
    (/empresas/rascunhos, /gerencia/dashboard) filtram por ele.

    Vira Lead quando o usuario abre a ficha, edita e salva."""
    with engine.begin() as conn:
        garantir_colunas_places(conn)
        garantir_campos_pipeline(conn)
        # Verifica duplicata por google_place_id dentro da mesma conta
        if rascunho.google_place_id:
            existing = conn.execute(
                text("SELECT empresa_id FROM empresas WHERE google_place_id = :gid AND conta_id = :cid"),
                {"gid": rascunho.google_place_id, "cid": auth["conta_id"]},
            ).fetchone()
            if existing:
                raise HTTPException(409, {"message": "Empresa já cadastrada", "empresa_id": str(existing[0])})
        empresa_id = str(uuid.uuid4())
        conn.execute(
            text("""
                INSERT INTO empresas (empresa_id, nome, cidade, endereco_completo, site, telefone_empresa,
                    google_place_id, latitude, longitude, google_rating, google_rating_count, business_status,
                    status, status_cadastro, origem_lead, temperatura, responsavel_principal,
                    conta_id, vendedor_id, ultima_interacao, status_atualizado_em)
                VALUES (:id, :nome, :cidade, :endereco_completo, :site, :telefone_empresa,
                    :google_place_id, :latitude, :longitude, :google_rating, :google_rating_count, :business_status,
                    'Rascunho', 'rascunho', 'Google Maps', 'Frio', :responsavel_principal,
                    :conta_id, :vendedor_id, NOW(), NOW())
            """),
            {
                "id": empresa_id,
                "nome": rascunho.nome,
                "cidade": rascunho.cidade,
                "endereco_completo": rascunho.endereco_completo,
                "site": rascunho.site,
                "telefone_empresa": rascunho.telefone_empresa,
                "google_place_id": rascunho.google_place_id,
                "latitude": rascunho.latitude,
                "longitude": rascunho.longitude,
                "google_rating": rascunho.google_rating,
                "google_rating_count": rascunho.google_rating_count,
                "business_status": rascunho.business_status,
                "responsavel_principal": auth["email"],
                "conta_id": auth["conta_id"],
                "vendedor_id": auth["usuario_id"],
            },
        )
    return {"empresa_id": empresa_id, "status_cadastro": "rascunho"}


ALERTA_EXPORTACAO_MASSA = int(os.getenv("ALERTA_EXPORTACAO_MASSA", "100"))


@app.get("/empresas")
def listar_empresas(request: Request, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        garantir_campos_pipeline(conn)
        # Vendedor: só a própria carteira. Supervisor: a da equipe. Gerente: tudo da conta.
        trecho, params = filtro_escopo(conn, auth, prefixo="e.")
        escopo = f"WHERE e.conta_id = :cid {trecho}"
        result = conn.execute(
            text(
                f"""
            SELECT e.*, c.email AS contato_email, c.celular AS contato_celular, c.whatsapp AS contato_whatsapp
            FROM empresas e
            LEFT JOIN LATERAL (
                SELECT email, celular, whatsapp FROM contatos WHERE empresa_id = e.empresa_id
                ORDER BY decisor DESC NULLS LAST, data_criacao ASC NULLS LAST LIMIT 1
            ) c ON TRUE
            {escopo}
            ORDER BY COALESCE(e.status_atualizado_em, e.ultima_interacao) DESC NULLS LAST, e.nome ASC
        """
            ),
            params,
        )
        rows = [dict(row._mapping) for row in result]
        # Auditoria de acesso à base de leads + alerta de leitura em massa (LGPD).
        acao = "LEADS_EXPORTACAO_MASSA" if len(rows) >= ALERTA_EXPORTACAO_MASSA else "LEADS_LISTADOS"
        registrar_auditoria(usuario=auth, acao=acao, recurso="empresas",
                            quantidade=len(rows), request=request, conn=conn)
        if len(rows) >= ALERTA_EXPORTACAO_MASSA:
            print(f"🚨 ALERTA: {auth['email']} listou {len(rows)} leads de uma vez (conta {auth['conta_id']}).")
        return rows


@app.get("/empresas/{empresa_id}")
def buscar_empresa(empresa_id: str, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        garantir_campos_pipeline(conn)
        checar_acesso_empresa(conn, empresa_id, auth)
        result = conn.execute(
            text(
                """
            SELECT e.*, c.email AS contato_email, c.celular AS contato_celular, c.whatsapp AS contato_whatsapp
            FROM empresas e
            LEFT JOIN LATERAL (
                SELECT email, celular, whatsapp FROM contatos WHERE empresa_id = e.empresa_id
                ORDER BY decisor DESC NULLS LAST, data_criacao ASC NULLS LAST LIMIT 1
            ) c ON TRUE WHERE e.empresa_id = :id
        """
            ),
            {"id": empresa_id},
        ).fetchone()
    if not result:
        raise HTTPException(404, "Empresa não encontrada")
    return dict(result._mapping)


@app.get("/empresas/{empresa_id}/historico-status")
def historico_status_empresa(empresa_id: str, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        garantir_campos_pipeline(conn)
        checar_acesso_empresa(conn, empresa_id, auth)
        result = conn.execute(
            text("SELECT * FROM empresa_status_historico WHERE empresa_id = :id ORDER BY alterado_em DESC"),
            {"id": empresa_id},
        )
        return [dict(row._mapping) for row in result]


@app.get("/funil/transicoes")
def funil_transicoes(
    meses: int = 6,
    vendedor_id: str | None = None,
    segmento: str | None = None,
    auth: dict = Depends(exigir_gestor),
):
    """Taxa de passagem entre as etapas do funil, por coorte de entrada.

    ── Por que esta rota existe ────────────────────────────────────────────────
    A tela de Insights media desfecho (fechou/perdeu) e o retrato do funil hoje,
    mas nao a PASSAGEM: "quanto de Lead vira Proposta". O dado sempre esteve em
    `empresa_status_historico`; faltava le-lo em lote. `GET
    /empresas/{id}/historico-status` serve uma empresa so, e a tela precisaria de
    uma chamada por empresa da carteira.

    A agregacao mora em `funil.py`, sem banco e sem FastAPI, para poder ser
    testada (`python test_funil.py`) — este processo recusa subir contra
    producao fora do Railway e o `.env.dev` esta sem banco, entao logica de
    coorte escrita aqui dentro seria inverificavel. Esta funcao so faz o SQL.

    ── O que o historico NAO cobre ────────────────────────────────────────────
    `POST /empresas` grava a linha inicial (`status_anterior` NULL) e o `PUT`
    grava toda mudanca. Mas:

    - **`POST /empresas/rascunho` NAO grava historico** — de proposito: rascunho
      esta fora do funil ate alguem completar. A empresa entra na contagem na
      transicao `Rascunho -> Lead`, que e o momento certo.
    - Empresa cadastrada antes de o historico existir nao tem linha nenhuma e e
      invisivel aqui. `cobertura.sem_historico` diz quantas sao, para a tela
      avisar em vez de apresentar uma taxa sobre metade da base como se fosse
      sobre a base inteira.
    """
    meses = max(1, min(int(meses or 6), 24))
    inicio, fim = janela_meses(meses)

    with engine.begin() as conn:
        garantir_campos_pipeline(conn)
        trecho, params = filtro_escopo(conn, auth, prefixo="e.")

        # Os mesmos recortes da barra de filtro da tela. Vem DEPOIS do escopo e
        # so estreita o que ele ja permitiu: `filtro_escopo` continua sendo a
        # unica coisa que decide o que este usuario pode ver, e um vendedor_id
        # de fora do escopo devolve lista vazia, nao dado de outra equipe.
        if vendedor_id:
            trecho += " AND e.vendedor_id = CAST(:vend AS uuid)"
            params["vend"] = vendedor_id
        if segmento:
            trecho += " AND COALESCE(NULLIF(TRIM(e.segmento), ''), :sem_seg) = :seg"
            params["seg"] = segmento
            # Espelha o rotulo que o frontend usa para empresa sem segmento
            # (`segmentosDisponiveis` em utils/metricas.ts); sem isso, filtrar
            # por "Nao informado" na tela nao casaria com nada aqui.
            params["sem_seg"] = "Não informado"

        # Denominador honesto: quantas empresas do escopo existem e de quantas
        # da para contar a historia. Rascunho fica fora das duas contas — ele
        # ainda nao entrou no funil, entao nao e "cobertura faltando".
        cobertura = conn.execute(
            text(
                f"""
                SELECT COUNT(*) AS total,
                       COUNT(*) FILTER (WHERE EXISTS (
                           SELECT 1 FROM empresa_status_historico h
                           WHERE h.empresa_id = e.empresa_id
                       )) AS com_historico
                FROM empresas e
                WHERE e.conta_id = :cid AND COALESCE(e.status, '') <> 'Rascunho' {trecho}
            """
            ),
            params,
        ).fetchone()

        primeiro = conn.execute(
            text(
                f"""
                SELECT MIN(h.alterado_em)
                FROM empresa_status_historico h
                JOIN empresas e ON e.empresa_id = h.empresa_id
                WHERE e.conta_id = :cid {trecho}
            """
            ),
            params,
        ).scalar()

        # Uma linha por mudanca de status. O volume e da ordem de
        # (empresas x mudancas) e a agregacao e feita em Python de proposito: a
        # logica de coorte tem condicoes temporais encadeadas que em SQL virariam
        # tres window functions aninhadas — ilegiveis, piores de depurar e sem
        # ganho nenhum nesta escala.
        linhas = conn.execute(
            text(
                f"""
                SELECT h.empresa_id, h.status_anterior, h.status_novo, h.alterado_em
                FROM empresa_status_historico h
                JOIN empresas e ON e.empresa_id = h.empresa_id
                WHERE e.conta_id = :cid {trecho}
                ORDER BY h.empresa_id, h.alterado_em ASC
            """
            ),
            params,
        ).fetchall()

    resultado = agregar_funil(linhas, inicio, fim)
    total = cobertura.total or 0
    com_hist = cobertura.com_historico or 0
    resultado["janela"] = {"inicio": inicio.isoformat(), "fim": fim.isoformat(), "meses": meses}
    resultado["cobertura"] = {
        "empresas_no_escopo": total,
        "com_historico": com_hist,
        # O numero que a tela precisa mostrar junto da taxa: empresa cadastrada
        # antes de o historico existir nao aparece em etapa nenhuma, e uma taxa
        # sobre metade da base apresentada como se fosse sobre a base inteira e
        # pior do que nao ter a taxa.
        "sem_historico": max(total - com_hist, 0),
        "na_coorte": resultado["coorte"]["entraram"],
        "primeiro_registro": primeiro.isoformat() if primeiro else None,
    }
    return resultado


@app.get("/empresas/{empresa_id}/contatos")
def listar_contatos_por_empresa(empresa_id: str, auth: dict = Depends(get_auth)):
    with engine.connect() as conn:
        checar_acesso_empresa(conn, empresa_id, auth)
        result = conn.execute(
            text("SELECT * FROM contatos WHERE empresa_id = :id ORDER BY data_criacao ASC NULLS LAST"),
            {"id": empresa_id},
        )
        return [dict(row._mapping) for row in result]


@app.get("/empresas/{empresa_id}/produtos")
def produtos_da_empresa(empresa_id: str, auth: dict = Depends(get_auth)):
    """O que esta empresa ja comprou, consolidado por item.

    Sai dos itens dos orcamentos APROVADOS -- e o mais perto de "venda
    faturada" que o sistema tem, ja que nao existe tabela de pedidos. Rascunho,
    enviado e recusado ficam de fora: intencao nao e compra.

    Agrupa por `descricao` e nao por `equipamento_id` porque item avulso nao tem
    equipamento: agrupar pelo id jogaria todo avulso num balde NULL so.
    """
    with engine.begin() as conn:
        garantir_vendas(conn)
        checar_acesso_empresa(conn, empresa_id, auth)
        rows = conn.execute(
            text(
                """SELECT i.descricao AS nome,
                          SUM(i.quantidade)::int AS quantidade,
                          SUM(i.quantidade * i.preco_unitario) AS valor,
                          COUNT(DISTINCT o.orcamento_id)::int AS compras,
                          MAX(COALESCE(o.data_decisao, o.data_envio, o.criado_em)) AS ultima_compra
                   FROM orcamento_itens i
                   JOIN orcamentos o ON o.orcamento_id = i.orcamento_id
                   WHERE o.empresa_id = :eid AND o.status = 'aprovado'
                   GROUP BY i.descricao
                   ORDER BY valor DESC"""
            ),
            {"eid": empresa_id},
        )
        return [dict(r._mapping) for r in rows]


@app.get("/empresas/{empresa_id}/observacoes")
def listar_observacoes(empresa_id: str, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        garantir_observacoes(conn)
        checar_acesso_empresa(conn, empresa_id, auth)
        rows = conn.execute(
            text(
                """SELECT ob.*, u.nome AS autor_nome
                   FROM empresa_observacoes ob
                   LEFT JOIN usuarios u ON u.usuario_id = ob.autor_id
                   WHERE ob.empresa_id = :eid
                   ORDER BY ob.criado_em DESC"""
            ),
            {"eid": empresa_id},
        )
        return [dict(r._mapping) for r in rows]


@app.post("/empresas/{empresa_id}/observacoes")
def criar_observacao(empresa_id: str, dados: ObservacaoCreate, auth: dict = Depends(get_auth)):
    texto = (dados.texto or "").strip()
    if not texto:
        raise HTTPException(400, "A observação não pode ficar vazia.")
    with engine.begin() as conn:
        garantir_observacoes(conn)
        checar_acesso_empresa(conn, empresa_id, auth)
        row = conn.execute(
            text(
                """INSERT INTO empresa_observacoes (empresa_id, autor_id, texto, marcador)
                   VALUES (:eid, :aid, :txt, :mkr)
                   RETURNING observacao_id, empresa_id, autor_id, texto, marcador, criado_em"""
            ),
            {"eid": empresa_id, "aid": auth["usuario_id"], "txt": texto,
             "mkr": (dados.marcador or None)},
        ).fetchone()
        criada = dict(row._mapping)
        # Devolve ja com o nome do autor: a tela insere a nota na lista sem
        # precisar rebuscar tudo so para descobrir quem escreveu.
        criada["autor_nome"] = conn.execute(
            text("SELECT nome FROM usuarios WHERE usuario_id = :id"), {"id": auth["usuario_id"]}
        ).scalar()
        return criada


@app.delete("/empresas/{empresa_id}/observacoes/{observacao_id}")
def excluir_observacao(empresa_id: str, observacao_id: str, auth: dict = Depends(get_auth)):
    """Apaga uma anotacao. Autor apaga a sua; gerente apaga qualquer uma."""
    with engine.begin() as conn:
        garantir_observacoes(conn)
        checar_acesso_empresa(conn, empresa_id, auth)
        dono = conn.execute(
            text("SELECT autor_id FROM empresa_observacoes WHERE observacao_id = :id AND empresa_id = :eid"),
            {"id": observacao_id, "eid": empresa_id},
        ).fetchone()
        if not dono:
            raise HTTPException(404, "Observação não encontrada")
        if auth.get("role") != "gerente" and str(dono.autor_id) != str(auth["usuario_id"]):
            raise HTTPException(403, "Só o autor ou um gerente pode excluir esta observação.")
        conn.execute(
            text("DELETE FROM empresa_observacoes WHERE observacao_id = :id"), {"id": observacao_id}
        )
        return {"ok": True}


# Empresas sem coordenada mas com algum endereço aproveitável
_SQL_SEM_COORD = """
    (latitude IS NULL OR longitude IS NULL)
    AND COALESCE(NULLIF(TRIM(endereco), ''), NULLIF(TRIM(cidade), '')) IS NOT NULL
"""


_NUMERO_NO_FIM = re.compile(r",\s*(\d+[A-Za-z]?)\s*$")


def _tentativas_nominatim(r) -> list[dict]:
    """Consultas do mais preciso ao mais generico, para tentar em cascata.

    A consulta livre (`q=rua, numero, bairro, cidade, cep, Brasil`) nao resolve
    endereco brasileiro: o Nominatim devolve vazio ate para uma rua que ele tem
    cadastrada. Medido com "Rua Benedito Gomes do Nascimento, 212, Nova Morada,
    Almirante Tamandare, 83504-680" -- sem resultado na consulta livre, ponto
    exato na estruturada.

    O `postalcode` fica de fora de proposito: sozinho nao acha CEP brasileiro e,
    somado aos outros campos, zera o resultado.

    A cidade sozinha e o ultimo recurso. Para "empresas num raio de X km" o
    centro da cidade ja e util -- melhor que a empresa sumir do mapa.
    """
    rua = (r.endereco or "").strip()
    cidade = (r.cidade or "").strip()
    numero = (getattr(r, "numero", None) or "").strip()

    # Cadastro antigo guardava tudo junto ("Rua X, 212"). Com a coluna `numero`
    # vazia, ainda vale tentar separar do proprio texto da rua.
    if not numero:
        m = _NUMERO_NO_FIM.search(rua)
        if m:
            numero = m.group(1)
            rua = rua[: m.start()].strip()

    tentativas: list[dict] = []

    if rua and cidade:
        if numero:
            # O Nominatim espera o numero na frente: "212 Rua X".
            tentativas.append({"street": f"{numero} {rua}", "city": cidade, "country": "Brasil"})
        # sem o numero: se o predio nao existe no OSM, a rua costuma existir
        tentativas.append({"street": rua, "city": cidade, "country": "Brasil"})

    if cidade:
        tentativas.append({"city": cidade, "country": "Brasil"})

    return tentativas


# User-Agent proprio e exigencia da politica do Nominatim -- cliente sem
# identificacao e bloqueado. Fica numa constante porque agora sao dois
# consumidores: o backfill em lote e a busca avulsa do planejador de rota.
NOMINATIM_UA = "CRM-Prospeccao/1.0 (https://frontend-crm-xi-plum.vercel.app)"

# Janela minima entre chamadas ao Nominatim, por processo. A politica do OSM e
# de no maximo 1 req/seg para o servico publico; o backfill respeita isso com
# sleep entre os itens do lote, e aqui o controle e por relogio porque a busca
# avulsa e disparada por gente digitando, sem lote nenhum.
_NOMINATIM_INTERVALO = 1.1
_nominatim_ultima_chamada = 0.0
_nominatim_trava = threading.Lock()


@app.get("/geo/buscar")
async def buscar_endereco(q: str = "", auth: dict = Depends(get_auth)):
    """Endereco digitado -> coordenada, para o planejador de rota.

    Custo zero: Nominatim/OpenStreetMap, o mesmo servico que ja faz o backfill
    de coordenadas das empresas. Nao encosta na cota do Google Places.

    Passa pelo backend em vez de ir direto do navegador para manter num lugar
    so o User-Agent exigido pela politica do OSM e o limite de 1 req/seg -- do
    cliente, cada aba abriria sua propria torneira.
    """
    termo = (q or "").strip()
    if len(termo) < 4:
        raise HTTPException(400, "Digite ao menos 4 caracteres")

    # Cache antes da fila: endereco nao muda de lugar, e sem isto a mesma busca
    # repetida paga 1s de espera de novo. Reaproveita o cache do proxy do OSRM
    # -- e o mesmo problema (servico comunitario, 1 req/s) e nao ha razao para
    # duas implementacoes de cache no arquivo.
    chave_busca = "nominatim|" + termo.lower()
    em_cache = _osrm_cache_ler(chave_busca)
    if em_cache is not None:
        return em_cache

    global _nominatim_ultima_chamada
    with _nominatim_trava:
        espera = _NOMINATIM_INTERVALO - (time.time() - _nominatim_ultima_chamada)
        _nominatim_ultima_chamada = time.time() + max(0.0, espera)
    if espera > 0:
        await asyncio.sleep(espera)

    try:
        async with httpx.AsyncClient(timeout=12.0, headers={"User-Agent": NOMINATIM_UA}) as client:
            resp = await client.get(
                "https://nominatim.openstreetmap.org/search",
                params={"format": "json", "limit": 1, "countrycodes": "br",
                        "addressdetails": 1, "q": termo},
            )
        dados = resp.json() if resp.status_code == 200 else []
    except Exception as e:  # noqa: BLE001 - servico publico: indisponivel nao e erro nosso
        print(f"[NOMINATIM] falha ao buscar {termo!r}: {e}")
        raise HTTPException(502, "Servico de endereco indisponivel no momento")

    if not dados:
        # 200 com achado=False, e nao 404: "nao encontrei este endereco" e uma
        # resposta valida da busca, nao um erro de rota. Vai para o cache pelo
        # mesmo motivo: quem digitou errado costuma tentar de novo igual.
        vazio = {"achado": False, "lat": None, "lon": None, "endereco": None}
        _osrm_cache_gravar(chave_busca, vazio)
        return vazio

    primeiro = dados[0]
    try:
        achado = {
            "achado": True,
            "lat": float(primeiro["lat"]),
            "lon": float(primeiro["lon"]),
            "endereco": primeiro.get("display_name") or termo,
        }
        _osrm_cache_gravar(chave_busca, achado)
        return achado
    except (KeyError, ValueError, TypeError):
        return {"achado": False, "lat": None, "lon": None, "endereco": None}




# =========================================================================
# PROXY DO OSRM -- rota e matriz de distancias para o planejador de rota
# =========================================================================
#
# Por que passar pelo backend em vez de chamar o OSRM do navegador (que e como
# isso nasceu): o demo publico do OSRM pede no maximo 1 req/s e bloqueia por
# ENDERECO IP, nao por aba nem por tela. Throttle no cliente nao e garantia
# nenhuma -- duas abas abertas ja sao 2 req/s, e o planejador dispara rajada.
# Com tudo saindo daqui existe uma torneira so, e ela e de verdade.
#
# Mesmo padrao do /geo/buscar (Nominatim), com tres diferencas:
#   - a fila reserva o horario e dorme FORA da trava, senao uma requisicao
#     lenta seguraria todas as outras enquanto espera resposta do OSRM;
#   - ha teto de espera: com fila grande, negar rapido e melhor que deixar o
#     usuario olhando para um spinner por meio minuto;
#   - ha cache, porque rua nao muda de lugar e o planejador repete muito a
#     mesma consulta (mexer no raio, tirar e repor a mesma parada).

# Base do OSRM. Configuravel para o dia em que o demo publico bloquear ou sair
# do ar: apontar para uma instancia propria vira preencher uma variavel no
# Railway e reiniciar, em vez de um commit e um deploy feitos as pressas.
#
# O demo publico e a politica dele -- 1 req/s, nao comercial, "nao recomendado
# para producao" -- sao o padrao apenas enquanto nao ha instancia propria. Uma
# instancia auto-hospedada fala a MESMA API, entao esta linha e a unica coisa
# que muda: /route e /table continuam identicos, e nada mais no proxy precisa
# ser tocado. Foi por isso que se manteve OSRM em vez de trocar de motor.
OSRM_BASE = os.getenv("OSRM_BASE", "https://router.project-osrm.org").rstrip("/")

# 1 req/s e a politica; a folga de 5% cobre a imprecisao do relogio.
_OSRM_INTERVALO = 1.05
# Horario (relogio monotonico) em que a proxima chamada pode sair.
_osrm_proxima_vaga = 0.0
_osrm_trava = asyncio.Lock()

# Acima disto a fila ja esta tao longa que a resposta chegaria tarde demais
# para ser util. 8s ~= 8 pedidos enfileirados.
_OSRM_ESPERA_MAXIMA = 8.0

# Teto do proprio OSRM demo, medido contra o servidor: 100 coordenadas passam,
# 120 voltam code TooBig. O frontend ja corta antes, mas quem publica a rota
# nao pode confiar em validacao que mora no cliente.
OSRM_MAX_COORD_MATRIZ = 100
# Rota desenhada: origem + 5 paradas + destino = 7. O teto folgado cobre a
# previa do hover sem virar porta para pedido absurdo.
OSRM_MAX_COORD_ROTA = 25

# Cache em memoria do processo. Geometria de rua nao muda em horas, entao TTL
# longo. O limite de tamanho existe para isto nao virar vazamento de memoria
# num processo que fica semanas de pe.
_OSRM_CACHE_TTL = 6 * 3600
_OSRM_CACHE_MAX = 400
_osrm_cache = {}


def _osrm_cache_ler(chave: str):
    item = _osrm_cache.get(chave)
    if not item:
        return None
    nascido, valor = item
    if time.monotonic() - nascido > _OSRM_CACHE_TTL:
        _osrm_cache.pop(chave, None)
        return None
    return valor


def _osrm_cache_gravar(chave: str, valor: dict) -> None:
    # Descarte burro (os mais antigos pela ordem de insercao do dict) em vez de
    # LRU: com 400 entradas e um punhado de vendedores, LRU seria complexidade
    # sem ganho observavel.
    if len(_osrm_cache) >= _OSRM_CACHE_MAX:
        for velho in list(_osrm_cache)[: _OSRM_CACHE_MAX // 4]:
            _osrm_cache.pop(velho, None)
    _osrm_cache[chave] = (time.monotonic(), valor)


async def _osrm_aguardar_vez() -> None:
    """Segura a chamada ate a vaga dela na fila de 1 req/s.

    A vaga e reservada dentro da trava e o sono acontece fora: manter a trava
    durante o sleep serializaria as ESPERAS, e ai N pedidos custariam
    N*intervalo de latencia acumulada em vez de se distribuirem no tempo.
    """
    global _osrm_proxima_vaga
    async with _osrm_trava:
        agora = time.monotonic()
        vaga = max(agora, _osrm_proxima_vaga)
        espera = vaga - agora
        if espera > _OSRM_ESPERA_MAXIMA:
            raise HTTPException(503, "Servico de rotas congestionado. Tente de novo em instantes.")
        _osrm_proxima_vaga = vaga + _OSRM_INTERVALO
    if espera > 0:
        await asyncio.sleep(espera)


def _osrm_pontos(bruto: str, maximo: int):
    """Converte "lng,lat;lng,lat;..." em lista validada. Formato ruim e 400.

    Validar aqui nao e paranoia com o nosso proprio frontend: e o que impede a
    rota de virar repassadora de string arbitraria para dentro da URL do OSRM.
    """
    pontos = []
    for parte in (bruto or "").split(";"):
        parte = parte.strip()
        if not parte:
            continue
        pedacos = parte.split(",")
        if len(pedacos) != 2:
            raise HTTPException(400, "Ponto invalido: use lng,lat")
        try:
            lng, lat = float(pedacos[0]), float(pedacos[1])
        except ValueError:
            raise HTTPException(400, "Coordenada nao numerica")
        if not (-180 <= lng <= 180) or not (-90 <= lat <= 90):
            raise HTTPException(400, "Coordenada fora do intervalo valido")
        pontos.append((lng, lat))
    if len(pontos) < 2:
        raise HTTPException(400, "Sao necessarios ao menos dois pontos")
    if len(pontos) > maximo:
        raise HTTPException(400, f"No maximo {maximo} pontos por chamada")
    return pontos


def _osrm_chave(pontos, sufixo: str) -> str:
    # 5 casas ~= 1 metro. Arredondar e o que faz o cache acertar: a mesma
    # empresa clicada duas vezes traz float identico, mas GPS e hover trazem
    # ruido nas ultimas casas, que geraria chave nova a cada passada.
    return sufixo + "|" + ";".join(f"{lng:.5f},{lat:.5f}" for lng, lat in pontos)


async def _osrm_get(caminho: str, params: dict) -> dict:
    """Chamada crua ao OSRM, ja enfileirada. Levanta HTTPException em falha.

    IMPORTANTE: nada aqui pode escapar como excecao nao tratada. Um 500 do
    FastAPI sai SEM os cabecalhos de CORS, e o navegador reporta isso como erro
    de CORS -- mandando quem for depurar para o middleware errado.
    """
    await _osrm_aguardar_vez()
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(f"{OSRM_BASE}{caminho}", params=params)
    except Exception as e:  # noqa: BLE001 - servidor publico sem SLA
        print(f"[OSRM] falha de rede em {caminho}: {e}")
        raise HTTPException(502, "Servico de rotas indisponivel no momento")

    if resp.status_code != 200:
        print(f"[OSRM] status {resp.status_code} em {caminho}")
        # 429 do OSRM significa que a nossa fila nao foi suficiente. Log
        # separado porque e o sintoma de que o processo deixou de ser unico.
        if resp.status_code == 429:
            print("[OSRM] ATENCAO: 429 mesmo com fila local. Ha mais de um worker saindo por este IP?")
        raise HTTPException(502, "Servico de rotas indisponivel no momento")

    try:
        dados = resp.json()
    except Exception:  # noqa: BLE001
        raise HTTPException(502, "Resposta invalida do servico de rotas")

    if dados.get("code") != "Ok":
        # NoRoute e resposta legitima (ilha, ponto no meio do mar), nao falha de
        # infraestrutura: 422 para o frontend distinguir e nao oferecer "tentar
        # de novo" para algo que nunca vai dar certo.
        if dados.get("code") in ("NoRoute", "NoSegment", "NoTable"):
            raise HTTPException(422, "Nao ha rota por ruas entre estes pontos")
        print(f"[OSRM] code={dados.get('code')} em {caminho}")
        raise HTTPException(502, "Servico de rotas indisponivel no momento")
    return dados


@app.get("/geo/rota")
async def geo_rota(
    pontos: str = "",
    overview: str = "full",
    _email: str = Depends(get_current_user),
):
    """Rota viaria passando pelos pontos na ordem dada.

    Devolve a geometria ja em [lat, lng] -- ordem do Leaflet, inversa a do
    GeoJSON que o OSRM manda. A conversao mora aqui para o cache guardar o
    formato final e nao repetir a transformacao a cada consumidor.

    Autenticacao por `get_current_user` (so o JWT) e nao por `get_auth`: esta
    rota nao filtra dado de carteira nenhum, e `get_auth` custa uma ida ao banco
    que se pagaria a cada traco de rota desenhado.
    """
    if overview not in ("full", "simplified", "false"):
        raise HTTPException(400, "overview invalido")
    lista = _osrm_pontos(pontos, OSRM_MAX_COORD_ROTA)

    chave = _osrm_chave(lista, f"rota:{overview}")
    em_cache = _osrm_cache_ler(chave)
    if em_cache is not None:
        return em_cache

    caminho = "/route/v1/driving/" + ";".join(f"{lng},{lat}" for lng, lat in lista)
    dados = await _osrm_get(caminho, {"overview": overview, "geometries": "geojson"})

    rotas = dados.get("routes") or []
    if not rotas:
        raise HTTPException(422, "Nao ha rota por ruas entre estes pontos")
    r = rotas[0]
    resultado = {
        "coords": [[c[1], c[0]] for c in (r.get("geometry") or {}).get("coordinates", [])],
        "km": (r.get("distance") or 0) / 1000.0,
        "min": (r.get("duration") or 0) / 60.0,
    }
    _osrm_cache_gravar(chave, resultado)
    return resultado


@app.get("/geo/matriz")
async def geo_matriz(
    pontos: str = "",
    _email: str = Depends(get_current_user),
):
    """Matriz de distancia/tempo entre todos os pontos, numa chamada so.

    E o que substitui o padrao antigo do planejador: uma chamada de /route por
    candidato (ate 12 em rajada) so para descobrir o custo do desvio de cada um.
    Com a matriz, o desvio de qualquer candidato -- e de qualquer posicao de
    insercao dele na rota -- sai de aritmetica local, sem tocar a rede de novo.

    `distancias_km[i][j]` e a distancia de i ate j. Nao e simetrica: mao unica
    faz a ida diferir da volta.
    """
    lista = _osrm_pontos(pontos, OSRM_MAX_COORD_MATRIZ)

    chave = _osrm_chave(lista, "matriz")
    em_cache = _osrm_cache_ler(chave)
    if em_cache is not None:
        return em_cache

    caminho = "/table/v1/driving/" + ";".join(f"{lng},{lat}" for lng, lat in lista)
    dados = await _osrm_get(caminho, {"annotations": "distance,duration"})

    distancias = dados.get("distances") or []
    duracoes = dados.get("durations") or []
    if not distancias or not duracoes:
        raise HTTPException(502, "Matriz incompleta do servico de rotas")

    # None aparece quando o OSRM nao acha caminho entre um par especifico. Vira
    # None no JSON tambem, e nao 0: zero seria lido como "coladinho" e o
    # candidato inalcancavel apareceria como o melhor de todos.
    def _km(v):
        return None if v is None else v / 1000.0

    def _minutos(v):
        return None if v is None else v / 60.0

    resultado = {
        "distancias_km": [[_km(v) for v in linha] for linha in distancias],
        "duracoes_min": [[_minutos(v) for v in linha] for linha in duracoes],
    }
    _osrm_cache_gravar(chave, resultado)
    return resultado


@app.post("/empresas/geocodificar")
async def geocodificar_empresas(limite: int = 15, usuario_email: str = Depends(get_current_user)):
    """Backfill de coordenadas (custo zero) via Nominatim/OpenStreetMap a partir do
    endereço já salvo. Processa em lote pequeno; o frontend chama em loop até
    'restantes' == 0. Respeita a política do OSM: <=1 req/seg e User-Agent próprio."""
    limite = max(1, min(limite, 30))
    with engine.connect() as conn:
        garantir_colunas_places(conn)
        rows = conn.execute(
            text(
                f"""
                SELECT empresa_id, endereco, numero, bairro, cidade, cep
                FROM empresas
                WHERE {_SQL_SEM_COORD}
                ORDER BY status_atualizado_em DESC NULLS LAST
                LIMIT :lim
            """
            ),
            {"lim": limite},
        ).fetchall()

    geocodificadas = 0
    falharam = 0
    headers = {"User-Agent": NOMINATIM_UA}
    async with httpx.AsyncClient(timeout=15.0, headers=headers) as client:
        primeira = True
        for r in rows:
            coord = None
            for params in _tentativas_nominatim(r):
                if not primeira:
                    await asyncio.sleep(1.1)  # política do Nominatim: <=1 req/seg
                primeira = False
                try:
                    resp = await client.get(
                        "https://nominatim.openstreetmap.org/search",
                        params={"format": "json", "limit": 1, "countrycodes": "br", **params},
                    )
                    data = resp.json() if resp.status_code == 200 else []
                except Exception:
                    data = []
                if data:
                    try:
                        coord = (float(data[0]["lat"]), float(data[0]["lon"]))
                        break
                    except (KeyError, ValueError, TypeError):
                        continue
            if coord is None:
                falharam += 1
                continue
            with engine.begin() as conn:
                conn.execute(
                    text("UPDATE empresas SET latitude = :lat, longitude = :lng WHERE empresa_id = :id"),
                    {"lat": coord[0], "lng": coord[1], "id": r.empresa_id},
                )
            geocodificadas += 1

    with engine.connect() as conn:
        restantes = conn.execute(text(f"SELECT COUNT(*) FROM empresas WHERE {_SQL_SEM_COORD}")).scalar()

    return {
        "processadas": len(rows),
        "geocodificadas": geocodificadas,
        "falharam": falharam,
        "restantes": restantes,
    }


@app.get("/empresas/{empresa_id}/google-refresh")
async def refresh_google_empresa(empresa_id: str, auth: dict = Depends(get_auth)):
    """Re-busca o snapshot volátil do Google (rating/contagem/status) usando o
    google_place_id já persistido na empresa. Usa Place Details by ID com field
    mask enxuto — mais barato que o text search, mas conta no mesmo teto mensal
    do usuário (ver reservar_chamada_places)."""
    if not GOOGLE_PLACES_API_KEY:
        raise HTTPException(503, "Google Places API não configurada")

    with engine.connect() as conn:
        garantir_colunas_places(conn)
        row = conn.execute(
            text("SELECT google_place_id FROM empresas WHERE empresa_id = :id"),
            {"id": empresa_id},
        ).fetchone()
    if not row:
        raise HTTPException(404, "Empresa não encontrada")
    place_id = row[0]
    if not place_id:
        raise HTTPException(422, "Empresa sem google_place_id — não foi importada do Google.")

    api_headers = {
        "X-Goog-Api-Key": GOOGLE_PLACES_API_KEY,
        "X-Goog-FieldMask": "id,rating,userRatingCount,businessStatus",
    }
    # Details by ID e mais barato que o text search, mas nao e de graca: entra
    # no mesmo teto mensal do usuario, senao o limite da busca seria contornavel.
    chave_cota = _chave_cota_places(auth)
    reservar_chamada_places(auth)
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.get(
                f"https://places.googleapis.com/v1/places/{place_id}",
                headers=api_headers,
                params={"languageCode": "pt-BR", "regionCode": "BR"},
            )
    except httpx.HTTPError as e:
        devolver_chamada_places(chave_cota)
        raise HTTPException(502, f"Google Places indisponivel: {e}")

    if resp.status_code != 200:
        devolver_chamada_places(chave_cota)

    if resp.status_code == 429 or (
        resp.status_code != 200 and
        any(k in resp.text.upper() for k in ("RESOURCE_EXHAUSTED", "QUOTA", "RATE_LIMIT"))
    ):
        raise HTTPException(429, "Cota da Google Places API esgotada. Tente mais tarde.")
    if resp.status_code != 200:
        raise HTTPException(502, f"Google Places erro: {resp.text}")

    p = resp.json()
    rating = p.get("rating")
    rating_count = p.get("userRatingCount")
    business_status = p.get("businessStatus")

    with engine.begin() as conn:
        updated = conn.execute(
            text(
                """
                UPDATE empresas
                SET google_rating = :rating,
                    google_rating_count = :rating_count,
                    business_status = :business_status,
                    google_synced_at = NOW()
                WHERE empresa_id = :id
                RETURNING google_synced_at
            """
            ),
            {
                "id": empresa_id,
                "rating": rating,
                "rating_count": rating_count,
                "business_status": business_status,
            },
        ).fetchone()

    synced_at = updated[0] if updated else None
    return {
        "google_rating": rating,
        "google_rating_count": rating_count,
        "business_status": business_status,
        "google_synced_at": synced_at.isoformat() if synced_at else None,
    }


@app.post("/empresas")
def criar_empresa(empresa: EmpresaCreate, auth: dict = Depends(get_auth)):
    empresa_id = str(uuid.uuid4())
    segmento = None
    is_rascunho = (empresa.status or "").lower() == "rascunho"
    if empresa.segmento and not is_rascunho:
        segmento = limpar_segmento(empresa.segmento)
        if not segmento_valido(segmento):
            raise HTTPException(400, "Segmento nao reconhecido.")
    elif empresa.segmento and is_rascunho:
        segmento = limpar_segmento(empresa.segmento) if empresa.segmento.strip() else None
    with engine.begin() as conn:
        garantir_campos_pipeline(conn)
        garantir_colunas_places(conn)
        if segmento:
            segmento = salvar_segmento(conn, segmento)
        conn.execute(
            text(
                """
            INSERT INTO empresas (empresa_id, nome, segmento, porte, cidade, endereco, numero, cep, bairro, regiao,
                observacoes, cnpj, site, linkedin_empresa, responsavel_principal,
                status, origem_lead, ultima_interacao, proxima_acao, data_proxima_acao, status_atualizado_em,
                motivo_perdido, temperatura, logo_url, criado_em, conta_id, vendedor_id,
                google_place_id, latitude, longitude, google_rating, google_rating_count, business_status, google_synced_at)
            VALUES (:id, :nome, :segmento, :porte, :cidade, :endereco, :numero, :cep, :bairro, :regiao,
                :observacoes, :cnpj, :site, :linkedin_empresa, :responsavel_principal,
                :status, :origem_lead, :ultima_interacao, :proxima_acao, :data_proxima_acao, NOW(),
                :motivo_perdido, :temperatura, :logo_url, NOW(), :conta_id, :vendedor_id,
                :google_place_id, :latitude, :longitude, :google_rating, :google_rating_count, :business_status, :google_synced_at)
        """
            ),
            {
                "id": empresa_id,
                "nome": empresa.nome,
                "segmento": segmento,
                "porte": empresa.porte,
                "cidade": empresa.cidade,
                "endereco": empresa.endereco,
                "numero": empresa.numero,
                "cep": empresa.cep,
                "bairro": empresa.bairro,
                "regiao": empresa.regiao,
                "observacoes": empresa.observacoes,
                "cnpj": empresa.cnpj,
                "site": empresa.site,
                "linkedin_empresa": empresa.linkedin_empresa,
                "responsavel_principal": empresa.responsavel_principal or auth["email"],
                "conta_id": auth["conta_id"],
                "vendedor_id": auth["usuario_id"],
                "status": empresa.status or "Lead",
                "origem_lead": empresa.origem_lead,
                "ultima_interacao": empresa.ultima_interacao or datetime.utcnow(),
                "proxima_acao": empresa.proxima_acao,
                "data_proxima_acao": empresa.data_proxima_acao,
                "motivo_perdido": empresa.motivo_perdido,
                "temperatura": empresa.temperatura,
                "logo_url": empresa.logo_url,
                "google_place_id": empresa.google_place_id,
                "latitude": empresa.latitude,
                "longitude": empresa.longitude,
                "google_rating": empresa.google_rating,
                "google_rating_count": empresa.google_rating_count,
                "business_status": empresa.business_status,
                "google_synced_at": empresa.google_synced_at,
            },
        )
        conn.execute(
            text(
                """
            INSERT INTO empresa_status_historico (historico_id, empresa_id, status_anterior, status_novo, observacao, alterado_em)
            VALUES (:id, :empresa_id, NULL, :status_novo, :observacao, NOW())
        """
            ),
            {
                "id": str(uuid.uuid4()),
                "empresa_id": empresa_id,
                "status_novo": empresa.status or "Lead",
                "observacao": "Rascunho salvo" if is_rascunho else "Cadastro inicial",
            },
        )
    return {"msg": "Empresa criada com sucesso 🚀", "empresa_id": empresa_id, "id": empresa_id}


@app.put("/empresas/{empresa_id}")
def atualizar_empresa(empresa_id: str, empresa: EmpresaUpdate, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        garantir_campos_pipeline(conn)
        checar_acesso_empresa(conn, empresa_id, auth)
        result = conn.execute(text("SELECT empresa_id, status FROM empresas WHERE empresa_id = :id"), {"id": empresa_id}).fetchone()
        if not result:
            raise HTTPException(404, "Empresa não encontrada")
        status_anterior = result._mapping.get("status")
        status_mudou = empresa.status is not None and empresa.status != status_anterior
        conn.execute(
            text(
                """
            UPDATE empresas SET nome=COALESCE(:nome,nome), segmento=COALESCE(:segmento,segmento),
                porte=COALESCE(:porte,porte), cidade=COALESCE(:cidade,cidade), endereco=COALESCE(:endereco,endereco),
                numero=COALESCE(:numero,numero),
                cep=COALESCE(:cep,cep), bairro=COALESCE(:bairro,bairro), regiao=COALESCE(:regiao,regiao),
                observacoes=COALESCE(:observacoes,observacoes), cnpj=COALESCE(:cnpj,cnpj), site=COALESCE(:site,site),
                linkedin_empresa=COALESCE(:linkedin_empresa,linkedin_empresa),
                responsavel_principal=COALESCE(:responsavel_principal,responsavel_principal),
                status=COALESCE(:status,status), status_cadastro=COALESCE(:status_cadastro,status_cadastro),
                origem_lead=COALESCE(:origem_lead,origem_lead),
                ultima_interacao=COALESCE(:ultima_interacao,ultima_interacao),
                proxima_acao=COALESCE(:proxima_acao,proxima_acao), data_proxima_acao=:data_proxima_acao,
                status_atualizado_em=CASE WHEN :status IS NOT NULL AND :status<>status THEN NOW() ELSE status_atualizado_em END,
                motivo_perdido=CASE WHEN :status IS NOT NULL AND :status<>'Perdido' THEN NULL ELSE COALESCE(:motivo_perdido,motivo_perdido) END,
                temperatura=COALESCE(:temperatura,temperatura),
                logo_url=COALESCE(:logo_url,logo_url)
            WHERE empresa_id=:id
        """
            ),
            {
                "id": empresa_id,
                "nome": empresa.nome,
                "segmento": empresa.segmento,
                "porte": empresa.porte,
                "cidade": empresa.cidade,
                "endereco": empresa.endereco,
                "numero": empresa.numero,
                "cep": empresa.cep,
                "bairro": empresa.bairro,
                "regiao": empresa.regiao,
                "observacoes": empresa.observacoes,
                "cnpj": empresa.cnpj,
                "site": empresa.site,
                "linkedin_empresa": empresa.linkedin_empresa,
                "responsavel_principal": empresa.responsavel_principal,
                "status": empresa.status,
                "status_cadastro": empresa.status_cadastro,
                "origem_lead": empresa.origem_lead,
                "ultima_interacao": empresa.ultima_interacao,
                "proxima_acao": empresa.proxima_acao,
                "data_proxima_acao": empresa.data_proxima_acao,
                "motivo_perdido": empresa.motivo_perdido,
                "temperatura": empresa.temperatura,
                "logo_url": empresa.logo_url,
            },
        )
        if status_mudou:
            conn.execute(
                text(
                    """
                INSERT INTO empresa_status_historico (historico_id, empresa_id, status_anterior, status_novo, observacao, alterado_em)
                VALUES (:id, :empresa_id, :status_anterior, :status_novo, :observacao, NOW())
            """
                ),
                {
                    "id": str(uuid.uuid4()),
                    "empresa_id": empresa_id,
                    "status_anterior": status_anterior,
                    "status_novo": empresa.status,
                    "observacao": empresa.motivo_perdido if empresa.status == "Perdido" else None,
                },
            )
    return {"msg": "Empresa atualizada com sucesso 🚀"}


@app.delete("/empresas/{empresa_id}")
def deletar_empresa(empresa_id: str, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        checar_acesso_empresa(conn, empresa_id, auth)
        conn.execute(text("DELETE FROM contatos WHERE empresa_id = :id"), {"id": empresa_id})
        conn.execute(text("DELETE FROM empresa_status_historico WHERE empresa_id = :id"), {"id": empresa_id})
        result = conn.execute(text("DELETE FROM empresas WHERE empresa_id = :id RETURNING empresa_id"), {"id": empresa_id}).fetchone()
    if not result:
        raise HTTPException(404, "Empresa não encontrada")
    return {"msg": "Empresa deletada com sucesso"}


# =========================
# CONTATOS
# =========================
@app.get("/contatos/{empresa_id}")
def listar_contatos_empresa(empresa_id: str, auth: dict = Depends(get_auth)):
    with engine.connect() as conn:
        checar_acesso_empresa(conn, empresa_id, auth)
        result = conn.execute(
            text("SELECT * FROM contatos WHERE empresa_id = :id ORDER BY data_criacao ASC NULLS LAST"),
            {"id": empresa_id},
        )
        return [dict(row._mapping) for row in result]


@app.post("/contatos")
def criar_contato(contato: dict, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        checar_acesso_empresa(conn, contato.get("empresa_id"), auth)
        conn.execute(
            text(
                """
            INSERT INTO contatos (contato_id, empresa_id, nome, funcao, email, celular, observacoes,
                prioridade, whatsapp, linkedin, nivel_influencia, decisor, data_ultimo_contato, canal_preferido)
            VALUES (:id, :empresa_id, :nome, :funcao, :email, :celular, :observacoes,
                :prioridade, :whatsapp, :linkedin, :nivel_influencia, :decisor, :data_ultimo_contato, :canal_preferido)
        """
            ),
            {
                "id": str(uuid.uuid4()),
                "empresa_id": contato.get("empresa_id"),
                "nome": contato.get("nome"),
                "funcao": contato.get("funcao"),
                "email": contato.get("email"),
                "celular": contato.get("celular"),
                "observacoes": contato.get("observacoes"),
                "prioridade": contato.get("prioridade"),
                "whatsapp": contato.get("whatsapp"),
                "linkedin": contato.get("linkedin"),
                "nivel_influencia": contato.get("nivel_influencia"),
                "decisor": contato.get("decisor"),
                "data_ultimo_contato": contato.get("data_ultimo_contato"),
                "canal_preferido": contato.get("canal_preferido"),
            },
        )
    return {"msg": "Contato criado com sucesso 🚀"}


@app.put("/contatos/{contato_id}")
def atualizar_contato(contato_id: str, contato: ContatoUpdate, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        result = conn.execute(text("SELECT empresa_id FROM contatos WHERE contato_id = :id"), {"id": contato_id}).fetchone()
        if not result:
            raise HTTPException(404, "Contato não encontrado")
        checar_acesso_empresa(conn, str(result.empresa_id), auth)
        conn.execute(
            text(
                """
            UPDATE contatos SET nome=COALESCE(:nome,nome), funcao=COALESCE(:funcao,funcao),
                email=COALESCE(:email,email), celular=COALESCE(:celular,celular),
                whatsapp=COALESCE(:whatsapp,whatsapp), linkedin=COALESCE(:linkedin,linkedin),
                observacoes=COALESCE(:observacoes,observacoes), prioridade=COALESCE(:prioridade,prioridade),
                nivel_influencia=COALESCE(:nivel_influencia,nivel_influencia), decisor=COALESCE(:decisor,decisor),
                canal_preferido=COALESCE(:canal_preferido,canal_preferido),
                data_ultimo_contato=COALESCE(:data_ultimo_contato,data_ultimo_contato)
            WHERE contato_id=:id
        """
            ),
            {
                "id": contato_id,
                "nome": contato.nome,
                "funcao": contato.funcao,
                "email": contato.email,
                "celular": contato.celular,
                "whatsapp": contato.whatsapp,
                "linkedin": contato.linkedin,
                "observacoes": contato.observacoes,
                "prioridade": contato.prioridade,
                "nivel_influencia": contato.nivel_influencia,
                "decisor": contato.decisor,
                "canal_preferido": contato.canal_preferido,
                "data_ultimo_contato": contato.data_ultimo_contato,
            },
        )
    return {"msg": "Contato atualizado com sucesso 🚀"}


@app.delete("/contatos/{contato_id}")
def deletar_contato(contato_id: str, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        alvo = conn.execute(text("SELECT empresa_id FROM contatos WHERE contato_id = :id"), {"id": contato_id}).fetchone()
        if not alvo:
            raise HTTPException(404, "Contato não encontrado")
        checar_acesso_empresa(conn, str(alvo.empresa_id), auth)
        conn.execute(text("DELETE FROM contatos WHERE contato_id = :id"), {"id": contato_id})
    return {"msg": "Contato deletado com sucesso"}


# =========================
# USUÁRIOS
# =========================
def _validar_supervisor(conn, supervisor_id: str, alvo_id: str | None, conta_id: str) -> str:
    """Regras de integridade do vínculo Vendedor → Supervisor, todas checadas
    aqui no servidor (o frontend só esconde botão, não autoriza nada):
    - o supervisor precisa existir e ser da MESMA conta;
    - precisa realmente ter a função 'supervisor';
    - ninguém pode ser supervisor de si mesmo."""
    if alvo_id and str(supervisor_id) == str(alvo_id):
        raise HTTPException(400, "Um usuário não pode ser supervisor de si mesmo")
    sup = conn.execute(
        text("SELECT usuario_id, conta_id, role FROM usuarios WHERE usuario_id = :id"),
        {"id": supervisor_id},
    ).fetchone()
    if not sup or str(sup.conta_id) != str(conta_id):
        raise HTTPException(404, "Supervisor não encontrado nesta conta")
    if normalizar_role(sup.role) != "supervisor":
        raise HTTPException(400, "O usuário escolhido não tem a função de Supervisor")
    return str(sup.usuario_id)


@app.get("/usuarios")
def listar_usuarios(auth: dict = Depends(exigir_gestor)):
    """Lista de usuários com função, status, supervisor e resumo da carteira.

    Gerente vê a conta inteira. Supervisor vê apenas ele mesmo e os vendedores
    atribuídos a ele — nunca os vendedores de outro supervisor."""
    with engine.connect() as conn:
        ids = escopo_vendedores(conn, auth)
        filtro = "" if ids is None else "AND u.usuario_id = ANY(CAST(:ids AS uuid[]))"
        params = {"cid": auth["conta_id"]}
        if ids is not None:
            params["ids"] = ids
        rows = conn.execute(
            text(
                f"""
                SELECT u.usuario_id, u.nome, u.email, u.telefone, u.role, u.ativo, u.data_criacao,
                       u.supervisor_id, s.nome AS supervisor_nome,
                       COUNT(e.empresa_id) AS total_empresas
                FROM usuarios u
                LEFT JOIN usuarios s ON s.usuario_id = u.supervisor_id
                LEFT JOIN empresas e ON e.vendedor_id = u.usuario_id
                WHERE u.conta_id = :cid {filtro}
                GROUP BY u.usuario_id, u.nome, u.email, u.telefone, u.role, u.ativo,
                         u.data_criacao, u.supervisor_id, s.nome
                ORDER BY u.role DESC, u.nome ASC
            """
            ),
            params,
        )
        return [dict(r._mapping) for r in rows]


@app.get("/equipe/estrutura")
def estrutura_equipe(auth: dict = Depends(exigir_gestor)):
    """Organograma da conta: gerentes, supervisores com seus vendedores e os
    vendedores ainda sem supervisor.

    O gerente recebe a estrutura inteira; o supervisor, só o próprio ramo."""
    with engine.connect() as conn:
        ids = escopo_vendedores(conn, auth)
        filtro = "" if ids is None else "AND u.usuario_id = ANY(CAST(:ids AS uuid[]))"
        params = {"cid": auth["conta_id"]}
        if ids is not None:
            params["ids"] = ids
        rows = conn.execute(
            text(
                f"""
                SELECT u.usuario_id, u.nome, u.email, u.role, u.ativo, u.supervisor_id,
                       COUNT(e.empresa_id) AS total_empresas
                FROM usuarios u
                LEFT JOIN empresas e ON e.vendedor_id = u.usuario_id
                WHERE u.conta_id = :cid {filtro}
                GROUP BY u.usuario_id, u.nome, u.email, u.role, u.ativo, u.supervisor_id
                ORDER BY u.nome ASC
            """
            ),
            params,
        ).fetchall()

    def pessoa(r):
        return {
            "usuario_id": str(r.usuario_id),
            "nome": r.nome,
            "email": r.email,
            "role": normalizar_role(r.role),
            "funcao": ROTULO_FUNCAO[normalizar_role(r.role)],
            "ativo": bool(r.ativo),
            "supervisor_id": str(r.supervisor_id) if r.supervisor_id else None,
            "total_empresas": int(r.total_empresas or 0),
        }

    pessoas = [pessoa(r) for r in rows]
    supervisores = [p for p in pessoas if p["role"] == "supervisor"]
    vendedores = [p for p in pessoas if p["role"] == "vendedor"]
    return {
        "gerentes": [p for p in pessoas if p["role"] == "gerente"],
        "supervisores": [
            {**s, "vendedores": [v for v in vendedores if v["supervisor_id"] == s["usuario_id"]]}
            for s in supervisores
        ],
        # Vendedores órfãos: existem de propósito (remover vínculo não apaga usuário).
        "sem_supervisor": [
            v for v in vendedores
            if not v["supervisor_id"]
            or v["supervisor_id"] not in {s["usuario_id"] for s in supervisores}
        ],
    }


@app.post("/usuarios", status_code=201)
async def criar_usuario(usuario: UsuarioCreate, auth: dict = Depends(exigir_gerente)):
    """Gerente adiciona um novo usuário (vendedor por padrão) já vinculado à sua
    conta, opcionalmente já atribuído a um supervisor. O usuário recebe email de
    ativação para criar a senha e então loga."""
    token_ativacao = str(uuid.uuid4())
    role = normalizar_role(usuario.role)
    novo_id = str(uuid.uuid4())
    try:
        with engine.begin() as conn:
            supervisor_id = None
            if usuario.supervisor_id:
                if role != "vendedor":
                    raise HTTPException(400, "Só vendedores podem ser atribuídos a um supervisor")
                supervisor_id = _validar_supervisor(conn, usuario.supervisor_id, novo_id, auth["conta_id"])
            conn.execute(
                text(
                    """
                INSERT INTO usuarios (usuario_id, nome, email, telefone, ativo, token_ativacao,
                    conta_id, role, supervisor_id, data_criacao)
                VALUES (:usuario_id, :nome, :email, :telefone, FALSE, :token,
                    :conta_id, :role, :supervisor_id, NOW())
            """
                ),
                {
                    "usuario_id": novo_id,
                    "nome": usuario.nome,
                    "email": usuario.email,
                    "telefone": usuario.telefone,
                    "token": token_ativacao,
                    "conta_id": auth["conta_id"],
                    "role": role,
                    "supervisor_id": supervisor_id,
                },
            )
    except IntegrityError:
        raise HTTPException(400, "Email já cadastrado")

    # O usuário já está gravado. Se o convite não sair, dizemos isso na cara e
    # devolvemos o link de ativação para o gerente repassar por outro canal —
    # em vez de fingir sucesso ou estourar 500 sobre um cadastro que existe.
    enviado, motivo = await enviar_email(usuario.email, token_ativacao)
    if enviado:
        return {"msg": "Usuário criado. O convite foi enviado por email 📩", "email_enviado": True}
    return {
        "msg": "Usuário criado, mas o convite por email não pôde ser enviado.",
        "email_enviado": False,
        "motivo": motivo,
        "link_ativacao": link_ativacao(token_ativacao),
    }


@app.post("/usuarios/{usuario_id}/reenviar-convite")
async def reenviar_convite(usuario_id: str, auth: dict = Depends(exigir_gerente)):
    """Gera um token de ativação novo e reenvia o convite.

    Serve para o caso em que o usuário foi criado mas o email não saiu: em vez
    de apagar e recadastrar (que esbarra em "Email já cadastrado"), o gerente
    reenvia. Só vale para quem ainda não ativou a conta — quem já tem senha não
    precisa de convite."""
    with engine.begin() as conn:
        alvo = conn.execute(
            text("SELECT usuario_id, nome, email, conta_id, ativo FROM usuarios WHERE usuario_id = :id"),
            {"id": usuario_id},
        ).fetchone()
        if not alvo or str(alvo.conta_id) != auth["conta_id"]:
            raise HTTPException(404, "Usuário não encontrado")
        if alvo.ativo:
            raise HTTPException(400, "Este usuário já ativou a conta e não precisa de convite")
        token_novo = str(uuid.uuid4())
        conn.execute(
            text("UPDATE usuarios SET token_ativacao = :t WHERE usuario_id = :id"),
            {"t": token_novo, "id": usuario_id},
        )
        registrar_auditoria(usuario=auth, acao="CONVITE_REENVIADO", recurso="usuarios",
                            recurso_id=usuario_id, conn=conn)

    enviado, motivo = await enviar_email(alvo.email, token_novo)
    if enviado:
        return {"msg": f"Convite reenviado para {alvo.email} 📩", "email_enviado": True}
    return {
        "msg": "Não foi possível enviar o email. Use o link abaixo para ativar a conta.",
        "email_enviado": False,
        "motivo": motivo,
        "link_ativacao": link_ativacao(token_novo),
    }


@app.patch("/usuarios/{usuario_id}")
def gerenciar_usuario(usuario_id: str, dados: UsuarioGerenciar, auth: dict = Depends(exigir_gerente)):
    """Gerente ativa/desativa um usuário, altera a função (vendedor/supervisor/
    gerente) e atribui ou remove o vínculo com um supervisor — sempre dentro da
    própria conta. Não pode rebaixar a si mesmo.

    Remover o vínculo NUNCA apaga o usuário: só zera `supervisor_id`."""
    with engine.begin() as conn:
        alvo = conn.execute(
            text("SELECT usuario_id, conta_id, role FROM usuarios WHERE usuario_id = :id"),
            {"id": usuario_id},
        ).fetchone()
        if not alvo or str(alvo.conta_id) != auth["conta_id"]:
            raise HTTPException(404, "Usuário não encontrado")

        role_atual = normalizar_role(alvo.role)
        nova_role = normalizar_role(dados.role, role_atual) if dados.role is not None else None
        if usuario_id == auth["usuario_id"] and nova_role and nova_role != "gerente":
            raise HTTPException(400, "Você não pode rebaixar a si mesmo")

        role_final = nova_role or role_atual

        # Vínculo com supervisor
        if dados.limpar_supervisor:
            novo_supervisor = None
            mexeu_no_vinculo = True
        elif dados.supervisor_id:
            if role_final != "vendedor":
                raise HTTPException(400, "Só vendedores podem ser atribuídos a um supervisor")
            novo_supervisor = _validar_supervisor(conn, dados.supervisor_id, usuario_id, auth["conta_id"])
            mexeu_no_vinculo = True
        else:
            novo_supervisor = None
            mexeu_no_vinculo = False

        # Quem deixa de ser vendedor não pode continuar pendurado num supervisor.
        if nova_role and nova_role != "vendedor":
            novo_supervisor = None
            mexeu_no_vinculo = True

        conn.execute(
            text(
                f"""
                UPDATE usuarios SET
                    ativo = COALESCE(:ativo, ativo),
                    role  = COALESCE(:role, role)
                    {", supervisor_id = :supervisor_id" if mexeu_no_vinculo else ""}
                WHERE usuario_id = :id
            """
            ),
            {
                "ativo": dados.ativo,
                "role": nova_role,
                "id": usuario_id,
                **({"supervisor_id": novo_supervisor} if mexeu_no_vinculo else {}),
            },
        )

        # Supervisor que perde a função deixaria vendedores apontando para alguém
        # que não é mais supervisor — desvincula em vez de deixar inconsistente.
        soltos = 0
        if role_atual == "supervisor" and nova_role and nova_role != "supervisor":
            soltos = conn.execute(
                text("UPDATE usuarios SET supervisor_id = NULL WHERE supervisor_id = :id"),
                {"id": usuario_id},
            ).rowcount or 0

    msg = "Usuário atualizado com sucesso"
    if soltos:
        msg += f". {soltos} vendedor(es) ficaram sem supervisor."
    return {"msg": msg, "vendedores_desvinculados": soltos}


@app.get("/gerencia/dashboard")
def dashboard_gerente(auth: dict = Depends(exigir_gestor)):
    """Visão geral para gerente e supervisor: totais e desempenho por vendedor
    (empresas, distribuição por status e ticket estimado).

    O gerente vê a conta inteira; o supervisor, apenas os vendedores atribuídos
    a ele — o recorte é feito aqui, não no frontend."""
    cid = auth["conta_id"]
    with engine.begin() as conn:
        garantir_campos_pipeline(conn)

        # Mesmo escopo usado em /empresas: gerente = conta toda, supervisor = equipe.
        trecho, params = filtro_escopo(conn, auth)
        ids = escopo_vendedores(conn, auth)
        filtro_usuarios = "" if ids is None else "AND u.usuario_id = ANY(CAST(:vids AS uuid[]))"
        # O app grava `Fechado`; estas consultas comparavam com `Ganho` e por
        # isso os dois contadores de "ganhos" devolviam zero desde sempre. A
        # lista vive em funil.py para as duas nao voltarem a divergir.
        params["ganhos"] = STATUS_GANHO

        totais = conn.execute(
            text(
                f"""
                SELECT
                    COUNT(*) AS total_empresas,
                    COUNT(*) FILTER (WHERE status = ANY(CAST(:ganhos AS text[]))) AS ganhos,
                    COUNT(*) FILTER (WHERE status = 'Perdido') AS perdidos,
                    COUNT(*) FILTER (WHERE status_cadastro = 'rascunho') AS rascunhos
                FROM empresas WHERE conta_id = :cid {trecho}
            """
            ),
            params,
        ).fetchone()

        # Pipeline em dinheiro. Era SUM(ticket_medio_estimado) — um campo que o
        # vendedor digitava no cadastro e ninguem revisava depois; o front parou
        # de preencher esse campo. Agora e o valor dos orcamentos que ja foram
        # para o cliente e continuam sem decisao. Rascunho fica de fora: enquanto
        # nao sai daqui, nao e dinheiro em jogo.
        trecho_emp, _ = filtro_escopo(conn, auth, prefixo="e.")
        ticket_total = conn.execute(
            text(
                f"""
                SELECT COALESCE(SUM(o.total), 0)
                FROM orcamentos o
                JOIN empresas e ON e.empresa_id = o.empresa_id
                WHERE o.conta_id = :cid
                  AND o.status IN ('enviado', 'em_negociacao')
                  {trecho_emp}
            """
            ),
            params,
        ).scalar()

        total_vendedores = conn.execute(
            text(
                f"""SELECT COUNT(*) FROM usuarios u
                    WHERE u.conta_id = :cid AND u.role = 'vendedor' {filtro_usuarios}"""
            ),
            params,
        ).scalar()

        # Desempenho por vendedor
        por_vendedor = conn.execute(
            text(
                f"""
                SELECT u.usuario_id, u.nome, u.email, u.ativo, u.supervisor_id,
                       COUNT(e.empresa_id) AS total_empresas,
                       COUNT(e.empresa_id) FILTER (WHERE e.status = ANY(CAST(:ganhos AS text[]))) AS ganhos,
                       COUNT(e.empresa_id) FILTER (WHERE e.status = 'Perdido') AS perdidos,
                       COUNT(e.empresa_id) FILTER (WHERE e.status_cadastro = 'rascunho') AS rascunhos,
                       COALESCE((
                           SELECT SUM(o.total) FROM orcamentos o
                           WHERE o.conta_id = :cid
                             AND o.status IN ('enviado', 'em_negociacao')
                             AND o.empresa_id IN (
                                 SELECT empresa_id FROM empresas
                                 WHERE conta_id = :cid AND vendedor_id = u.usuario_id
                             )
                       ), 0) AS ticket_total,
                       MAX(e.status_atualizado_em) AS ultima_atividade
                FROM usuarios u
                LEFT JOIN empresas e ON e.vendedor_id = u.usuario_id AND e.conta_id = :cid
                WHERE u.conta_id = :cid AND u.role = 'vendedor' {filtro_usuarios}
                GROUP BY u.usuario_id, u.nome, u.email, u.ativo, u.supervisor_id
                ORDER BY total_empresas DESC, u.nome ASC
            """
            ),
            params,
        )
        vendedores = [dict(r._mapping) for r in por_vendedor]

        # Distribuição por status (dentro do escopo do solicitante)
        por_status = conn.execute(
            text(
                f"""
                SELECT COALESCE(status, 'Sem status') AS status, COUNT(*) AS total
                FROM empresas WHERE conta_id = :cid {trecho}
                GROUP BY status ORDER BY total DESC
            """
            ),
            params,
        )
        distribuicao_status = [dict(r._mapping) for r in por_status]

    return {
        "conta": {
            "total_empresas": totais.total_empresas,
            "ganhos": totais.ganhos,
            "perdidos": totais.perdidos,
            "rascunhos": totais.rascunhos,
            "ticket_total": float(ticket_total or 0),
            "total_vendedores": total_vendedores,
        },
        "distribuicao_status": distribuicao_status,
        "vendedores": vendedores,
    }


@app.post("/signup", status_code=201)
@limiter.limit("10/hour")
async def signup_conta(dados: ContaSignup, request: Request):
    """Cadastro de uma NOVA assinatura: cria a conta e o primeiro usuário como
    gerente (ADM). O gerente recebe email para ativar a conta e definir a senha.
    Pagamento/cobrança fica fora deste fluxo por enquanto."""
    token_ativacao = str(uuid.uuid4())
    conta_id = str(uuid.uuid4())
    try:
        with engine.begin() as conn:
            garantir_multiusuario(conn)
            existe = conn.execute(
                text("SELECT 1 FROM usuarios WHERE LOWER(email) = LOWER(:e)"),
                {"e": dados.email},
            ).fetchone()
            if existe:
                raise HTTPException(400, "Email já cadastrado")
            conn.execute(
                text("INSERT INTO contas (conta_id, nome) VALUES (:id, :nome)"),
                {"id": conta_id, "nome": dados.empresa_nome.strip() or "Minha empresa"},
            )
            conn.execute(
                text(
                    """
                INSERT INTO usuarios (usuario_id, nome, email, telefone, ativo, token_ativacao,
                    conta_id, role, empresa_nome, data_criacao)
                VALUES (:uid, :nome, :email, :tel, FALSE, :token,
                    :cid, 'gerente', :empnome, NOW())
            """
                ),
                {
                    "uid": str(uuid.uuid4()),
                    "nome": dados.nome,
                    "email": dados.email,
                    "tel": dados.telefone,
                    "token": token_ativacao,
                    "cid": conta_id,
                    "empnome": dados.empresa_nome.strip() or None,
                },
            )
    except IntegrityError:
        raise HTTPException(400, "Email já cadastrado")

    # Mesma regra do convite: a conta já existe, então falha de email não pode
    # virar 500. O link vai na resposta para quem acabou de se cadastrar não
    # ficar preso sem nenhuma forma de ativar.
    enviado, motivo = await enviar_email(dados.email, token_ativacao)
    if enviado:
        return {"msg": "Conta criada! Verifique seu email para ativar. 📩", "conta_id": conta_id,
                "email_enviado": True}
    return {
        "msg": "Conta criada, mas o email de ativação não pôde ser enviado.",
        "conta_id": conta_id,
        "email_enviado": False,
        "motivo": motivo,
        "link_ativacao": link_ativacao(token_ativacao),
    }


@app.post("/ativar-conta")
def ativar_conta(dados: AtivarConta):
    senha_hash = hash_senha(dados.senha)
    with engine.begin() as conn:
        result = conn.execute(
            text(
                """
            UPDATE usuarios SET senha_hash = :senha, ativo = TRUE, token_ativacao = NULL
            WHERE token_ativacao = :token RETURNING usuario_id
        """
            ),
            {"senha": senha_hash, "token": dados.token},
        ).fetchone()
        if not result:
            raise HTTPException(400, "Token inválido")
    return {"msg": "Conta ativada com sucesso 🚀"}


# Cookie do refresh token. Frontend (vercel.app) e backend (railway.app) são
# cross-site → exige SameSite=None + Secure para o cookie ser enviado.
REFRESH_COOKIE = "refresh_token"


def set_refresh_cookie(response: Response, token: str):
    response.set_cookie(
        key=REFRESH_COOKIE,
        value=token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=REFRESH_TOKEN_EXPIRE_DAYS * 24 * 60 * 60,
        path="/",
    )


def clear_refresh_cookie(response: Response):
    response.set_cookie(
        key=REFRESH_COOKIE, value="", httponly=True, secure=True,
        samesite="none", max_age=0, path="/",
    )


def emitir_sessao(conn, usuario: dict, request: Request, response: Response,
                  familia: Optional[str] = None) -> dict:
    """Emite access token (JWT curto) + refresh token (cookie httpOnly)."""
    access = criar_token_acesso({"sub": usuario["email"]})
    raw_refresh = criar_refresh_token(conn, str(usuario["usuario_id"]), familia, request)
    set_refresh_cookie(response, raw_refresh)
    return {"access_token": access, "token_type": "bearer"}


@app.post("/login")
@limiter.limit("10/15minutes")
def login(dados: Login, request: Request, response: Response):
    # Lockout progressivo por conta (além do rate limit por IP do slowapi).
    checar_lockout(dados.email)

    with engine.begin() as conn:
        garantir_seguranca(conn)
        usuario = conn.execute(
            text("SELECT * FROM usuarios WHERE email = :email"), {"email": dados.email}
        ).fetchone()

        # Anti-enumeração: mensagem única e verificação de senha em tempo ~constante
        # mesmo quando o e-mail não existe (evita timing attack).
        if not usuario:
            verificar_senha(dados.senha, _DUMMY_HASH)
            registrar_falha_login(dados.email)
            registrar_auditoria(email=dados.email, acao="LOGIN_FALHOU",
                                request=request, meta={"motivo": "email_inexistente"}, conn=conn)
            raise HTTPException(401, CREDENCIAIS_INVALIDAS)

        usuario = dict(usuario._mapping)

        if not usuario.get("senha_hash") or not verificar_senha(dados.senha, usuario["senha_hash"]):
            registrar_falha_login(dados.email)
            registrar_auditoria(email=dados.email, acao="LOGIN_FALHOU",
                                request=request, meta={"motivo": "senha_invalida"}, conn=conn)
            raise HTTPException(401, CREDENCIAIS_INVALIDAS)

        if not usuario["ativo"]:
            raise HTTPException(401, "Conta não ativada. Verifique seu e-mail de ativação.")

        # MFA obrigatório quando ativado na conta.
        if usuario.get("mfa_ativado"):
            code = (dados.mfa_code or "").strip()
            if not code:
                # Senha OK mas falta o 2º fator — sinaliza ao frontend sem emitir token.
                return {"mfa_required": True}
            ok = mfa_verificar_totp(usuario.get("mfa_secret"), code)
            if not ok:
                # Tenta como código de backup (uso único).
                backups = usuario.get("mfa_backup_codes") or []
                h = _hash_token(code)
                if h in backups:
                    backups = [b for b in backups if b != h]
                    conn.execute(
                        text("UPDATE usuarios SET mfa_backup_codes = CAST(:b AS JSONB) WHERE usuario_id = :id"),
                        {"b": json.dumps(backups), "id": usuario["usuario_id"]},
                    )
                    ok = True
            if not ok:
                registrar_falha_login(dados.email)
                registrar_auditoria(usuario={"usuario_id": str(usuario["usuario_id"]),
                                             "email": usuario["email"],
                                             "conta_id": str(usuario.get("conta_id")) if usuario.get("conta_id") else None},
                                    acao="MFA_FALHOU", request=request, conn=conn)
                raise HTTPException(401, "Código de verificação inválido.")

        limpar_falhas_login(dados.email)
        sessao = emitir_sessao(conn, usuario, request, response)
        registrar_auditoria(usuario={"usuario_id": str(usuario["usuario_id"]),
                                     "email": usuario["email"],
                                     "conta_id": str(usuario.get("conta_id")) if usuario.get("conta_id") else None},
                            acao="LOGIN_OK", request=request, conn=conn)
    return sessao


@app.post("/refresh")
def refresh_token_endpoint(request: Request, response: Response):
    """Rotação de refresh token: valida o cookie, revoga o antigo e emite novos.
    Reuso de token já usado/revogado → revoga a família inteira (proteção contra roubo)."""
    raw = request.cookies.get(REFRESH_COOKIE)
    if not raw or "." not in raw:
        raise HTTPException(401, "Sessão expirada, faça login novamente.")
    familia, _, _ = raw.partition(".")
    h = _hash_token(raw.split(".", 1)[1])
    with engine.begin() as conn:
        garantir_seguranca(conn)
        row = conn.execute(
            text(
                """
                SELECT rt.id, rt.usuario_id, rt.familia, rt.revogado, rt.expira_em, rt.usado_em,
                       u.email, u.conta_id, u.ativo
                FROM refresh_tokens rt
                JOIN usuarios u ON u.usuario_id = rt.usuario_id
                WHERE rt.token_hash = :h
                """
            ),
            {"h": h},
        ).fetchone()
        if not row or str(row.familia) != familia:
            clear_refresh_cookie(response)
            raise HTTPException(401, "Sessão expirada, faça login novamente.")
        if row.revogado or row.usado_em is not None:
            # Reuso detectado → revoga toda a família (possível token roubado).
            revogar_refresh_familia(conn, familia)
            registrar_auditoria(email=row.email, acao="REFRESH_REUSO_DETECTADO",
                                request=request, conn=conn)
            clear_refresh_cookie(response)
            raise HTTPException(401, "Sessão inválida, faça login novamente.")
        if row.expira_em < datetime.utcnow() or not row.ativo:
            clear_refresh_cookie(response)
            raise HTTPException(401, "Sessão expirada, faça login novamente.")

        # Marca o atual como usado/revogado e emite um novo na mesma família.
        conn.execute(
            text("UPDATE refresh_tokens SET revogado = TRUE, usado_em = NOW() WHERE id = :id"),
            {"id": row.id},
        )
        usuario = {"usuario_id": str(row.usuario_id), "email": row.email}
        sessao = emitir_sessao(conn, usuario, request, response, familia=familia)
    return sessao


@app.post("/logout")
def logout(request: Request, response: Response):
    raw = request.cookies.get(REFRESH_COOKIE)
    if raw and "." in raw:
        familia = raw.split(".", 1)[0]
        with engine.begin() as conn:
            revogar_refresh_familia(conn, familia)
    clear_refresh_cookie(response)
    return {"msg": "Sessão encerrada."}


@app.post("/logout-all")
def logout_all(request: Request, response: Response, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        revogar_refresh_usuario(conn, auth["usuario_id"])
        registrar_auditoria(usuario=auth, acao="LOGOUT_TODOS", request=request, conn=conn)
    clear_refresh_cookie(response)
    return {"msg": "Todas as sessões foram encerradas."}


# =========================
# MFA / TOTP
# =========================
@app.post("/mfa/setup")
def mfa_setup(auth: dict = Depends(get_auth)):
    """Gera um segredo TOTP (ainda não ativado) e devolve o QR code para
    escanear no Google Authenticator/Authy."""
    secret = pyotp.random_base32()
    otpauth = pyotp.totp.TOTP(secret).provisioning_uri(
        name=auth["email"], issuer_name="CRM Prospecção"
    )
    with engine.begin() as conn:
        conn.execute(
            text("UPDATE usuarios SET mfa_secret = :s WHERE usuario_id = :id"),
            {"s": secret, "id": auth["usuario_id"]},
        )
    qr_data_url = None
    try:
        import qrcode
        import io
        img = qrcode.make(otpauth)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        qr_data_url = "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()
    except Exception as e:
        print(f"⚠️ QR não gerado: {e}")
    return {"secret": secret, "otpauth_url": otpauth, "qr_code": qr_data_url}


@app.post("/mfa/ativar")
def mfa_ativar(dados: MFAAtivar, request: Request, auth: dict = Depends(get_auth)):
    """Confirma o código do app autenticador, ativa o MFA e devolve os
    códigos de backup (mostrados UMA única vez)."""
    with engine.begin() as conn:
        row = conn.execute(
            text("SELECT mfa_secret FROM usuarios WHERE usuario_id = :id"),
            {"id": auth["usuario_id"]},
        ).fetchone()
        if not row or not row.mfa_secret:
            raise HTTPException(400, "Inicie a configuração do MFA primeiro.")
        if not mfa_verificar_totp(row.mfa_secret, dados.code):
            raise HTTPException(400, "Código inválido. Tente novamente.")
        codes, hashes = mfa_gerar_backup_codes()
        conn.execute(
            text(
                "UPDATE usuarios SET mfa_ativado = TRUE, mfa_backup_codes = CAST(:b AS JSONB) "
                "WHERE usuario_id = :id"
            ),
            {"b": json.dumps(hashes), "id": auth["usuario_id"]},
        )
        registrar_auditoria(usuario=auth, acao="MFA_ATIVADO", request=request, conn=conn)
    return {"msg": "MFA ativado com sucesso.", "backup_codes": codes}


@app.post("/mfa/desativar")
def mfa_desativar(dados: MFADesativar, request: Request, auth: dict = Depends(get_auth)):
    """Desativa o MFA (exige a senha da conta como confirmação)."""
    with engine.begin() as conn:
        row = conn.execute(
            text("SELECT senha_hash FROM usuarios WHERE usuario_id = :id"),
            {"id": auth["usuario_id"]},
        ).fetchone()
        if not row or not verificar_senha(dados.senha, row.senha_hash):
            raise HTTPException(401, "Senha inválida.")
        conn.execute(
            text(
                "UPDATE usuarios SET mfa_ativado = FALSE, mfa_secret = NULL, mfa_backup_codes = NULL "
                "WHERE usuario_id = :id"
            ),
            {"id": auth["usuario_id"]},
        )
        registrar_auditoria(usuario=auth, acao="MFA_DESATIVADO", request=request, conn=conn)
    return {"msg": "MFA desativado."}


@app.get("/gerencia/auditoria")
def listar_auditoria(auth: dict = Depends(exigir_gerente), limite: int = 200):
    """Logs de auditoria da conta (acesso a leads, autenticação, exportações).
    Restrito ao gerente. Atende rastreabilidade exigida pela LGPD."""
    limite = max(1, min(limite, 1000))
    with engine.begin() as conn:
        garantir_seguranca(conn)
        rows = conn.execute(
            text(
                """
                SELECT id, usuario_email, acao, recurso, recurso_id, quantidade,
                       ip, criado_em
                FROM audit_log
                WHERE conta_id = :cid
                   OR (conta_id IS NULL AND usuario_email IN (
                        SELECT email FROM usuarios WHERE conta_id = :cid))
                ORDER BY criado_em DESC
                LIMIT :lim
                """
            ),
            {"cid": auth["conta_id"], "lim": limite},
        )
        return [dict(r._mapping) for r in rows]


# =========================
# VENDAS: EQUIPAMENTOS
# =========================
def _escopo_vendas(conn, auth: dict, alias: str = "o"):
    """Vendedor enxerga só o que é dele; supervisor, o da própria equipe;
    gerente, tudo da conta."""
    trecho, params = filtro_escopo(conn, auth, prefixo=f"{alias}.")
    return f"{alias}.conta_id = :cid {trecho}", params



# ── Busca tolerante de itens (autocomplete do orcamento) ─────────────────────
#
# POR QUE NAO pg_trgm: exigiria CREATE EXTENSION no Postgres de producao para
# ganhar performance que este volume nao pede. Substring normalizado resolve em
# milissegundos, e a ordenacao fina acontece em Python.
#
# TODO o SQL de busca de item mora em buscar_itens_similares(). Migrar para
# pg_trgm um dia e reescrever o corpo daquela funcao, e nada mais.

# Mapa de acentos do lado do SQL: os DOIS lados da comparacao (coluna e termo)
# passam pelo mesmo lower() + translate(), senao a comparacao e assimetrica.
#
# O destino e TODO minusculo, inclusive para as 24 maiusculas. Em Postgres com
# locale C, lower('Á') devolve 'Á' intacto -- mapear para 'A' faria "ÁGUA"
# nunca casar com "agua" digitado, e o erro so apareceria nesse item.
_ACENTOS_DE = "áàâãäéèêëíìîïóòôõöúùûüçñÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇÑ"
_ACENTOS_PARA = "aaaaaeeeeiiiiooooouuuucn" * 2
# translate() ignora em silencio o excedente do lado maior: um mapa desalinhado
# nao levanta erro, so para de normalizar de um caractere em diante.
assert len(_ACENTOS_DE) == len(_ACENTOS_PARA) == 48, "mapa de acentos desalinhado"

# Uma letra casa com quase tudo, e a pessoa ainda esta digitando.
MIN_TERMO_SUGESTAO = 2
# Piso da passada tolerante a erro de digitacao.
SIMILARIDADE_MINIMA = 0.6
# A partir daqui a UI oferece o item existente em vez de deixar criar outro.
SIMILARIDADE_AVISO = 0.85


def _sem_acento_sql(expr: str) -> str:
    """Expressao SQL equivalente ao normalizar_texto() do lado Python."""
    return f"translate(lower({expr}), :ac_de, :ac_para)"


def _ordem_sugestao(alvo: str, texto: str):
    """Chave de ordenacao: prefixo exato > comeca com > contem > parecenca.

    Empate resolve por uso mais recente -- o item que a equipe acabou de usar e
    o palpite melhor entre dois nomes igualmente parecidos.
    """
    if texto == alvo:
        faixa = 0
    elif texto.startswith(alvo):
        faixa = 1
    elif alvo in texto:
        faixa = 2
    else:
        faixa = 3
    return faixa


def buscar_itens_similares(conn, auth: dict, termo: str,
                           tipo: Optional[str] = None, limite: int = 10) -> list[dict]:
    """Sugestoes de item para o autocomplete do orcamento.

    Une o catalogo da conta com as descricoes de itens AVULSOS ja usadas em
    orcamentos anteriores, deduplicadas pela descricao normalizada.

    Escopo: catalogo e por conta (e compartilhado); historico passa pelo
    _escopo_vendas, senao vendedor veria o que o colega digitou.

    Devolve, por sugestao:
        descricao       str
        tipo            "equipamento" | "servico"
        ultimo_preco    float | None   -- do orcamento mais recente, nao a media
        origem          "catalogo" | "historico"
        equipamento_id  str | None     -- so quando origem == "catalogo"
        similaridade    float          -- 0..1, para a UI avisar sobre duplicata
    """
    alvo = normalizar_texto(termo or "")
    if len(alvo) < MIN_TERMO_SUGESTAO:
        return []

    escopo, params = _escopo_vendas(conn, auth)
    params = dict(params)
    params.update({
        "cid": auth["conta_id"],
        "ac_de": _ACENTOS_DE,
        "ac_para": _ACENTOS_PARA,
        "q": alvo,
    })

    filtro_cat = filtro_hist = ""
    if tipo is not None:
        params["tipo"] = _tipo_catalogo(tipo)
        filtro_cat = " AND COALESCE(e.tipo, 'equipamento') = :tipo"
        filtro_hist = " AND h.tipo = :tipo"

    nome_norm = _sem_acento_sql("e.nome")
    desc_norm = _sem_acento_sql("u.descricao")

    def consultar(predicado_cat: str, predicado_hist: str) -> list:
        sql = f"""
            WITH usados AS (
                SELECT i.equipamento_id, i.descricao, i.preco_unitario,
                       COALESCE(i.tipo, 'equipamento') AS tipo, o.criado_em
                FROM orcamento_itens i
                JOIN orcamentos o ON o.orcamento_id = i.orcamento_id
                WHERE {escopo}
            )
            SELECT e.equipamento_id::text          AS equipamento_id,
                   e.nome                          AS descricao,
                   COALESCE(e.tipo, 'equipamento') AS tipo,
                   e.preco_base                    AS ultimo_preco,
                   'catalogo'                      AS origem,
                   (SELECT MAX(u2.criado_em) FROM usados u2
                     WHERE u2.equipamento_id = e.equipamento_id) AS usado_em
            FROM equipamentos e
            WHERE e.conta_id = :cid AND e.ativo = TRUE{filtro_cat}
              AND {predicado_cat}
            UNION ALL
            SELECT NULL AS equipamento_id, h.descricao, h.tipo,
                   h.preco_unitario AS ultimo_preco,
                   'historico'      AS origem,
                   h.criado_em      AS usado_em
            FROM (
                SELECT DISTINCT ON ({desc_norm})
                       u.descricao, u.tipo, u.preco_unitario, u.criado_em
                FROM usados u
                WHERE u.equipamento_id IS NULL
                ORDER BY {desc_norm}, u.criado_em DESC NULLS LAST
            ) h
            WHERE TRUE{filtro_hist}
              AND {predicado_hist}
            LIMIT 500
        """
        return conn.execute(text(sql), params).fetchall()

    # 1a passada: substring normalizado. Cobre "almoco" -> "Almoço", que e o caso
    # que motivou a feature, e devolve pouca linha.
    #
    # POSITION e nao LIKE, por dois motivos. O SQLAlchemy dobra sozinho o '%'
    # literal ao compilar para o paramstyle do psycopg2, entao escrever '%%' na
    # fonte vira '%%' de verdade na consulta e nao casa com nada -- falha calada,
    # ainda por cima coberta pela 2a passada. E LIKE trataria '%' ou '_'
    # digitados pela pessoa como curinga: buscar "50%" devolveria o catalogo
    # inteiro. POSITION compara texto literal e nao tem escape nenhum.
    linhas = consultar(
        f"POSITION(:q IN {nome_norm}) > 0",
        f"POSITION(:q IN {_sem_acento_sql('h.descricao')}) > 0",
    )
    # 2a passada, so quando a primeira nao encheu a lista: erro de digitacao
    # ("almco") nao casa por substring, e sem isto o typeahead ficaria mudo
    # justamente para quem mais precisa dele. Custa uma varredura do catalogo,
    # que e barata nesta ordem de grandeza -- e nao roda no caminho comum.
    if len(linhas) < limite:
        vistos = {(r.origem, normalizar_texto(r.descricao)) for r in linhas}
        for r in consultar("TRUE", "TRUE"):
            if (r.origem, normalizar_texto(r.descricao)) not in vistos:
                linhas.append(r)

    # Dedupe entre as duas fontes: mesmo nome no catalogo e no historico e o
    # mesmo item, e o catalogo ganha -- ele tem preco de tabela e equipamento_id.
    por_nome: dict[str, dict] = {}
    for r in linhas:
        chave = normalizar_texto(r.descricao)
        if not chave:
            continue
        parecenca = difflib.SequenceMatcher(None, alvo, chave).ratio()
        faixa = _ordem_sugestao(alvo, chave)
        # Fora da faixa de substring, so entra quem se parece de verdade.
        if faixa == 3 and parecenca < SIMILARIDADE_MINIMA:
            continue
        atual = por_nome.get(chave)
        if atual and not (atual["origem"] == "historico" and r.origem == "catalogo"):
            continue
        por_nome[chave] = {
            "equipamento_id": r.equipamento_id,
            "descricao": r.descricao,
            "tipo": r.tipo or "equipamento",
            "ultimo_preco": float(r.ultimo_preco) if r.ultimo_preco is not None else None,
            "origem": r.origem,
            "similaridade": round(parecenca, 3),
            "_faixa": faixa,
            "_usado_em": r.usado_em,
        }

    ordenadas = sorted(
        por_nome.values(),
        key=lambda s: (
            s["_faixa"],
            -s["similaridade"],
            # Mais usado recentemente primeiro; nunca usado vai para o fim.
            -(s["_usado_em"].timestamp() if s["_usado_em"] else 0),
            s["descricao"].lower(),
        ),
    )
    for s in ordenadas:
        s.pop("_faixa", None)
        s.pop("_usado_em", None)
    return ordenadas[:limite]


@app.get("/catalogo/sugestoes")
def sugestoes_catalogo(q: str = "", tipo: Optional[str] = None, limite: int = 10,
                       auth: dict = Depends(get_auth)):
    """Autocomplete da descricao do item no editor de orcamento.

    Existe para padronizar nomenclatura na ENTRADA: sem ela, "almoço", "almoco"
    e "Almoço equipe" viram tres itens distintos e os graficos de historico
    agrupam errado. Orienta, nao bloqueia -- texto novo continua permitido.
    """
    try:
        teto = int(limite)
    except (TypeError, ValueError):
        teto = 10
    teto = max(1, min(teto, 25))
    with engine.begin() as conn:
        garantir_vendas(conn)
        return buscar_itens_similares(conn, auth, q, tipo=tipo, limite=teto)


@app.get("/equipamentos")
def listar_equipamentos(auth: dict = Depends(get_auth), incluir_inativos: bool = False,
                        tipo: Optional[str] = None):
    """Catálogo da conta usado para montar orçamentos.

    Sem `tipo`, devolve equipamentos E serviços — é assim que a tela de vendas
    busca uma vez só e separa as abas do lado dela, em vez de fazer duas
    requisições para dois recortes da mesma tabela."""
    with engine.begin() as conn:
        garantir_vendas(conn)
        filtro = "" if incluir_inativos else " AND ativo = TRUE"
        params = {"cid": auth["conta_id"]}
        if tipo is not None:
            filtro += " AND COALESCE(tipo, 'equipamento') = :tipo"
            params["tipo"] = _tipo_catalogo(tipo)
        rows = conn.execute(
            text(
                f"""SELECT equipamento_id, codigo, nome, descricao, preco_base,
                           COALESCE(quantidade, 0) AS quantidade,
                           COALESCE(tipo, 'equipamento') AS tipo, ativo, criado_em
                    FROM equipamentos WHERE conta_id = :cid{filtro} ORDER BY nome ASC"""
            ),
            params,
        )
        return [dict(r._mapping) for r in rows]


@app.post("/equipamentos")
def criar_equipamento(dados: EquipamentoCreate, auth: dict = Depends(get_auth)):
    tipo = _tipo_catalogo(dados.tipo)
    nome = (dados.nome or "").strip()
    if not nome:
        rotulo = "serviço" if tipo == "servico" else "equipamento"
        raise HTTPException(400, f"Nome do {rotulo} é obrigatório")
    with engine.begin() as conn:
        garantir_vendas(conn)
        codigo = (dados.codigo or "").strip() or None
        if codigo:
            # Codigo e unico DENTRO do catalogo, nao na conta inteira: o usuario
            # ve duas listas separadas na tela, e recusar "SRV-01" porque existe
            # um equipamento com esse codigo seria uma colisao invisivel para
            # ele. A importacao casa pelo mesmo par (tipo, codigo).
            ja_existe = conn.execute(
                text("""SELECT 1 FROM equipamentos
                        WHERE conta_id = :cid AND lower(codigo) = lower(:c)
                          AND COALESCE(tipo, 'equipamento') = :t"""),
                {"cid": auth["conta_id"], "c": codigo, "t": tipo},
            ).fetchone()
            if ja_existe:
                raise HTTPException(400, f"Já existe um item com o código '{codigo}'")
        row = conn.execute(
            text(
                """INSERT INTO equipamentos (conta_id, codigo, nome, descricao, preco_base, quantidade, tipo)
                   VALUES (:cid, :c, :n, :d, :p, :q, :t)
                   RETURNING equipamento_id, codigo, nome, descricao, preco_base,
                             COALESCE(quantidade, 0) AS quantidade,
                             COALESCE(tipo, 'equipamento') AS tipo, ativo, criado_em"""
            ),
            {
                "cid": auth["conta_id"], "c": codigo, "n": nome, "d": dados.descricao,
                "p": dados.preco_base or 0,
                # Servico nao tem estoque: gravar 0 evita que a tela mostre um
                # saldo que ninguem controla.
                "q": 0 if tipo == "servico" else (dados.quantidade or 0),
                "t": tipo,
            },
        ).fetchone()
        return dict(row._mapping)


@app.put("/equipamentos/{equipamento_id}")
def atualizar_equipamento(equipamento_id: str, dados: EquipamentoUpdate, auth: dict = Depends(get_auth)):
    campos = {k: v for k, v in dados.dict().items() if v is not None}
    if "tipo" in campos:
        campos["tipo"] = _tipo_catalogo(campos["tipo"])
    if not campos:
        raise HTTPException(400, "Nada para atualizar")
    with engine.begin() as conn:
        garantir_vendas(conn)
        sets = ", ".join(f"{k} = :{k}" for k in campos)
        params = {**campos, "eid": equipamento_id, "cid": auth["conta_id"]}
        row = conn.execute(
            text(
                f"""UPDATE equipamentos SET {sets}
                    WHERE equipamento_id = :eid AND conta_id = :cid
                    RETURNING equipamento_id, codigo, nome, descricao, preco_base,
                              COALESCE(quantidade, 0) AS quantidade,
                              COALESCE(tipo, 'equipamento') AS tipo, ativo, criado_em"""
            ),
            params,
        ).fetchone()
        if not row:
            raise HTTPException(404, "Equipamento não encontrado")
        return dict(row._mapping)


@app.delete("/equipamentos/{equipamento_id}")
def desativar_equipamento(equipamento_id: str, auth: dict = Depends(get_auth)):
    """Desativa em vez de apagar: orçamentos antigos continuam apontando pro item."""
    with engine.begin() as conn:
        garantir_vendas(conn)
        row = conn.execute(
            text(
                """UPDATE equipamentos SET ativo = FALSE
                   WHERE equipamento_id = :eid AND conta_id = :cid RETURNING equipamento_id"""
            ),
            {"eid": equipamento_id, "cid": auth["conta_id"]},
        ).fetchone()
        if not row:
            raise HTTPException(404, "Equipamento não encontrado")
        return {"msg": "Equipamento desativado"}


# =========================
# VENDAS: IMPORTAÇÃO DE CATÁLOGO/ESTOQUE (EXCEL)
# =========================
# O mapeamento é SEMPRE por nome de cabeçalho, nunca por posição da coluna:
# se o usuário reorganizar as colunas no Excel, o preço continua indo para
# preço e a quantidade para quantidade.
CAMPOS_IMPORTACAO = [
    # (campo, obrigatório, rótulo no modelo, sinônimos aceitos no cabeçalho)
    ("codigo",     False, "Código",            ["codigo", "code", "sku", "referencia", "ref", "id", "identificador"]),
    ("nome",       True,  "Nome",              ["nome", "item", "produto", "equipamento", "titulo"]),
    ("descricao",  False, "Descrição",         ["descricao", "detalhe", "detalhes", "observacao", "observacoes"]),
    ("quantidade", False, "Quantidade",        ["quantidade", "qtd", "qtde", "estoque", "saldo"]),
    ("preco_base", True,  "Preço unitário",    ["preco unitario", "preco", "valor unitario", "valor",
                                                "preco base", "preco de venda", "unitario"]),
]
CAMPOS_OBRIGATORIOS = [c for c, obrig, _, _ in CAMPOS_IMPORTACAO if obrig]
ROTULOS_IMPORTACAO = {c: r for c, _, r, _ in CAMPOS_IMPORTACAO}
LIMITE_LINHAS_IMPORTACAO = 5000


def _normalizar_cabecalho(valor) -> str:
    """'Preço Unitário (R$)' → 'preco unitario'. Tira acento, caixa, pontuação
    e o que estiver entre parênteses, para casar cabeçalhos escritos à mão."""
    txt = str(valor or "").strip()
    txt = re.sub(r"\(.*?\)", " ", txt)
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode()
    txt = re.sub(r"[^a-zA-Z0-9]+", " ", txt).strip().lower()
    return re.sub(r"\s+", " ", txt)


# Palavras de enfeite que aparecem no cabeçalho sem mudar o significado:
# "Nome do item" é a mesma coluna que "Nome".
_RUIDO_CABECALHO = {"do", "da", "de", "dos", "das", "o", "a", "em", "por",
                    "item", "itens", "produto", "produtos", "equipamento", "equipamentos"}


def _cabecalho_sem_ruido(titulo: str) -> str:
    restante = [t for t in titulo.split() if t not in _RUIDO_CABECALHO]
    return " ".join(restante)


def _mapear_colunas(cabecalho: list) -> tuple[dict, list]:
    """Descobre em que índice cada campo está, pelo NOME do cabeçalho.

    Duas passadas: primeiro casamento exato com os sinônimos, depois o mesmo
    casamento ignorando palavras de enfeite ("Nome do item" → "nome"). Nunca é
    aproximado por prefixo: "Valor total" não vira "Valor unitário por engano.
    Devolve ({campo: indice}, faltando)."""
    normalizados = [_normalizar_cabecalho(c) for c in cabecalho]
    mapa: dict = {}
    for chave in ("exato", "sem_ruido"):
        for campo, _obrig, _rotulo, sinonimos in CAMPOS_IMPORTACAO:
            if campo in mapa:
                continue
            for idx, titulo in enumerate(normalizados):
                if not titulo or idx in mapa.values():
                    continue
                candidato = titulo if chave == "exato" else _cabecalho_sem_ruido(titulo)
                if candidato in sinonimos:
                    mapa[campo] = idx
                    break
    faltando = [ROTULOS_IMPORTACAO[c] for c in CAMPOS_OBRIGATORIOS if c not in mapa]
    return mapa, faltando


def _num_br(valor):
    """Aceita 1.234,56 (pt-BR), 1234.56 (en) e números vindos do próprio Excel."""
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    txt = str(valor).strip()
    txt = re.sub(r"[R$\s ]", "", txt, flags=re.IGNORECASE)
    if "," in txt and "." in txt:            # 1.234,56 → 1234.56
        txt = txt.replace(".", "").replace(",", ".")
    elif "," in txt:                          # 12,50 → 12.50
        txt = txt.replace(",", ".")
    try:
        return float(txt)
    except ValueError:
        return None


def _ler_planilha(conteudo: bytes) -> list:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(503, "Leitura de Excel indisponível no servidor (openpyxl ausente)")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(400, "Arquivo inválido: envie uma planilha .xlsx")
    linhas = [list(l) for l in wb[wb.sheetnames[0]].iter_rows(values_only=True)]
    wb.close()
    return linhas


def _analisar_planilha(conteudo: bytes, existentes: dict) -> dict:
    """Valida a planilha inteira ANTES de qualquer escrita e devolve a prévia.

    `existentes` mapeia a chave de duplicidade (código, ou nome quando não há
    código) para o item já cadastrado — é o que decide criar x atualizar."""
    linhas = _ler_planilha(conteudo)
    if not linhas:
        raise HTTPException(400, "A planilha está vazia")

    # A primeira linha não-vazia é o cabeçalho (tolera título/logo acima dela).
    idx_cab, cabecalho = None, None
    for i, linha in enumerate(linhas[:10]):
        mapa_teste, _ = _mapear_colunas(linha)
        if "nome" in mapa_teste:
            idx_cab, cabecalho = i, linha
            break
    if cabecalho is None:
        idx_cab, cabecalho = 0, linhas[0]

    mapa, faltando = _mapear_colunas(cabecalho)
    if faltando:
        encontradas = [str(c).strip() for c in cabecalho if c is not None and str(c).strip()]
        raise HTTPException(
            400,
            "Coluna obrigatória ausente: " + ", ".join(faltando)
            + ". Cabeçalhos encontrados: " + (", ".join(encontradas) or "nenhum")
            + ". Baixe o modelo de importação para conferir os nomes esperados.",
        )

    def celula(linha, campo):
        idx = mapa.get(campo)
        if idx is None or idx >= len(linha):
            return None
        v = linha[idx]
        return v.strip() if isinstance(v, str) else v

    validos, erros, vistos = [], [], {}
    corpo = linhas[idx_cab + 1:]
    if len(corpo) > LIMITE_LINHAS_IMPORTACAO:
        raise HTTPException(400, f"A planilha tem mais de {LIMITE_LINHAS_IMPORTACAO} linhas. Divida o arquivo.")

    for offset, linha in enumerate(corpo):
        num = idx_cab + 2 + offset  # número da linha como o usuário vê no Excel
        if not any(str(c).strip() for c in linha if c is not None):
            continue  # linha em branco: ignora em silêncio

        problemas = []
        nome = celula(linha, "nome")
        nome = str(nome).strip() if nome is not None else ""
        if not nome:
            problemas.append(f"Linha {num} - {ROTULOS_IMPORTACAO['nome']}: obrigatório e está vazio.")

        bruto_preco = celula(linha, "preco_base")
        preco = _num_br(bruto_preco)
        if preco is None:
            problemas.append(
                f"Linha {num} - {ROTULOS_IMPORTACAO['preco_base']}: valor \"{bruto_preco}\" não é numérico."
            )
        elif preco < 0:
            problemas.append(f"Linha {num} - {ROTULOS_IMPORTACAO['preco_base']}: não pode ser negativo.")

        bruto_qtd = celula(linha, "quantidade")
        if bruto_qtd is None or (isinstance(bruto_qtd, str) and not bruto_qtd.strip()):
            quantidade = 0
        else:
            n = _num_br(bruto_qtd)
            if n is None:
                problemas.append(
                    f"Linha {num} - {ROTULOS_IMPORTACAO['quantidade']}: valor \"{bruto_qtd}\" não é numérico."
                )
                quantidade = 0
            elif n < 0:
                problemas.append(f"Linha {num} - {ROTULOS_IMPORTACAO['quantidade']}: não pode ser negativa.")
                quantidade = 0
            else:
                quantidade = int(round(n))

        codigo = celula(linha, "codigo")
        codigo = str(codigo).strip() if codigo not in (None, "") else None
        if isinstance(codigo, str) and codigo.endswith(".0") and codigo[:-2].isdigit():
            codigo = codigo[:-2]  # Excel devolve código numérico como 1234.0

        descricao = celula(linha, "descricao")
        descricao = str(descricao).strip() if descricao not in (None, "") else None

        chave = f"cod:{codigo.lower()}" if codigo else f"nome:{nome.lower()}"
        if chave in vistos:
            problemas.append(
                f"Linha {num} - {ROTULOS_IMPORTACAO['codigo'] if codigo else ROTULOS_IMPORTACAO['nome']}: "
                f"repetido na linha {vistos[chave]} do próprio arquivo."
            )

        if problemas:
            erros.extend(problemas)
            continue

        vistos[chave] = num
        ja_existe = existentes.get(chave)
        validos.append({
            "linha": num,
            "codigo": codigo,
            "nome": nome,
            "descricao": descricao,
            "quantidade": quantidade,
            "preco_base": round(preco, 2),
            "acao": "atualizar" if ja_existe else "criar",
            "equipamento_id": ja_existe,
        })

    return {
        "colunas_reconhecidas": {ROTULOS_IMPORTACAO[c]: int(i) for c, i in mapa.items()},
        "linha_cabecalho": idx_cab + 1,
        "total_linhas": len([l for l in corpo if any(str(c).strip() for c in l if c is not None)]),
        "validos": validos,
        "erros": erros,
        "resumo": {
            "validos": len(validos),
            "com_erro": len(erros),
            "criar": sum(1 for v in validos if v["acao"] == "criar"),
            "atualizar": sum(1 for v in validos if v["acao"] == "atualizar"),
        },
    }


def _catalogo_existente(conn, conta_id: str, tipo: str = "equipamento") -> dict:
    """Chaves de duplicidade do catálogo atual → equipamento_id.
    Prioriza o código (SKU); só cai no nome quando o item não tem código.

    Recortado por `tipo`: importar a planilha de serviços não pode atualizar um
    equipamento que por acaso tem o mesmo nome — "Instalação" pode existir nos
    dois catálogos e significar coisas diferentes."""
    rows = conn.execute(
        text("""SELECT equipamento_id, codigo, nome FROM equipamentos
                WHERE conta_id = :cid AND COALESCE(tipo, 'equipamento') = :t"""),
        {"cid": conta_id, "t": tipo},
    ).fetchall()
    mapa = {}
    for r in rows:
        if r.codigo:
            mapa[f"cod:{str(r.codigo).strip().lower()}"] = str(r.equipamento_id)
        else:
            mapa.setdefault(f"nome:{(r.nome or '').strip().lower()}", str(r.equipamento_id))
    return mapa


@app.get("/equipamentos/modelo-importacao")
def modelo_importacao_equipamentos(auth: dict = Depends(get_auth), tipo: Optional[str] = None):
    """Modelo .xlsx gerado a partir dos campos reais do catálogo, com uma linha
    de exemplo. É o arquivo que o usuário preenche e devolve na importação."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(503, "Geração de Excel indisponível no servidor (openpyxl ausente)")
    from openpyxl.styles import Font, PatternFill, Alignment

    tipo = _tipo_catalogo(tipo)
    servico = tipo == "servico"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Serviços" if servico else "Catálogo"
    # Servico nao tem estoque: a coluna sai do modelo em vez de vir e ser
    # ignorada — planilha com coluna morta e convite para preencher errado.
    campos = [c for c in CAMPOS_IMPORTACAO if not (servico and c[0] == "quantidade")]
    cabecalhos = [ROTULOS_IMPORTACAO[c] for c, _, _, _ in campos]
    ws.append(cabecalhos)
    for i, _ in enumerate(cabecalhos, start=1):
        cel = ws.cell(row=1, column=i)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="2980B9")
        cel.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cel.column_letter].width = 26
    ws.append(
        ["SRV-001", "Instalação em campo", "Instalação e comissionamento no cliente", 800.00]
        if servico else
        ["EQ-001", "Gerador 15 kVA", "Gerador a diesel silenciado", 3, 1250.00]
    )
    ws.freeze_panes = "A2"

    ajuda = wb.create_sheet("Instruções")
    for linha in [
        ["Como importar"],
        [""],
        [f"1. Preencha uma linha por item, a partir da linha 2 da aba '{ws.title}'."],
        ["2. As colunas são reconhecidas PELO NOME do cabeçalho, não pela posição."],
        ["   Você pode reordenar as colunas à vontade — não deixe de renomear o cabeçalho."],
        ["3. Colunas obrigatórias: " + ", ".join(ROTULOS_IMPORTACAO[c] for c in CAMPOS_OBRIGATORIOS) + "."],
        ["4. Código (SKU) é opcional, mas é ele que identifica o item numa reimportação:"],
        ["   mesmo código = atualiza o item existente; código novo = cria."],
        ["   Sem código, a identificação cai para o Nome do item."],
        ["5. Preço aceita 1.234,56 ou 1234.56. Quantidade deve ser um número inteiro."]
        if not servico else
        ["5. Preço aceita 1.234,56 ou 1234.56. Serviço não tem estoque, por isso não há coluna Quantidade."],
        ["6. Linhas em branco são ignoradas."],
        ["7. Esta planilha importa " + ("SERVIÇOS" if servico else "EQUIPAMENTOS")
         + " — use a aba correspondente do sistema para enviá-la."],
    ]:
        ajuda.append(linha)
    ajuda["A1"].font = Font(bold=True, size=13)
    ajuda.column_dimensions["A"].width = 95

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="modelo-{"servicos" if servico else "catalogo"}-prospectageo.xlsx"'},
    )


@app.post("/equipamentos/importar")
async def importar_equipamentos(
    arquivo: UploadFile = File(...),
    confirmar: bool = False,
    tipo: Optional[str] = None,
    auth: dict = Depends(get_auth),
):
    """Importa catálogo/estoque de um .xlsx.

    Dois passos, com o MESMO arquivo:
    - `confirmar=false` (padrão): só valida e devolve a prévia. Nada é gravado.
    - `confirmar=true`: grava tudo dentro de uma única transação. Se qualquer
      linha falhar, nada é salvo — não existe meia-importação silenciosa.

    Com erros de validação e `confirmar=true`, a gravação é recusada (400) e o
    usuário recebe a lista de problemas por linha/coluna."""
    nome_arq = (arquivo.filename or "").lower()
    if not nome_arq.endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Formato não suportado. Envie um arquivo .xlsx (Excel).")
    conteudo = await arquivo.read()
    if not conteudo:
        raise HTTPException(400, "Arquivo vazio")
    if len(conteudo) > 5 * 1024 * 1024:
        raise HTTPException(400, "Arquivo muito grande (máximo 5 MB)")

    tipo = _tipo_catalogo(tipo)
    with engine.begin() as conn:
        garantir_vendas(conn)
        analise = _analisar_planilha(conteudo, _catalogo_existente(conn, auth["conta_id"], tipo))
        analise["tipo"] = tipo

        if not confirmar:
            analise["gravado"] = False
            # Na prévia devolvemos no máximo 50 linhas — o resumo tem os números completos.
            analise["validos"] = analise["validos"][:50]
            analise["erros"] = analise["erros"][:50]
            return analise

        if analise["erros"]:
            raise HTTPException(
                400,
                "A planilha tem linhas inválidas. Corrija antes de importar: "
                + " | ".join(analise["erros"][:5])
                + (" ..." if len(analise["erros"]) > 5 else ""),
            )
        if not analise["validos"]:
            raise HTTPException(400, "Nenhuma linha válida para importar.")

        criados = atualizados = 0
        for item in analise["validos"]:
            if item["equipamento_id"]:
                conn.execute(
                    text(
                        """UPDATE equipamentos SET
                               nome = :n,
                               codigo = COALESCE(:c, codigo),
                               descricao = COALESCE(:d, descricao),
                               preco_base = :p,
                               quantidade = :q,
                               ativo = TRUE
                           WHERE equipamento_id = :eid AND conta_id = :cid"""
                    ),
                    {
                        "n": item["nome"], "c": item["codigo"], "d": item["descricao"],
                        "p": item["preco_base"],
                        "q": 0 if tipo == "servico" else item["quantidade"],
                        "eid": item["equipamento_id"], "cid": auth["conta_id"],
                    },
                )
                atualizados += 1
            else:
                conn.execute(
                    text(
                        """INSERT INTO equipamentos
                               (conta_id, codigo, nome, descricao, preco_base, quantidade, tipo)
                           VALUES (:cid, :c, :n, :d, :p, :q, :t)"""
                    ),
                    {
                        "cid": auth["conta_id"], "c": item["codigo"], "n": item["nome"],
                        "d": item["descricao"], "p": item["preco_base"],
                        "q": 0 if tipo == "servico" else item["quantidade"],
                        "t": tipo,
                    },
                )
                criados += 1

        registrar_auditoria(
            usuario=auth,
            acao="SERVICOS_IMPORTADOS" if tipo == "servico" else "CATALOGO_IMPORTADO",
            recurso="equipamentos", quantidade=criados + atualizados, conn=conn,
        )

    return {
        "gravado": True,
        "tipo": tipo,
        "criados": criados,
        "atualizados": atualizados,
        "total": criados + atualizados,
        "resumo": analise["resumo"],
    }


# =========================
# VENDAS: ORÇAMENTOS
# =========================
def _carregar_itens(conn, orcamento_id: str):
    rows = conn.execute(
        text(
            """SELECT item_id, equipamento_id, descricao, quantidade, preco_unitario, tipo
               FROM orcamento_itens WHERE orcamento_id = :oid ORDER BY descricao ASC"""
        ),
        {"oid": orcamento_id},
    )
    return [dict(r._mapping) for r in rows]


def _regravar_itens(conn, orcamento_id: str, itens: list) -> float:
    """Troca os itens do orçamento e devolve o novo total."""
    conn.execute(text("DELETE FROM orcamento_itens WHERE orcamento_id = :oid"), {"oid": orcamento_id})
    total = 0.0
    for item in itens:
        qtd = max(1, int(item.quantidade or 1))
        preco = float(item.preco_unitario or 0)
        total += qtd * preco
        # Tipo so do avulso (ver comentario em OrcamentoItemIn.tipo). Valor
        # desconhecido vira NULL em vez de 400: o tipo e uma dica de
        # classificacao vinda da UI, e derrubar o salvamento do orcamento
        # inteiro por causa dela seria desproporcional.
        tipo_item = None
        if not item.equipamento_id:
            bruto = (getattr(item, "tipo", None) or "").strip().lower()
            tipo_item = bruto if bruto in TIPOS_CATALOGO else None
        conn.execute(
            text(
                """INSERT INTO orcamento_itens
                   (orcamento_id, equipamento_id, descricao, quantidade, preco_unitario, tipo)
                   VALUES (:oid, :eq, :d, :q, :p, :tp)"""
            ),
            {
                "oid": orcamento_id,
                "eq": item.equipamento_id or None,
                "d": (item.descricao or "").strip() or "Item",
                "q": qtd,
                "p": preco,
                "tp": tipo_item,
            },
        )
    conn.execute(
        text("UPDATE orcamentos SET total = :t, atualizado_em = NOW() WHERE orcamento_id = :oid"),
        {"t": total, "oid": orcamento_id},
    )
    return total


def _orcamento_do_usuario(conn, orcamento_id: str, auth: dict):
    """Carrega o orçamento respeitando a carteira. 404 se não for visível."""
    escopo, params = _escopo_vendas(conn, auth)
    row = conn.execute(
        text(f"SELECT * FROM orcamentos o WHERE o.orcamento_id = :oid AND {escopo}"),
        {**params, "oid": orcamento_id},
    ).fetchone()
    if not row:
        raise HTTPException(404, "Orçamento não encontrado")
    return dict(row._mapping)


@app.get("/orcamentos")
def listar_orcamentos(auth: dict = Depends(get_auth), status: Optional[str] = None,
                      empresa_id: Optional[str] = None):
    """Lista orçamentos da carteira, com empresa, vendedor e resumo dos itens
    já resolvidos.

    O vendedor e os itens vêm no mesmo SELECT de propósito: a ficha da empresa
    mostra uma linha por orçamento com quem vendeu e do que se trata, e buscar
    isso por linha seria N+1 num endpoint que o front relê de 5 em 5 segundos.
    `item_principal` é o item de maior valor do orçamento — é ele que diz do que
    a proposta trata, não o primeiro cadastrado."""
    with engine.begin() as conn:
        garantir_vendas(conn)
        escopo, params = _escopo_vendas(conn, auth)
        if status:
            escopo += " AND o.status = :st"
            params["st"] = status
        if empresa_id:
            escopo += " AND o.empresa_id = :eid"
            params["eid"] = empresa_id
        rows = conn.execute(
            text(
                f"""SELECT o.*, e.nome AS empresa_nome, u.nome AS vendedor_nome,
                           COALESCE(it.qtd_itens, 0) AS qtd_itens,
                           COALESCE(it.qtd_pecas, 0) AS qtd_pecas,
                           it.item_principal
                    FROM orcamentos o
                    LEFT JOIN empresas e ON e.empresa_id = o.empresa_id
                    LEFT JOIN usuarios u ON u.usuario_id = o.vendedor_id
                    LEFT JOIN LATERAL (
                        SELECT COUNT(*)::int AS qtd_itens,
                               COALESCE(SUM(i.quantidade), 0)::int AS qtd_pecas,
                               (array_agg(i.descricao
                                    ORDER BY i.quantidade * i.preco_unitario DESC))[1]
                                    AS item_principal
                        FROM orcamento_itens i
                        WHERE i.orcamento_id = o.orcamento_id
                    ) it ON TRUE
                    WHERE {escopo}
                    ORDER BY o.atualizado_em DESC NULLS LAST, o.criado_em DESC"""
            ),
            params,
        )
        return [dict(r._mapping) for r in rows]


@app.get("/orcamentos/{orcamento_id}")
def obter_orcamento(orcamento_id: str, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        garantir_vendas(conn)
        orc = _orcamento_do_usuario(conn, orcamento_id, auth)
        orc["itens"] = _carregar_itens(conn, orcamento_id)
        return orc


@app.post("/orcamentos")
def criar_orcamento(dados: OrcamentoCreate, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        garantir_vendas(conn)
        emp = conn.execute(
            text("SELECT empresa_id FROM empresas WHERE empresa_id = :eid AND conta_id = :cid"),
            {"eid": dados.empresa_id, "cid": auth["conta_id"]},
        ).fetchone()
        if not emp:
            raise HTTPException(404, "Empresa não encontrada")
        row = conn.execute(
            text(
                """INSERT INTO orcamentos (conta_id, empresa_id, vendedor_id, titulo, observacoes)
                   VALUES (:cid, :eid, :vid, :t, :o) RETURNING orcamento_id"""
            ),
            {
                "cid": auth["conta_id"],
                "eid": dados.empresa_id,
                "vid": auth["usuario_id"],
                "t": (dados.titulo or "").strip() or "Orçamento",
                "o": dados.observacoes,
            },
        ).fetchone()
        oid = str(row.orcamento_id)
        _regravar_itens(conn, oid, dados.itens or [])
        orc = _orcamento_do_usuario(conn, oid, auth)
        orc["itens"] = _carregar_itens(conn, oid)
        return orc


@app.put("/orcamentos/{orcamento_id}")
def atualizar_orcamento(orcamento_id: str, dados: OrcamentoUpdate, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        garantir_vendas(conn)
        atual = _orcamento_do_usuario(conn, orcamento_id, auth)
        if atual["status"] in ("aprovado", "recusado"):
            raise HTTPException(400, "Orçamento já decidido não pode ser editado")
        campos = {}
        if dados.titulo is not None:
            campos["titulo"] = dados.titulo
        if dados.observacoes is not None:
            campos["observacoes"] = dados.observacoes
        if campos:
            sets = ", ".join(f"{k} = :{k}" for k in campos)
            conn.execute(
                text(f"UPDATE orcamentos SET {sets}, atualizado_em = NOW() WHERE orcamento_id = :oid"),
                {**campos, "oid": orcamento_id},
            )
        if dados.itens is not None:
            _regravar_itens(conn, orcamento_id, dados.itens)
        orc = _orcamento_do_usuario(conn, orcamento_id, auth)
        orc["itens"] = _carregar_itens(conn, orcamento_id)
        return orc


@app.delete("/orcamentos/{orcamento_id}")
def excluir_orcamento(orcamento_id: str, auth: dict = Depends(get_auth)):
    with engine.begin() as conn:
        garantir_vendas(conn)
        _orcamento_do_usuario(conn, orcamento_id, auth)
        conn.execute(text("DELETE FROM orcamentos WHERE orcamento_id = :oid"), {"oid": orcamento_id})
        return {"msg": "Orçamento excluído"}


def _brl(v) -> str:
    """Formata em Real: 1234.5 -> 'R$ 1.234,50'."""
    return f"R$ {float(v or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _html_orcamento(empresa_nome: str, orc: dict, itens: list) -> str:
    """Corpo do email do orçamento."""
    linhas = "".join(
        "<tr>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>{i['descricao']}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eee;text-align:center'>{i['quantidade']}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eee;text-align:right'>{_brl(i['preco_unitario'])}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eee;text-align:right'>"
        f"{_brl(int(i['quantidade'] or 1) * float(i['preco_unitario'] or 0))}</td>"
        "</tr>"
        for i in itens
    )
    obs = f"<p style='color:#555'>{orc.get('observacoes')}</p>" if orc.get("observacoes") else ""
    titulo = orc.get("titulo") or "Orçamento"
    return (
        "<div style=\"font-family:Arial,sans-serif;max-width:640px\">"
        f"<h2 style=\"color:#0f2133\">{titulo}</h2>"
        f"<p style=\"color:#555\">Olá, {empresa_nome}!</p>"
        "<p style=\"color:#555\">Segue o orçamento solicitado:</p>"
        "<table style=\"width:100%;border-collapse:collapse;font-size:14px\">"
        "<thead><tr style=\"background:#f4f7fa\">"
        "<th style=\"padding:8px 10px;text-align:left\">Item</th>"
        "<th style=\"padding:8px 10px;text-align:center\">Qtd</th>"
        "<th style=\"padding:8px 10px;text-align:right\">Unit.</th>"
        "<th style=\"padding:8px 10px;text-align:right\">Total</th>"
        "</tr></thead>"
        f"<tbody>{linhas}</tbody></table>"
        "<p style=\"font-size:16px;font-weight:bold;text-align:right;color:#0f2133\">"
        f"Total: {_brl(orc.get('total'))}</p>"
        f"{obs}</div>"
    )


def _texto_orcamento(empresa_nome: str, orc: dict, itens: list) -> str:
    """Versão em texto puro do orçamento — para o corpo de um mailto ou do
    WhatsApp, onde HTML não funciona."""
    linhas = "\n".join(
        f"- {i['descricao']} | {i['quantidade']} x {_brl(i['preco_unitario'])} = "
        f"{_brl(int(i['quantidade'] or 1) * float(i['preco_unitario'] or 0))}"
        for i in itens
    )
    partes = [
        f"Olá, {empresa_nome}!",
        "",
        "Segue o orçamento solicitado:",
        "",
        linhas,
        "",
        f"Total: {_brl(orc.get('total'))}",
    ]
    if orc.get("observacoes"):
        partes += ["", str(orc["observacoes"])]
    return "\n".join(partes)


@app.get("/orcamentos/{orcamento_id}/previa-email")
def previa_email_orcamento(orcamento_id: str, auth: dict = Depends(get_auth)):
    """Conteúdo do orçamento pronto para o vendedor enviar por conta própria.

    Existe para o envio automático não ser um beco sem saída: quando o Resend
    recusa, o vendedor abre o próprio email/WhatsApp com tudo preenchido. Não
    altera o status — quem enviou por fora marca em PUT /orcamentos/{id}/status."""
    with engine.begin() as conn:
        garantir_vendas(conn)
        orc = _orcamento_do_usuario(conn, orcamento_id, auth)
        itens = _carregar_itens(conn, orcamento_id)
        emp = conn.execute(
            text(
                """SELECT e.nome, c.email, c.celular, c.whatsapp
                   FROM empresas e
                   LEFT JOIN LATERAL (
                       SELECT email, celular, whatsapp FROM contatos
                       WHERE empresa_id = e.empresa_id
                       ORDER BY decisor DESC NULLS LAST, data_criacao ASC NULLS LAST LIMIT 1
                   ) c ON TRUE
                   WHERE e.empresa_id = :eid"""
            ),
            {"eid": orc["empresa_id"]},
        ).fetchone()
    if not emp:
        raise HTTPException(404, "Empresa do orçamento não encontrada")
    return {
        "destino": emp.email,
        "telefone": emp.whatsapp or emp.celular,
        "empresa_nome": emp.nome,
        "assunto": orc.get("titulo") or "Orçamento",
        "texto": _texto_orcamento(emp.nome, orc, itens),
        "html": _html_orcamento(emp.nome, orc, itens),
    }


@app.post("/orcamentos/{orcamento_id}/enviar")
def enviar_orcamento(orcamento_id: str, request: Request, auth: dict = Depends(get_auth)):
    """Envia o orçamento por email ao contato da empresa e marca como enviado."""
    with engine.begin() as conn:
        garantir_vendas(conn)
        orc = _orcamento_do_usuario(conn, orcamento_id, auth)
        itens = _carregar_itens(conn, orcamento_id)
        if not itens:
            raise HTTPException(400, "Orçamento sem itens não pode ser enviado")
        emp = conn.execute(
            text(
                """SELECT e.nome, c.email
                   FROM empresas e
                   LEFT JOIN LATERAL (
                       SELECT email FROM contatos
                       WHERE empresa_id = e.empresa_id AND email IS NOT NULL
                       ORDER BY decisor DESC NULLS LAST, data_criacao ASC NULLS LAST LIMIT 1
                   ) c ON TRUE
                   WHERE e.empresa_id = :eid"""
            ),
            {"eid": orc["empresa_id"]},
        ).fetchone()
        if not emp or not emp.email:
            raise HTTPException(400, "A empresa não tem contato com email cadastrado")
        try:
            resend.Emails.send(
                {
                    "from": REMETENTE_EMAIL,
                    "to": emp.email,
                    "reply_to": endereco_de_resposta(conn, auth["email"]),
                    "subject": orc.get("titulo") or "Orçamento",
                    "html": _html_orcamento(emp.nome, orc, itens),
                }
            )
        except Exception as e:
            # Diferente do convite: aqui o 502 é proposital, para a transação
            # inteira dar rollback. Marcar o orçamento como "enviado" quando o
            # email não saiu seria mentira — o cliente não recebeu nada. O que
            # muda é a mensagem: agora diz POR QUE falhou, e o vendedor tem
            # GET /orcamentos/{id}/previa-email para mandar por conta própria.
            print(f"❌ Falha ao enviar orçamento {orcamento_id} (from={REMETENTE_EMAIL}): {e}")
            raise HTTPException(502, motivo_falha_email(e))
        conn.execute(
            text(
                """UPDATE orcamentos
                   SET status = CASE WHEN status = 'rascunho' THEN 'enviado' ELSE status END,
                       data_envio = NOW(), atualizado_em = NOW()
                   WHERE orcamento_id = :oid"""
            ),
            {"oid": orcamento_id},
        )
        registrar_auditoria(usuario=auth, acao="ORCAMENTO_ENVIADO", recurso="orcamentos",
                            recurso_id=orcamento_id, request=request, conn=conn)
        atualizado = _orcamento_do_usuario(conn, orcamento_id, auth)
        atualizado["itens"] = itens
        return atualizado


@app.put("/orcamentos/{orcamento_id}/status")
def atualizar_status_orcamento(orcamento_id: str, dados: OrcamentoStatusUpdate,
                               auth: dict = Depends(get_auth)):
    """Move o orçamento no fluxo: enviado -> em_negociacao -> aprovado/recusado."""
    if dados.status not in ORCAMENTO_STATUS:
        raise HTTPException(400, f"Status inválido. Use um de: {', '.join(ORCAMENTO_STATUS)}")
    with engine.begin() as conn:
        garantir_vendas(conn)
        _orcamento_do_usuario(conn, orcamento_id, auth)
        decidido = dados.status in ("aprovado", "recusado")
        conn.execute(
            text(
                """UPDATE orcamentos
                   SET status = :st,
                       motivo_recusa = :mr,
                       data_decisao = CASE WHEN :dec THEN NOW() ELSE data_decisao END,
                       -- Quem enviou o orçamento por fora (email próprio, WhatsApp) marca
                       -- 'enviado' na mão; sem isto a data de envio ficaria vazia.
                       data_envio = CASE WHEN :st = 'enviado' AND data_envio IS NULL
                                         THEN NOW() ELSE data_envio END,
                       atualizado_em = NOW()
                   WHERE orcamento_id = :oid"""
            ),
            {
                "st": dados.status,
                "mr": dados.motivo_recusa if dados.status == "recusado" else None,
                "dec": decidido,
                "oid": orcamento_id,
            },
        )
        orc = _orcamento_do_usuario(conn, orcamento_id, auth)
        orc["itens"] = _carregar_itens(conn, orcamento_id)
        return orc


@app.get("/vendas/insights")
def insights_vendas(auth: dict = Depends(get_auth)):
    """Números do dashboard de vendas: funil por status, valores, tempo de resposta
    e desempenho por equipamento (ofertado x aprovado x recusado)."""
    with engine.begin() as conn:
        garantir_vendas(conn)
        escopo, params = _escopo_vendas(conn, auth)
        por_status = conn.execute(
            text(
                f"""SELECT status, COUNT(*) AS total, COALESCE(SUM(total),0) AS valor
                    FROM orcamentos o WHERE {escopo} GROUP BY status"""
            ),
            params,
        )
        status_map = {r.status: {"total": r.total, "valor": float(r.valor or 0)} for r in por_status}
        resumo = {s: status_map.get(s, {"total": 0, "valor": 0.0}) for s in ORCAMENTO_STATUS}

        ranking = conn.execute(
            text(
                f"""SELECT COALESCE(eq.nome, i.descricao) AS nome,
                           SUM(i.quantidade) AS quantidade,
                           COALESCE(SUM(i.quantidade * i.preco_unitario),0) AS valor
                    FROM orcamento_itens i
                    JOIN orcamentos o ON o.orcamento_id = i.orcamento_id
                    LEFT JOIN equipamentos eq ON eq.equipamento_id = i.equipamento_id
                    WHERE {escopo}
                    GROUP BY COALESCE(eq.nome, i.descricao)
                    ORDER BY quantidade DESC
                    LIMIT 10"""
            ),
            params,
        ).fetchall()
        # Metricas POR EQUIPAMENTO com desfecho, nao so popularidade.
        # `equipamentos_mais_orcados` acima responde "o que a gente mais oferta";
        # esta responde "o que a gente mais GANHA", que e outra pergunta e a que
        # decide catalogo e desconto. Um item muito orcado e pouco aprovado e o
        # sinal mais barato de preco fora do mercado que este CRM consegue dar.
        #
        # O LIMIT e folgado de proposito: o front monta "mais vendidos" E "menos
        # vendidos" a partir desta lista, e um limite apertado cortaria
        # justamente a cauda que a segunda metade precisa. So itens que ja
        # apareceram em algum orcamento entram (INNER JOIN) -- "nunca ofertado"
        # e outra pergunta, e sai do catalogo, nao daqui.
        por_equipamento = conn.execute(
            text(
                f"""SELECT COALESCE(eq.nome, i.descricao) AS nome,
                           COALESCE(eq.tipo, i.tipo, 'equipamento') AS tipo,
                           SUM(i.quantidade) AS quantidade,
                           COALESCE(SUM(i.quantidade * i.preco_unitario),0) AS valor,
                           COALESCE(SUM(i.quantidade)
                               FILTER (WHERE o.status = 'aprovado'),0) AS qtd_aprovada,
                           COALESCE(SUM(i.quantidade * i.preco_unitario)
                               FILTER (WHERE o.status = 'aprovado'),0) AS valor_aprovado,
                           COALESCE(SUM(i.quantidade)
                               FILTER (WHERE o.status = 'recusado'),0) AS qtd_recusada,
                           COALESCE(SUM(i.quantidade)
                               FILTER (WHERE o.status IN ('enviado','em_negociacao')),0) AS qtd_aberta,
                           COALESCE(SUM(i.quantidade * i.preco_unitario)
                               FILTER (WHERE o.status IN ('enviado','em_negociacao')),0) AS valor_aberto
                    FROM orcamento_itens i
                    JOIN orcamentos o ON o.orcamento_id = i.orcamento_id
                    LEFT JOIN equipamentos eq ON eq.equipamento_id = i.equipamento_id
                    WHERE {escopo}
                    GROUP BY COALESCE(eq.nome, i.descricao), COALESCE(eq.tipo, i.tipo, 'equipamento')
                    ORDER BY valor_aprovado DESC, quantidade DESC
                    LIMIT 200"""
            ),
            params,
        )
        equipamentos = []
        for r in por_equipamento:
            ganhou, perdeu = int(r.qtd_aprovada or 0), int(r.qtd_recusada or 0)
            decidido = ganhou + perdeu
            equipamentos.append({
                "nome": r.nome,
                # Item avulso (sem equipamento_id) conta como equipamento: e
                # onde ele sempre apareceu, e mudar isso escondia historico.
                "tipo": r.tipo,
                "quantidade": int(r.quantidade or 0),
                "valor": float(r.valor or 0),
                "qtd_aprovada": ganhou,
                "valor_aprovado": float(r.valor_aprovado or 0),
                "qtd_recusada": perdeu,
                "qtd_aberta": int(r.qtd_aberta or 0),
                "valor_aberto": float(r.valor_aberto or 0),
                # None (nao 0) quando nada foi decidido: "0% de aprovacao" e
                # "ninguem respondeu ainda" sao leituras opostas, e o front
                # precisa distinguir para nao acusar item que so foi ofertado.
                "taxa_aprovacao": round(ganhou / decidido * 100, 1) if decidido else None,
            })

        # Quanto tempo o cliente leva para responder uma proposta enviada. E o
        # numero que diz quando cobrar: sem ele o vendedor chuta o follow-up.
        resposta = conn.execute(
            text(
                f"""SELECT AVG(EXTRACT(EPOCH FROM (o.data_decisao - o.data_envio)) / 86400.0)
                    FROM orcamentos o
                    WHERE {escopo} AND o.data_envio IS NOT NULL
                      AND o.data_decisao IS NOT NULL AND o.data_decisao >= o.data_envio"""
            ),
            params,
        ).scalar()

        aprovados = resumo["aprovado"]["total"]
        decididos = aprovados + resumo["recusado"]["total"]
        return {
            "por_status": resumo,
            "total_orcamentos": sum(v["total"] for v in resumo.values()),
            "valor_em_aberto": resumo["enviado"]["valor"] + resumo["em_negociacao"]["valor"],
            "valor_aprovado": resumo["aprovado"]["valor"],
            "taxa_conversao": round(aprovados / decididos * 100, 1) if decididos else 0.0,
            "ticket_medio": round(resumo["aprovado"]["valor"] / aprovados, 2) if aprovados else None,
            "tempo_medio_resposta_dias": round(float(resposta), 1) if resposta is not None else None,
            "equipamentos": equipamentos,
            "equipamentos_mais_orcados": [
                {"nome": r.nome, "quantidade": int(r.quantidade or 0), "valor": float(r.valor or 0)}
                for r in ranking
            ],
        }

# deploy marker: forca o webhook do Railway a reconhecer o modulo de vendas (bf01191)
